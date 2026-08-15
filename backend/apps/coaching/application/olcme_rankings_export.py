"""Ölçme değerlendirme — sıralama listesi dışa aktarma (Excel/CSV).

Tablo düzeni PDF sıralama ile aynıdır: ders başına D/Y/Net, T.Net, Puan
(veya AYT puan türleri), K%, tahmini sıra (yıl ile) ve TR%.
"""
from __future__ import annotations

import csv
import io
from django.http import HttpResponse

ALAN_LABELS = {
    'SAYISAL': 'Sayısal',
    'SOZEL': 'Sözel',
    'ESIT_AGIRLIK': 'Eşit Ağırlık',
}

PT_KEYS = ('SAY', 'EA', 'SOZ')
PT_LABELS = {'SAY': 'SAY', 'EA': 'EA', 'SOZ': 'SÖZ'}

# Eski düz kolon adları — geriye dönük test/import için korunur
EXPORT_COLUMNS = {
    'sira': 'Sıra',
    'student_name': 'Öğrenci Adı',
    'sinif': 'Sınıf/Şube',
    'alan_display': 'Alan',
    'toplam_net': 'T.Net',
    'total_correct': 'Doğru',
    'total_wrong': 'Yanlış',
    'total_empty': 'Boş',
    'puan': 'Puan',
    'kurum_ici_yuzdelik': 'K%',
    'tahmini_siralama': 'Tahmini TR Sıralaması',
    'yuzdelik_dilim': 'TR%',
}


def _subs_by_parent(sections: list[dict]) -> dict:
    mapping: dict = {}
    for sec in sections or []:
        if sec.get('is_sub_section') and sec.get('parent_id'):
            mapping.setdefault(sec['parent_id'], []).append(sec)
    return mapping


def ordered_export_sections(sections: list[dict] | None) -> list[dict]:
    """PDF ile aynı: ana bölüm + alt dersler; residual yoksa ana sütun atlanır."""
    sections = sections or []
    mains = [s for s in sections if not s.get('is_sub_section')]
    subs = _subs_by_parent(sections)
    ordered = []
    for main in mains:
        kids = subs.get(main['id'], [])
        residual = (main.get('question_count') or 0) - sum(k.get('question_count') or 0 for k in kids)
        if not kids or residual > 0:
            ordered.append(main)
        ordered.extend(kids)
    return ordered


def _section_question_count(sec: dict, subs_map: dict) -> int:
    if sec.get('is_sub_section'):
        return sec.get('question_count') or 0
    kids = subs_map.get(sec['id'], [])
    if not kids:
        return sec.get('question_count') or 0
    return (sec.get('question_count') or 0) - sum(k.get('question_count') or 0 for k in kids)


def _section_triplet(row: dict, sec: dict, subs_map: dict) -> tuple:
    nets = row.get('section_nets') or {}
    data = nets.get(str(sec['id'])) or {}
    kids = [] if sec.get('is_sub_section') else subs_map.get(sec['id'], [])
    if not kids:
        return data.get('correct') or 0, data.get('wrong') or 0, data.get('net') or 0
    sub_c = sub_w = sub_n = 0
    for kid in kids:
        kd = nets.get(str(kid['id'])) or {}
        sub_c += kd.get('correct') or 0
        sub_w += kd.get('wrong') or 0
        sub_n += kd.get('net') or 0
    return (
        (data.get('correct') or 0) - sub_c,
        (data.get('wrong') or 0) - sub_w,
        round((data.get('net') or 0) - sub_n, 2),
    )


def _avg_triplet(section_avgs: dict | None, sec: dict, subs_map: dict) -> tuple | None:
    if not section_avgs:
        return None
    data = section_avgs.get(str(sec['id']))
    if not data:
        return None
    kids = [] if sec.get('is_sub_section') else subs_map.get(sec['id'], [])
    if not kids:
        return data.get('avg_correct') or 0, data.get('avg_wrong') or 0, data.get('avg_net') or 0
    sub_c = sub_w = sub_n = 0.0
    for kid in kids:
        kd = section_avgs.get(str(kid['id'])) or {}
        sub_c += kd.get('avg_correct') or 0
        sub_w += kd.get('avg_wrong') or 0
        sub_n += kd.get('avg_net') or 0
    return (
        round((data.get('avg_correct') or 0) - sub_c, 1),
        round((data.get('avg_wrong') or 0) - sub_w, 1),
        round((data.get('avg_net') or 0) - sub_n, 2),
    )


def _pt_ranks(ranking_list: list[dict]) -> dict:
    ranks: dict[str, dict] = {}
    for pt in PT_KEYS:
        scored = [
            r for r in ranking_list
            if ((r.get('puan_turleri') or {}).get(pt) or {}).get('puan') is not None
        ]
        scored.sort(
            key=lambda r: ((r.get('puan_turleri') or {}).get(pt) or {}).get('puan') or 0,
            reverse=True,
        )
        ranks[pt] = {r.get('answer_id'): i + 1 for i, r in enumerate(scored)}
    return ranks


def _parent_name(sec: dict, sections: list[dict]) -> str:
    if not sec.get('is_sub_section'):
        return sec.get('name') or 'Ders'
    for item in sections:
        if item.get('id') == sec.get('parent_id'):
            return item.get('name') or 'Ders'
    return sec.get('name') or 'Ders'


def _short_label(name: str, qc: int, limit: int = 12) -> str:
    label = name or 'Ders'
    if len(label) > limit:
        label = label[: limit - 1] + '.'
    return f'{label} ({qc})' if qc else label


def _d_key(sid) -> str:
    return f'd_{sid}'


def _y_key(sid) -> str:
    return f'y_{sid}'


def _net_key(sid) -> str:
    return f'net_{sid}'


def tahmini_sira_label(ranking_year: int | None) -> str:
    if ranking_year:
        return f'Tah.Sıra ({ranking_year})'
    return 'Tah.Sıra'


def build_export_columns(*, is_ayt: bool, sections: list[dict] | None = None, ranking_year: int | None = None):
    from shared.export.style_manager import ExportColumn

    ordered = ordered_export_sections(sections)
    subs_map = _subs_by_parent(sections or [])
    cols = [
        ExportColumn(key='sira', label='Sıra', type='integer', width=6),
        ExportColumn(key='student_name', label='İsim', type='text', width=22),
        ExportColumn(key='sinif', label='Sınıf', type='text', width=12),
    ]
    for sec in ordered:
        sid = sec.get('id')
        qc = _section_question_count(sec, subs_map)
        base = _short_label(sec.get('name') or 'Ders', qc)
        cols.append(ExportColumn(key=_d_key(sid), label=f'{base} D', type='integer', width=5))
        cols.append(ExportColumn(key=_y_key(sid), label=f'{base} Y', type='integer', width=5))
        cols.append(ExportColumn(key=_net_key(sid), label=f'{base} Net', type='decimal', width=7))
    cols.append(ExportColumn(key='toplam_net', label='T.Net', type='decimal', width=8))
    if is_ayt:
        for pt in PT_KEYS:
            cols.append(ExportColumn(key=f'puan_{pt.lower()}', label=f'{PT_LABELS[pt]} Puan', type='decimal', width=9))
            cols.append(ExportColumn(key=f'kurs_{pt.lower()}', label=f'{PT_LABELS[pt]} Kurs', type='integer', width=7))
            cols.append(ExportColumn(key=f'genel_{pt.lower()}', label=f'{PT_LABELS[pt]} Genel', type='integer', width=9))
    else:
        cols.append(ExportColumn(key='puan', label='Puan', type='decimal', width=9))
    cols.extend([
        ExportColumn(key='kurum_ici_yuzdelik', label='K%', type='percent', width=7),
        ExportColumn(key='tahmini_siralama', label=tahmini_sira_label(ranking_year), type='integer', width=14),
        ExportColumn(key='yuzdelik_dilim', label='TR%', type='percent', width=7),
    ])
    return cols


def build_export_rows(
    ranking_list: list[dict],
    *,
    is_ayt: bool,
    sections: list[dict] | None = None,
    section_avgs: dict | None = None,
    avg_net: float | None = None,
    avg_score: float | None = None,
    puan_turleri_avgs: dict | None = None,
    include_avg_row: bool = True,
) -> list[dict]:
    ordered = ordered_export_sections(sections)
    subs_map = _subs_by_parent(sections or [])
    ranks = _pt_ranks(ranking_list) if is_ayt else {}
    rows: list[dict] = []

    if include_avg_row and (section_avgs or avg_net is not None):
        avg_row = {
            'sira': None,
            'student_name': 'Kurs Ortalaması',
            'sinif': '',
            'alan_display': '',
            'toplam_net': avg_net,
            'puan': avg_score,
            'kurum_ici_yuzdelik': None,
            'tahmini_siralama': None,
            'yuzdelik_dilim': None,
            '_is_avg': True,
        }
        for sec in ordered:
            sid = sec.get('id')
            trip = _avg_triplet(section_avgs, sec, subs_map)
            if trip:
                avg_row[_d_key(sid)], avg_row[_y_key(sid)], avg_row[_net_key(sid)] = trip
            if is_ayt:
                for pt in PT_KEYS:
                    avg_row[f'puan_{pt.lower()}'] = (puan_turleri_avgs or {}).get(pt)
        rows.append(avg_row)

    for r in ranking_list:
        row = {
            'sira': r.get('kurum_ici_sira'),
            'student_name': r.get('student_name') or '',
            'sinif': r.get('sinif') or '',
            'alan_display': ALAN_LABELS.get(r.get('alan') or '', r.get('alan') or ''),
            'toplam_net': r.get('toplam_net'),
            'total_correct': r.get('total_correct') or 0,
            'total_wrong': r.get('total_wrong') or 0,
            'total_empty': r.get('total_empty') or 0,
            'puan': r.get('puan'),
            'kurum_ici_yuzdelik': r.get('kurum_ici_yuzdelik'),
            'tahmini_siralama': r.get('tahmini_siralama'),
            'yuzdelik_dilim': r.get('yuzdelik_dilim'),
        }
        for sec in ordered:
            sid = sec.get('id')
            correct, wrong, net = _section_triplet(r, sec, subs_map)
            if correct or wrong:
                row[_d_key(sid)] = correct
                row[_y_key(sid)] = wrong
                row[_net_key(sid)] = net
        if is_ayt:
            pt = r.get('puan_turleri') or {}
            for key in PT_KEYS:
                info = pt.get(key) or {}
                row[f'puan_{key.lower()}'] = info.get('puan')
                row[f'kurs_{key.lower()}'] = ranks.get(key, {}).get(r.get('answer_id'))
                row[f'genel_{key.lower()}'] = info.get('tahmini_siralama') or r.get('tahmini_siralama')
        rows.append(row)
    return rows


def build_export_meta(request, exam, *, report_title: str | None = None, ranking_year: int | None = None):
    from shared.export.style_manager import ReportMeta

    kurum_ad = getattr(exam.kurum, 'ad', '') if exam.kurum_id else ''
    sube_ad = getattr(exam.sube, 'ad', '') if exam.sube_id else ''
    egitim_yili = str(exam.egitim_yili) if exam.egitim_yili_id else ''

    user = getattr(request, 'user', None)
    generated_by = ''
    if user and getattr(user, 'is_authenticated', False):
        generated_by = user.get_full_name() or user.get_username()

    extra = {}
    if ranking_year:
        extra['tahmini_siralama_yili'] = ranking_year

    title = report_title or f'{exam.name} — Sıralama Sonuçları'
    return ReportMeta(
        report_title=title,
        kurum_ad=kurum_ad,
        sube_ad=sube_ad,
        egitim_yili=egitim_yili,
        generated_by=generated_by,
        extra=extra,
    )


def build_export_stats(ranking_list: list[dict], *, avg_net: float, avg_score: float, ranking_year: int | None = None):
    from shared.export.style_manager import ExportStat

    toplam = len(ranking_list)
    en_yuksek_puan = max((r.get('puan') or 0 for r in ranking_list), default=0)
    stats = [
        ExportStat(label='Katılımcı Sayısı', value=toplam, type='integer'),
        ExportStat(label='Ortalama Net', value=avg_net, type='decimal'),
        ExportStat(label='Ortalama Puan', value=avg_score, type='decimal'),
        ExportStat(label='En Yüksek Puan', value=en_yuksek_puan, type='decimal'),
    ]
    if ranking_year:
        stats.append(ExportStat(label='Tahmini Sıralama Yılı', value=ranking_year, type='year'))
    return stats


def export_rankings_file(
    export_format: str,
    ranking_list: list[dict],
    *,
    exam,
    request,
    is_ayt: bool,
    sections: list[dict] | None,
    section_avgs: dict | None,
    sinif_avgs: dict | None,
    puan_turleri_avgs: dict | None,
    ranking_year: int | None,
    avg_net: float,
    avg_score: float,
):
    rows = build_export_rows(
        ranking_list,
        is_ayt=is_ayt,
        sections=sections,
        section_avgs=section_avgs,
        avg_net=avg_net,
        avg_score=avg_score,
        puan_turleri_avgs=puan_turleri_avgs,
    )
    columns = build_export_columns(is_ayt=is_ayt, sections=sections, ranking_year=ranking_year)
    meta = build_export_meta(request, exam, report_title=f'{exam.name} — Sıralama Sonuçları', ranking_year=ranking_year)
    stats = build_export_stats(ranking_list, avg_net=avg_net, avg_score=avg_score, ranking_year=ranking_year)
    filename = f'{exam.name}_siralama'
    chart_ctx = {
        'sections': sections or [],
        'ordered_sections': ordered_export_sections(sections),
        'section_avgs': section_avgs or {},
        'sinif_avgs': sinif_avgs or {},
        'puan_turleri_avgs': puan_turleri_avgs or {},
        'is_ayt': is_ayt,
        'ranking_year': ranking_year,
        'exam_name': exam.name,
        'avg_net': avg_net,
        'avg_score': avg_score,
        'student_count': len(ranking_list),
    }
    if export_format == 'xlsx':
        return _export_rankings_xlsx(rows, columns, meta=meta, stats=stats, filename=filename, chart_ctx=chart_ctx)
    return _export_rankings_csv(rows, columns, meta=meta, stats=stats, filename=filename, chart_ctx=chart_ctx)


# ---------------------------------------------------------------------------
# Excel — PDF benzeri 3 satırlı başlık + grafik sayfası
# ---------------------------------------------------------------------------

def _header_groups(columns, sections: list[dict] | None, is_ayt: bool):
    """Excel 3 satırlı başlık için grup bilgisi."""
    sections = sections or []
    ordered = ordered_export_sections(sections)
    subs_map = _subs_by_parent(sections)
    groups = []
    # prefix
    for key, label in (('sira', 'Sıra'), ('student_name', 'İsim'), ('sinif', 'Sınıf')):
        groups.append({'kind': 'single', 'keys': [key], 'top': label, 'mid': '', 'bot': label})
    # section groups by parent
    i = 0
    while i < len(ordered):
        sec = ordered[i]
        parent_id = sec['id'] if not sec.get('is_sub_section') else sec.get('parent_id')
        members = [sec]
        j = i + 1
        while j < len(ordered):
            nxt = ordered[j]
            nxt_parent = nxt['id'] if not nxt.get('is_sub_section') else nxt.get('parent_id')
            if nxt_parent != parent_id:
                break
            members.append(nxt)
            j += 1
        top_name = _parent_name(members[0], sections)
        groups.append({
            'kind': 'section',
            'top': top_name,
            'members': [
                {
                    'keys': [_d_key(m['id']), _y_key(m['id']), _net_key(m['id'])],
                    'mid': _short_label(m.get('name') or 'Ders', _section_question_count(m, subs_map)),
                }
                for m in members
            ],
        })
        i = j
    groups.append({'kind': 'single', 'keys': ['toplam_net'], 'top': 'T.Net', 'mid': '', 'bot': 'T.Net'})
    if is_ayt:
        for pt in PT_KEYS:
            groups.append({
                'kind': 'pt',
                'top': PT_LABELS[pt],
                'members': [
                    {'key': f'puan_{pt.lower()}', 'mid': 'Puan'},
                    {'key': f'kurs_{pt.lower()}', 'mid': 'Kurs'},
                    {'key': f'genel_{pt.lower()}', 'mid': 'Genel'},
                ],
            })
    else:
        groups.append({'kind': 'single', 'keys': ['puan'], 'top': 'Puan', 'mid': '', 'bot': 'Puan'})
    year_label = next((c.label for c in columns if c.key == 'tahmini_siralama'), 'Tah.Sıra')
    for key, label in (
        ('kurum_ici_yuzdelik', 'K%'),
        ('tahmini_siralama', year_label),
        ('yuzdelik_dilim', 'TR%'),
    ):
        groups.append({'kind': 'single', 'keys': [key], 'top': label, 'mid': '', 'bot': label})
    return groups


def _export_rankings_xlsx(rows, columns, *, meta, stats, filename, chart_ctx):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from shared.export import style_manager as sm
    from shared.export.excel_export_service import ExcelExportService

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sıralama'

    cols = list(columns)
    key_to_col = {c.key: i + 1 for i, c in enumerate(cols)}
    num_cols = len(cols)
    header_row = ExcelExportService._write_letterhead(ws, meta, num_cols=num_cols)
    stats_end = ExcelExportService._write_stats(ws, stats, start_row=header_row, num_cols=num_cols)

    groups = _header_groups(cols, chart_ctx.get('sections'), chart_ctx.get('is_ayt'))
    r0, r1, r2 = stats_end, stats_end + 1, stats_end + 2

    thin = Side(style='thin', color=sm.BORDER_COLOR_HEX)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor=sm.BRAND_PRIMARY_HEX)
    header_font = Font(name=sm.FONT_NAME, size=9, bold=True, color=sm.BRAND_TEXT_ON_PRIMARY_HEX)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def style_header_cell(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    col_idx = 1
    for group in groups:
        if group['kind'] == 'single':
            keys = group['keys']
            start = col_idx
            end = col_idx + len(keys) - 1
            ws.merge_cells(start_row=r0, start_column=start, end_row=r2, end_column=end)
            cell = ws.cell(row=r0, column=start, value=group['top'])
            style_header_cell(cell)
            for rr in (r0, r1, r2):
                for cc in range(start, end + 1):
                    style_header_cell(ws.cell(row=rr, column=cc))
            col_idx = end + 1
        elif group['kind'] == 'section':
            members = group['members']
            start = col_idx
            end = col_idx + len(members) * 3 - 1
            if len(members) == 1:
                ws.merge_cells(start_row=r0, start_column=start, end_row=r1, end_column=end)
                cell = ws.cell(row=r0, column=start, value=members[0]['mid'] or group['top'])
                style_header_cell(cell)
                for rr in (r0, r1):
                    for cc in range(start, end + 1):
                        style_header_cell(ws.cell(row=rr, column=cc))
                for offset, bot in enumerate(('D', 'Y', 'Net')):
                    style_header_cell(ws.cell(row=r2, column=col_idx + offset, value=bot))
                col_idx += 3
            else:
                ws.merge_cells(start_row=r0, start_column=start, end_row=r0, end_column=end)
                cell = ws.cell(row=r0, column=start, value=group['top'])
                style_header_cell(cell)
                for cc in range(start, end + 1):
                    style_header_cell(ws.cell(row=r0, column=cc))
                for member in members:
                    ws.merge_cells(start_row=r1, start_column=col_idx, end_row=r1, end_column=col_idx + 2)
                    mid = ws.cell(row=r1, column=col_idx, value=member['mid'])
                    style_header_cell(mid)
                    for offset, bot in enumerate(('D', 'Y', 'Net')):
                        style_header_cell(ws.cell(row=r2, column=col_idx + offset, value=bot))
                        style_header_cell(ws.cell(row=r1, column=col_idx + offset))
                    col_idx += 3
        elif group['kind'] == 'pt':
            start = col_idx
            end = col_idx + 2
            ws.merge_cells(start_row=r0, start_column=start, end_row=r0, end_column=end)
            cell = ws.cell(row=r0, column=start, value=group['top'])
            style_header_cell(cell)
            for cc in range(start, end + 1):
                style_header_cell(ws.cell(row=r0, column=cc))
            for member in group['members']:
                ws.merge_cells(start_row=r1, start_column=col_idx, end_row=r2, end_column=col_idx)
                mid = ws.cell(row=r1, column=col_idx, value=member['mid'])
                style_header_cell(mid)
                style_header_cell(ws.cell(row=r2, column=col_idx))
                col_idx += 1

    ws.row_dimensions[r0].height = 18
    ws.row_dimensions[r1].height = 18
    ws.row_dimensions[r2].height = 16

    zebra_fill = PatternFill('solid', fgColor=sm.ZEBRA_FILL_HEX)
    avg_fill = PatternFill('solid', fgColor='DCEEFF')
    avg_font = Font(name=sm.FONT_NAME, size=sm.FONT_SIZE, bold=True, color=sm.BRAND_PRIMARY_HEX)
    data_font = Font(name=sm.FONT_NAME, size=sm.FONT_SIZE)

    data_start = r2 + 1
    for i, row in enumerate(rows):
        sheet_row = data_start + i
        is_avg = bool(row.get('_is_avg'))
        is_zebra = (not is_avg) and (i % 2 == 1)
        for col in cols:
            raw = row.get(col.key)
            excel_value = sm.excel_cell_value(raw, col)
            cell = ws.cell(row=sheet_row, column=key_to_col[col.key], value=excel_value)
            numfmt = sm.excel_number_format(col)
            if numfmt:
                cell.number_format = numfmt
            cell.font = avg_font if is_avg else data_font
            cell.alignment = Alignment(horizontal=col.alignment(), vertical='center')
            cell.border = border
            if is_avg:
                cell.fill = avg_fill
            elif is_zebra:
                cell.fill = zebra_fill

    last_row = data_start + len(rows) - 1 if rows else r2
    last_col_letter = get_column_letter(max(num_cols, 1))
    ws.auto_filter.ref = f'A{r0}:{last_col_letter}{max(last_row, r2)}'
    ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate

    for idx, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = col.width or 10

    ExcelExportService._apply_page_setup(
        ws, orientation='landscape', header_row=r0, last_row=max(last_row, r2),
        last_col=num_cols, report_title=meta.report_title,
    )
    ws.print_title_rows = f'{r0}:{r2}'

    _add_charts_sheet(wb, chart_ctx)

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{sm.safe_filename(filename)}.xlsx"'
    return response


def _add_charts_sheet(wb, ctx: dict):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from shared.export import style_manager as sm

    ws = wb.create_sheet('İstatistik Grafikleri')
    year = ctx.get('ranking_year') or ''
    exam_name = ctx.get('exam_name') or ''

    # Banner — landscape A4'e sığacak 3 kart + grafikler
    ws.merge_cells('A1:L1')
    title = ws['A1']
    title.value = f"{exam_name} — İstatistik Grafikleri"
    title.font = Font(name=sm.FONT_NAME, size=16, bold=True, color='FFFFFF')
    title.fill = PatternFill('solid', fgColor=sm.BRAND_PRIMARY_HEX)
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:L2')
    sub = ws['A2']
    sub.value = f"Tahmini sıralama yılı: {year}" if year else 'İstatistik özeti'
    sub.font = Font(name=sm.FONT_NAME, size=11, bold=True, color=sm.BRAND_PRIMARY_HEX)
    sub.fill = PatternFill('solid', fgColor='E8F1FA')
    sub.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Özet kartları
    cards = [
        ('Katılımcı', ctx.get('student_count') or 0),
        ('Ort. Net', ctx.get('avg_net') or 0),
        ('Ort. Puan', ctx.get('avg_score') or 0),
    ]
    thin = Side(style='thin', color=sm.BORDER_COLOR_HEX)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (label, value) in enumerate(cards):
        col = 1 + i * 2
        lc = ws.cell(row=4, column=col, value=label)
        lc.font = Font(name=sm.FONT_NAME, size=8, bold=True, color='64748B')
        lc.fill = PatternFill('solid', fgColor='F8FAFC')
        lc.alignment = Alignment(horizontal='center')
        lc.border = border
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        vc = ws.cell(row=5, column=col, value=value)
        vc.font = Font(name=sm.FONT_NAME, size=14, bold=True, color=sm.BRAND_PRIMARY_HEX)
        vc.fill = PatternFill('solid', fgColor='E8F1FA')
        vc.alignment = Alignment(horizontal='center')
        vc.border = border
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.cell(row=4, column=col + 1).border = border
        ws.cell(row=5, column=col + 1).border = border
        ws.cell(row=4, column=col + 1).fill = PatternFill('solid', fgColor='F8FAFC')
        ws.cell(row=5, column=col + 1).fill = PatternFill('solid', fgColor='E8F1FA')

    # Grafik verileri sağda (yazdırılmaz)
    data_col = 20
    datasets = _chart_datasets(ctx)
    chart_specs = []
    cursor = 1
    for title_text, pairs, color in datasets:
        if not pairs:
            continue
        ws.cell(row=cursor, column=data_col, value='Etiket')
        ws.cell(row=cursor, column=data_col + 1, value=title_text)
        for i, (label, value) in enumerate(pairs, start=1):
            ws.cell(row=cursor + i, column=data_col, value=label)
            ws.cell(row=cursor + i, column=data_col + 1, value=value)
        chart_specs.append((title_text, cursor, len(pairs), color))
        cursor += len(pairs) + 3

    ws.column_dimensions[get_column_letter(data_col)].hidden = True
    ws.column_dimensions[get_column_letter(data_col + 1)].hidden = True

    # Grafik boyutları — yatay A4 (~27×18 cm kullanılabilir alan)
    n = len(chart_specs)
    if n == 1:
        sizes = [(18, 11)]
        anchors = ['A7']
    elif n == 2:
        sizes = [(18, 7), (18, 7)]
        anchors = ['A7', 'A22']
    else:
        sizes = [(18, 6.2), (8.8, 6.2), (8.8, 6.2)]
        anchors = ['A7', 'A20', 'G20']

    for (title_text, start_row, count, color), (w, h), anchor in zip(chart_specs, sizes, anchors):
        chart = BarChart()
        chart.type = 'bar'
        chart.style = 10
        chart.title = title_text
        chart.y_axis.delete = False
        chart.x_axis.delete = False
        chart.legend = None
        data = Reference(ws, min_col=data_col + 1, min_row=start_row, max_row=start_row + count)
        cats = Reference(ws, min_col=data_col, min_row=start_row + 1, max_row=start_row + count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.gapWidth = 60
        labels = DataLabelList()
        labels.showVal = True
        chart.dataLabels = labels
        if chart.series:
            try:
                chart.series[0].graphicalProperties.solidFill = color
            except Exception:
                pass
        chart.width = w
        chart.height = h
        ws.add_chart(chart, anchor)

    for idx in range(1, 13):
        ws.column_dimensions[get_column_letter(idx)].width = 12

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = 'A1:L36'
    ws.page_setup.horizontalCentered = True
    ws.oddHeader.center.text = 'İSTATİSTİK GRAFİKLERİ'
    ws.oddFooter.right.text = 'Sayfa &P / &N'
    if year:
        ws.oddFooter.left.text = f'Tahmini sıralama yılı: {year}'


def _chart_datasets(ctx: dict) -> list[tuple[str, list[tuple[str, float]], str]]:
    sections = ctx.get('ordered_sections') or ordered_export_sections(ctx.get('sections'))
    section_avgs = ctx.get('section_avgs') or {}
    tests = []
    for sec in sections:
        avg = section_avgs.get(str(sec.get('id'))) or {}
        name = sec.get('name') or 'Ders'
        if len(name) > 18:
            name = name[:17] + '.'
        tests.append((name, float(avg.get('avg_net') or 0)))

    sinif_pairs = []
    for name in sorted((ctx.get('sinif_avgs') or {}).keys()):
        info = ctx['sinif_avgs'][name]
        label = name if len(name) <= 16 else name[:15] + '.'
        sinif_pairs.append((label, float(info.get('avg_net') or 0)))

    pt_pairs = []
    if ctx.get('is_ayt'):
        for pt in PT_KEYS:
            val = (ctx.get('puan_turleri_avgs') or {}).get(pt)
            if val is not None:
                pt_pairs.append((PT_LABELS[pt], float(val)))

    datasets = [('Test Bazlı Net Ortalamaları', tests, '0261A6')]
    if pt_pairs:
        datasets.append(('Puan Türü Ortalamaları', pt_pairs, 'CA8A04'))
    if len(sinif_pairs) > 1:
        datasets.append(('Şube Karşılaştırma (Ort. Net)', sinif_pairs, '0D9488'))
    return datasets


# ---------------------------------------------------------------------------
# CSV — PDF kolonları + metin istatistik bloğu
# ---------------------------------------------------------------------------

def _unicode_bar(value: float, max_v: float, width: int = 18) -> str:
    if max_v <= 0:
        return ''
    filled = int(round((float(value) / max_v) * width))
    filled = max(0, min(width, filled))
    return '█' * filled + '░' * (width - filled)


def _export_rankings_csv(rows, columns, *, meta, stats, filename, chart_ctx):
    from shared.export import style_manager as sm
    from shared.export.csv_export_service import CsvExportService
    from shared.export.style_manager import ExportColumn

    delimiter = ';'
    decimal_comma = True
    buf = io.StringIO()
    buf.write('\ufeff')
    writer = csv.writer(buf, delimiter=delimiter, lineterminator='\r\n')

    CsvExportService.write_letterhead_rows(writer, meta)

    for stat in stats or []:
        display = stat.value
        if stat.type in ('integer', 'decimal', 'percent', 'year'):
            display = sm.format_cell_display(
                stat.value,
                ExportColumn(key='_', label='_', type=stat.type),
            )
        writer.writerow([stat.label, display])
    writer.writerow([])

    year = chart_ctx.get('ranking_year')
    writer.writerow(['İSTATİSTİK GRAFİKLERİ'])
    if year:
        writer.writerow(['Tahmini sıralama yılı', year])
    writer.writerow([])

    for title_text, pairs, _color in _chart_datasets(chart_ctx):
        if not pairs:
            continue
        writer.writerow([title_text])
        writer.writerow(['Etiket', 'Değer', 'Grafik'])
        max_v = max((v for _, v in pairs), default=1) or 1
        for label, value in pairs:
            writer.writerow([label, sm.format_number_display(value, decimals=2), _unicode_bar(value, max_v)])
        writer.writerow([])

    writer.writerow([c.label for c in columns])
    for row in rows:
        writer.writerow([
            CsvExportService._csv_cell(row.get(c.key), c, decimal_comma=decimal_comma)
            for c in columns
        ])

    response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{sm.safe_filename(filename)}.csv"'
    return response
