"""Toplu kaynak kitap içe aktarma — şube kapsamlı + Excel açılır listeler."""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import BinaryIO

from django.db import transaction

from apps.egitim_tanimlari.models import Ders, SinifSeviyesi
from apps.resources.models import BookType, ResourceBook
from apps.resources.utils import generate_book_kod, normalize_kod

MAX_AD_LENGTH = 200
BULK_BATCH_SIZE = 200
LABEL_SEP = ' — '
ZORLUK_CHOICES = tuple(str(i) for i in range(0, 11))

HEADER_ALIASES = {
    'ad': ('kitap adı', 'kitap adi', 'ad', 'kitap', 'book', 'book name', 'title'),
    'kod': ('kod', 'code', 'kitap kodu'),
    'book_type': ('kitap türü', 'kitap turu', 'tür', 'tur', 'book_type', 'type'),
    'ders': ('ders', 'ders kodu', 'ders adı', 'ders adi', 'subject'),
    'sinif': ('sınıf', 'sinif', 'sınıf seviyesi', 'sinif seviyesi', 'sinif_seviyesi', 'grade'),
    'yayinevi': ('yayınevi', 'yayinevi', 'publisher'),
    'yazar': ('yazar', 'author'),
    'yayin_yili': ('yayın yılı', 'yayin yili', 'yayin_yili', 'yıl', 'yil', 'year'),
    'isbn': ('isbn',),
    'zorluk_min': ('zorluk min', 'zorluk_min', 'min zorluk'),
    'zorluk_max': ('zorluk max', 'zorluk_max', 'max zorluk'),
    'aciklama': ('açıklama', 'aciklama', 'description', 'not'),
}


@dataclass
class BulkImportResult:
    toplam_satir: int = 0
    eklenen: int = 0
    guncellenen: int = 0
    atlanan: int = 0
    hatali: int = 0
    hatalar: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            'toplam_satir': self.toplam_satir,
            'eklenen': self.eklenen,
            'guncellenen': self.guncellenen,
            'atlanan': self.atlanan,
            'hatali': self.hatali,
            'hatalar': self.hatalar,
        }


def _cell_str(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_at(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ''
    return _cell_str(row[index])


def _normalize_header(value: str) -> str:
    return (value or '').strip().casefold()


def _map_columns(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    normalized = [_normalize_header(h) for h in headers]
    for field_name, aliases in HEADER_ALIASES.items():
        alias_fold = {_normalize_header(a) for a in aliases}
        for idx, header in enumerate(normalized):
            if header in alias_fold or any(alias in header for alias in alias_fold):
                col_map[field_name] = idx
                break
    return col_map


def _chain_first_row(first_row, rows_iter):
    yield first_row
    yield from rows_iter


def _parse_int(raw: str) -> int | None:
    if not raw:
        return None
    try:
        return int(float(str(raw).replace(',', '.')))
    except (TypeError, ValueError):
        return None


def _format_kod_ad(kod: str, ad: str) -> str:
    kod_clean = (kod or '').strip()
    ad_clean = (ad or '').strip()
    if kod_clean and ad_clean:
        return f'{kod_clean}{LABEL_SEP}{ad_clean}'
    return kod_clean or ad_clean


def _split_label_candidates(raw: str) -> list[str]:
    """'KOD — Ad', 'KOD - Ad', 'Ad (KOD)' gibi hücrelerden aday parçalar çıkarır."""
    text = (raw or '').strip()
    if not text:
        return []
    candidates = [text]
    for sep in (LABEL_SEP, ' - ', ' – ', ' | ', '|'):
        if sep in text:
            left, right = text.split(sep, 1)
            candidates.extend([left.strip(), right.strip()])
            break
    if '(' in text and text.endswith(')'):
        before, _, inside = text[:-1].partition('(')
        candidates.extend([before.strip(), inside.strip()])
    # yinelenenleri koru, boşları at
    seen: set[str] = set()
    out: list[str] = []
    for c in candidates:
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    return out


def _lookup_by_kod_or_ad(raw: str, by_kod: dict, by_ad: dict):
    for part in _split_label_candidates(raw):
        keyed = normalize_kod(part)
        if keyed and keyed in by_kod:
            return by_kod[keyed]
        ad_key = _normalize_header(part)
        if ad_key in by_ad:
            return by_ad[ad_key]
    return None


class BulkBookImportService:
    def __init__(self, *, kurum_id: int, sube_id: int):
        self.kurum_id = kurum_id
        self.sube_id = sube_id
        book_types = list(BookType.objects.filter(aktif_mi=True).order_by('sira', 'ad'))
        dersler = list(Ders.objects.filter(aktif_mi=True, sube_id=sube_id).order_by('ad'))
        siniflar = list(
            SinifSeviyesi.objects.filter(aktif_mi=True, sube_id=sube_id).order_by('sira', 'ad')
        )
        self._book_types = {normalize_kod(bt.kod): bt for bt in book_types if bt.kod}
        self._book_types_by_ad = {_normalize_header(bt.ad): bt for bt in book_types if bt.ad}
        self._dersler = {normalize_kod(d.kod): d for d in dersler if d.kod}
        self._dersler_by_ad = {_normalize_header(d.ad): d for d in dersler if d.ad}
        self._siniflar = {normalize_kod(s.kod): s for s in siniflar if s.kod}
        self._siniflar_by_ad = {_normalize_header(s.ad): s for s in siniflar if s.ad}
        self._book_type_labels = [_format_kod_ad(bt.kod, bt.ad) for bt in book_types]
        self._ders_labels = [_format_kod_ad(d.kod, d.ad) for d in dersler]
        self._sinif_labels = [_format_kod_ad(s.kod, s.ad) for s in siniflar]
        self._existing_kods = set(
            ResourceBook.objects.filter(sube_id=sube_id).values_list('kod', flat=True)
        )

    def _resolve_book_type(self, raw: str) -> BookType | None:
        return _lookup_by_kod_or_ad(raw, self._book_types, self._book_types_by_ad)

    def _resolve_ders(self, raw: str) -> Ders | None:
        return _lookup_by_kod_or_ad(raw, self._dersler, self._dersler_by_ad)

    def _resolve_sinif(self, raw: str) -> SinifSeviyesi | None:
        if not raw:
            return None
        parts = [p.strip() for p in raw.replace(';', ',').split(',') if p.strip()]
        if not parts:
            return None
        return _lookup_by_kod_or_ad(parts[0], self._siniflar, self._siniflar_by_ad)

    def _resolve_sinif_list(self, raw: str) -> list[SinifSeviyesi]:
        if not raw:
            return []
        parts = [p.strip() for p in raw.replace(';', ',').split(',') if p.strip()]
        found: list[SinifSeviyesi] = []
        for part in parts:
            sinif = _lookup_by_kod_or_ad(part, self._siniflar, self._siniflar_by_ad)
            if sinif and sinif not in found:
                found.append(sinif)
        return found

    def _available_hint(self, labels: list[str], limit: int = 8) -> str:
        if not labels:
            return ' (sistemde kayıt yok)'
        shown = ', '.join(labels[:limit])
        more = f' …(+{len(labels) - limit})' if len(labels) > limit else ''
        return f' Örnekler: {shown}{more}'

    def import_rows(self, rows: list[dict]) -> BulkImportResult:
        result = BulkImportResult()
        seen_kods: set[str] = set()
        to_create: list[tuple[ResourceBook, list[SinifSeviyesi]]] = []

        for idx, raw in enumerate(rows, start=1):
            ad = str(raw.get('ad') or '').strip()
            if not ad:
                if any(str(raw.get(k) or '').strip() for k in ('ders', 'book_type', 'sinif', 'kod')):
                    result.toplam_satir += 1
                    result.hatali += 1
                    result.hatalar.append({'satir': idx, 'ad': '', 'neden': 'Kitap adı boş'})
                continue

            # Yalnızca şablonun varsayılan örnek satırını yoksay
            # (kullanıcı örnek satırı düzenlediyse artık atlanmaz)
            if _is_template_sample_row(ad, str(raw.get('aciklama') or '')):
                continue

            result.toplam_satir += 1
            if len(ad) > MAX_AD_LENGTH:
                result.hatali += 1
                result.hatalar.append({'satir': idx, 'ad': ad, 'neden': 'Kitap adı çok uzun'})
                continue

            book_type_raw = str(raw.get('book_type') or '').strip()
            book_type = self._resolve_book_type(book_type_raw)
            if not book_type:
                neden = (
                    'Kitap türü boş — Excel’de Kitap Türü sütunundan seçin.'
                    if not book_type_raw
                    else f'Kitap türü bulunamadı: “{book_type_raw}”.'
                )
                neden += self._available_hint(self._book_type_labels)
                result.hatali += 1
                result.hatalar.append({'satir': idx, 'ad': ad, 'neden': neden})
                continue

            ders_raw = str(raw.get('ders') or '').strip()
            ders = self._resolve_ders(ders_raw)
            if not ders:
                neden = (
                    'Ders boş — Excel’de Ders sütunundan seçin.'
                    if not ders_raw
                    else f'Ders bulunamadı: “{ders_raw}” (bu şube).'
                )
                neden += self._available_hint(self._ders_labels)
                result.hatali += 1
                result.hatalar.append({'satir': idx, 'ad': ad, 'neden': neden})
                continue

            sinif_raw = str(raw.get('sinif') or '')
            sinif_list = self._resolve_sinif_list(sinif_raw)
            sinif = sinif_list[0] if sinif_list else self._resolve_sinif(sinif_raw)
            if not sinif:
                neden = (
                    'Sınıf boş — Excel’de Sınıf sütunundan seçin.'
                    if not sinif_raw.strip()
                    else f'Sınıf seviyesi bulunamadı: “{sinif_raw}” (bu şube).'
                )
                neden += self._available_hint(self._sinif_labels)
                result.hatali += 1
                result.hatalar.append({'satir': idx, 'ad': ad, 'neden': neden})
                continue
            if not sinif_list:
                sinif_list = [sinif]

            zorluk_min = _parse_int(str(raw.get('zorluk_min') or ''))
            zorluk_max = _parse_int(str(raw.get('zorluk_max') or ''))
            zorluk_error = None
            if zorluk_min is not None and (zorluk_min < 0 or zorluk_min > 10):
                zorluk_error = f'Zorluk Min 0-10 arasında olmalı (girilen: {zorluk_min})'
            elif zorluk_max is not None and (zorluk_max < 0 or zorluk_max > 10):
                zorluk_error = f'Zorluk Max 0-10 arasında olmalı (girilen: {zorluk_max})'
            elif (
                zorluk_min is not None
                and zorluk_max is not None
                and zorluk_min > zorluk_max
            ):
                zorluk_error = 'Zorluk Min, Zorluk Max’tan büyük olamaz'
            if zorluk_error:
                result.hatali += 1
                result.hatalar.append({'satir': idx, 'ad': ad, 'neden': zorluk_error})
                continue

            kod_raw = str(raw.get('kod') or '').strip()
            kod = normalize_kod(kod_raw) if kod_raw else generate_book_kod(
                self.kurum_id, book_type, ders, sube_id=self.sube_id,
            )
            if not kod_raw:
                base = kod
                n = 1
                while kod in self._existing_kods or kod in seen_kods:
                    n += 1
                    kod = f'{base.rsplit("_", 1)[0]}_{n:03d}' if '_' in base else f'{base}_{n:03d}'

            if kod in self._existing_kods:
                result.atlanan += 1
                seen_kods.add(kod)
                continue
            if kod in seen_kods:
                result.hatali += 1
                result.hatalar.append({
                    'satir': idx, 'ad': ad,
                    'neden': 'Aynı Excel dosyasında tekrar eden kod',
                })
                continue

            seen_kods.add(kod)
            book = ResourceBook(
                ad=ad,
                kod=kod,
                kurum_id=self.kurum_id,
                sube_id=self.sube_id,
                book_type=book_type,
                ders=ders,
                sinif_seviyesi=sinif,
                yayinevi=str(raw.get('yayinevi') or '').strip()[:200],
                yazar=str(raw.get('yazar') or '').strip()[:200],
                yayin_yili=_parse_int(str(raw.get('yayin_yili') or '')),
                isbn=str(raw.get('isbn') or '').strip()[:20],
                zorluk_min=zorluk_min,
                zorluk_max=zorluk_max,
                aciklama=str(raw.get('aciklama') or '').strip(),
                aktif_mi=True,
            )
            to_create.append((book, sinif_list))

        if to_create:
            with transaction.atomic():
                books = [b for b, _ in to_create]
                ResourceBook.objects.bulk_create(books, batch_size=BULK_BATCH_SIZE)
                for book, sinif_list in to_create:
                    if book.pk:
                        book.sinif_seviyeleri.set(sinif_list)
            result.eklenen = len(to_create)

        if (
            result.eklenen == 0
            and result.atlanan == 0
            and result.hatali == 0
            and result.toplam_satir == 0
        ):
            result.hatalar.append({
                'satir': 0,
                'ad': '',
                'neden': (
                    'İşlenecek satır bulunamadı. Veriyi "Kitaplar" sayfasına girin; '
                    'Kitap Adı, Kitap Türü, Ders ve Sınıf doldurulmalı. '
                    'Şablondaki örnek satırı kullanıyorsanız Kitap Adı’nı değiştirin '
                    '(“Örnek Kitap…” ile başlayan adlar yok sayılır).'
                ),
            })
            result.hatali = 1

        return result

    def parse_excel(self, file_obj: BinaryIO) -> list[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = _pick_books_sheet(wb)
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if not first_row:
            wb.close()
            return []

        headers = [_cell_str(c) for c in first_row]
        col_map = _map_columns(headers)
        start_iter = rows_iter

        if col_map.get('ad') is None:
            col_map = {
                'ad': 0, 'kod': 1, 'book_type': 2, 'ders': 3, 'sinif': 4,
                'yayinevi': 5, 'yazar': 6, 'yayin_yili': 7,
                'isbn': 8, 'zorluk_min': 9, 'zorluk_max': 10, 'aciklama': 11,
            }
            start_iter = _chain_first_row(first_row, rows_iter)

        parsed_rows: list[dict] = []
        for row in start_iter:
            if not row or all(_cell_str(c) == '' for c in row):
                continue
            first_cell = _cell_at(row, col_map.get('ad'))
            if first_cell.casefold().startswith('not:') or first_cell.casefold().startswith('açıklama:'):
                continue
            parsed_rows.append({
                'ad': _cell_at(row, col_map.get('ad')),
                'kod': _cell_at(row, col_map.get('kod')),
                'book_type': _cell_at(row, col_map.get('book_type')),
                'ders': _cell_at(row, col_map.get('ders')),
                'sinif': _cell_at(row, col_map.get('sinif')),
                'yayinevi': _cell_at(row, col_map.get('yayinevi')),
                'yazar': _cell_at(row, col_map.get('yazar')),
                'yayin_yili': _cell_at(row, col_map.get('yayin_yili')),
                'isbn': _cell_at(row, col_map.get('isbn')),
                'zorluk_min': _cell_at(row, col_map.get('zorluk_min')),
                'zorluk_max': _cell_at(row, col_map.get('zorluk_max')),
                'aciklama': _cell_at(row, col_map.get('aciklama')),
            })
        wb.close()
        return parsed_rows

    def import_excel(self, file_obj: BinaryIO) -> BulkImportResult:
        return self.import_rows(self.parse_excel(file_obj))


def _is_template_sample_row(ad: str, aciklama: str = '') -> bool:
    """Şablondaki varsayılan örnek satır mı? Yalnızca ad 'Örnek Kitap…' ise atlanır."""
    ad_l = (ad or '').strip().casefold()
    return ad_l.startswith('örnek kitap')


def _pick_books_sheet(wb):
    """Kitaplar sayfasını bul; yoksa 'Kitap Adı' başlıklı sayfayı ara."""
    if 'Kitaplar' in wb.sheetnames:
        return wb['Kitaplar']
    for name in wb.sheetnames:
        ws = wb[name]
        if getattr(ws, 'sheet_state', 'visible') == 'hidden':
            continue
        try:
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        except Exception:
            continue
        if not first:
            continue
        headers = ' '.join(_cell_str(c).casefold() for c in first)
        if 'kitap adı' in headers or 'kitap adi' in headers:
            return ws
    return wb.active


def _write_list_column(ws, col: int, header: str, values: list[str]) -> int:
    """Listeler sayfasına sütun yazar; döner: son satır numarası (>=1)."""
    from openpyxl.utils import get_column_letter

    ws.cell(1, col, header)
    for i, val in enumerate(values, start=2):
        ws.cell(i, col, val)
    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = max(18, min(48, max((len(v) for v in values), default=12) + 2))
    return max(1, len(values) + 1)


def build_excel_template(*, kurum_id: int | None = None, sube_id: int | None = None) -> bytes:
    """
    Aktif şubedeki ders/sınıf ve sistemdeki kitap türleri ile
    açılır liste (data validation) içeren şablon üretir.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    book_types = list(BookType.objects.filter(aktif_mi=True).order_by('sira', 'ad'))
    ders_qs = Ders.objects.filter(aktif_mi=True)
    sinif_qs = SinifSeviyesi.objects.filter(aktif_mi=True)
    if sube_id:
        ders_qs = ders_qs.filter(sube_id=sube_id)
        sinif_qs = sinif_qs.filter(sube_id=sube_id)
    elif kurum_id:
        ders_qs = ders_qs.filter(kurum_id=kurum_id)
        sinif_qs = sinif_qs.filter(kurum_id=kurum_id)
    dersler = list(ders_qs.order_by('ad'))
    siniflar = list(sinif_qs.order_by('sira', 'ad'))

    type_labels = [_format_kod_ad(bt.kod, bt.ad) for bt in book_types]
    ders_labels = [_format_kod_ad(d.kod, d.ad) for d in dersler]
    sinif_labels = [_format_kod_ad(s.kod, s.ad) for s in siniflar]

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # —— Talimat ——
    ws_help = wb.create_sheet('Talimat', 0)
    ws_help['A1'] = 'Toplu Kitap Yükleme Talimatı'
    ws_help['A1'].font = Font(bold=True, size=14)
    help_lines = [
        '',
        '1. "Kitaplar" sayfasında satırlara kitap girin.',
        '2. Kitap Türü, Ders, Sınıf ve Zorluk sütunlarında hücreye tıklayınca açılır listeden seçin.',
        '3. Kod boş bırakılırsa sistem otomatik üretir.',
        '4. Birden fazla sınıf için Sınıf hücresine virgülle ayırarak yazabilirsiniz (listeden seçtikten sonra düzenleyin).',
        '5. Zorluk Min / Max: 0–10 arası.',
        '6. "Kitaplar" sayfasında 2. satırdan itibaren doldurun (Kitap Adı zorunlu).',
        '7. Dosyayı kaydedip sisteme yükleyin.',
        '',
        f'Kitap türü sayısı: {len(type_labels)}',
        f'Ders sayısı (aktif şube): {len(ders_labels)}',
        f'Sınıf seviyesi sayısı (aktif şube): {len(sinif_labels)}',
        '',
        'Eşleşme: Açılır listedeki "KOD — Ad" biçimi, yalnız kod veya yalnız ad da kabul edilir.',
    ]
    for i, line in enumerate(help_lines, start=2):
        ws_help.cell(i, 1, line)
    ws_help.column_dimensions['A'].width = 96

    # —— Kitaplar ——
    ws = wb.create_sheet('Kitaplar', 1)

    # —— Listeler (açılır kaynak, gizli) ——
    ws_lists = wb.create_sheet('Listeler', 2)
    end_type = _write_list_column(ws_lists, 1, 'Kitap Türleri', type_labels)
    end_ders = _write_list_column(ws_lists, 2, 'Dersler', ders_labels)
    end_sinif = _write_list_column(ws_lists, 3, 'Sınıf Seviyeleri', sinif_labels)
    end_zorluk = _write_list_column(ws_lists, 4, 'Zorluk (0-10)', list(ZORLUK_CHOICES))
    ws_lists.sheet_state = 'hidden'
    headers = [
        'Kitap Adı', 'Kod', 'Kitap Türü', 'Ders', 'Sınıf',
        'Yayınevi', 'Yazar', 'Yayın Yılı', 'ISBN', 'Zorluk Min', 'Zorluk Max', 'Açıklama',
    ]
    header_fill = PatternFill('solid', fgColor='1F3C88')
    header_font = Font(color='FFFFFF', bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Boş veri satırları (açılır listeler hazır); örnek kitap satırı yok —
    # kullanıcı doğrudan 2. satırdan doldurur.
    for _ in range(5):
        ws.append(['', '', '', '', '', '', '', '', '', '', '', ''])

    widths = [36, 16, 28, 22, 22, 16, 16, 12, 14, 12, 12, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Açılır listeler (1000 satıra kadar)
    max_row = 1000

    def add_list_dv(col_letter: str, list_col: str, end_row: int, prompt: str):
        if end_row < 2:
            # Liste boş — sabit uyarı listesi
            formula = '"Liste boş — sistemde kayıt yok"'
        else:
            formula = f'Listeler!${list_col}$2:${list_col}${end_row}'
        dv = DataValidation(
            type='list',
            formula1=formula,
            allow_blank=True,
            showDropDown=False,  # False = göster (openpyxl ters anlamlı)
            showErrorMessage=True,
            errorTitle='Geçersiz seçim',
            error='Lütfen listeden bir değer seçin.',
            promptTitle=prompt,
            prompt='Listeden seçin',
            showInputMessage=True,
        )
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{max_row}')

    add_list_dv('C', 'A', end_type, 'Kitap Türü')
    add_list_dv('D', 'B', end_ders, 'Ders')
    add_list_dv('E', 'C', end_sinif, 'Sınıf')
    add_list_dv('J', 'D', end_zorluk, 'Zorluk Min')
    add_list_dv('K', 'D', end_zorluk, 'Zorluk Max')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
