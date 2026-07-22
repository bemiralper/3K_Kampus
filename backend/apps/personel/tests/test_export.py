"""Personel listesi dışa aktarma (Excel/CSV) — kurumsal export entegrasyonu."""
from django.contrib.auth import get_user_model
from django.test import Client, TestCase

from apps.egitim_yili.domain.models import EgitimYili
from apps.kurum.domain.models import Kurum
from apps.personel.domain.models import Personel, PersonelGorevlendirme
from apps.roller.models import Permission, Role, RolePermission, UserRole
from apps.sube.domain.models import Sube

User = get_user_model()


def _assign_personel_read(user, kurum):
    role, _ = Role.objects.get_or_create(
        code='personel_export_test',
        defaults={'name': 'Personel Export Test', 'level': 100, 'is_system_role': True},
    )
    perm, _ = Permission.objects.get_or_create(
        code='personel.read',
        defaults={'name': 'personel.read', 'module': 'personel', 'permission_type': 'read'},
    )
    RolePermission.objects.get_or_create(role=role, permission=perm)
    UserRole.objects.update_or_create(user=user, defaults={'role': role, 'kurum': kurum})


class PersonelExportAPITest(TestCase):
    def setUp(self):
        self.client = Client()
        self.kurum = Kurum.objects.create(ad='Export Kurum', kod='EXK')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez Şube', kod='EXK-M')
        self.yil = EgitimYili.objects.create(
            baslangic_yil=2025,
            bitis_yil=2026,
            aktif_mi=True,
        )
        self.ogretmen_rol, _ = Role.objects.get_or_create(
            code='ogretmen',
            defaults={'name': 'Öğretmen', 'level': 50, 'is_system_role': True},
        )
        self.mudur_rol, _ = Role.objects.get_or_create(
            code='mudur',
            defaults={'name': 'Müdür', 'level': 10, 'is_system_role': True},
        )

        self.ogretmen = Personel.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            ad='Ayşe',
            soyad='Yılmaz',
            tc_kimlik_no='11111111110',
            telefon='5551112233',
            cep_telefon='5551112233',
            email='ayse@example.com',
            aktif_mi=True,
        )
        PersonelGorevlendirme.objects.create(
            kurum=self.kurum,
            personel=self.ogretmen,
            egitim_yili=self.yil,
            gorev_sube=self.sube,
            rol=self.ogretmen_rol,
            gorev_baslangic='2025-09-01',
            aktif_mi=True,
        )

        self.mudur = Personel.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            ad='Mehmet',
            soyad='Demir',
            tc_kimlik_no='22222222220',
            telefon='5552223344',
            email='mehmet@example.com',
            aktif_mi=True,
        )
        PersonelGorevlendirme.objects.create(
            kurum=self.kurum,
            personel=self.mudur,
            egitim_yili=self.yil,
            gorev_sube=self.sube,
            rol=self.mudur_rol,
            aktif_mi=True,
        )

        self.pasif_personel = Personel.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            ad='Pasif',
            soyad='Personel',
            aktif_mi=False,
        )

        self.user = User.objects.create_user(username='exportuser', password='test')
        _assign_personel_read(self.user, self.kurum)
        self.client.force_login(self.user)

    def _ctx_headers(self):
        return {
            'HTTP_X_KURUM_ID': str(self.kurum.id),
            'HTTP_X_SUBE_ID': str(self.sube.id),
            'HTTP_X_EGITIMYILI_ID': str(self.yil.id),
        }

    def test_export_csv_returns_200_with_headers(self):
        res = self.client.get('/personel/api/export/?format=csv', **self._ctx_headers())
        self.assertEqual(res.status_code, 200)
        self.assertEqual(res['Content-Type'], 'text/csv; charset=utf-8')
        content = res.content.decode('utf-8-sig')
        self.assertIn('Ad Soyad', content)
        self.assertIn('TC Kimlik No', content)
        self.assertIn('Rol / Pozisyon', content)
        self.assertIn('Ayşe Yılmaz', content)
        self.assertIn('Mehmet Demir', content)
        # Pasif personel show_inactive=false (varsayılan) iken dahil edilmemeli
        self.assertNotIn('Pasif Personel', content)

    def test_export_csv_includes_inactive_when_requested(self):
        res = self.client.get(
            '/personel/api/export/?format=csv&show_inactive=true', **self._ctx_headers(),
        )
        self.assertEqual(res.status_code, 200)
        content = res.content.decode('utf-8-sig')
        self.assertIn('Pasif Personel', content)

    def test_export_xlsx_returns_200_and_nonempty_file(self):
        res = self.client.get('/personel/api/export/?format=xlsx', **self._ctx_headers())
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            res['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertGreater(len(res.content), 5000)

        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        all_text = '\n'.join(
            str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
        )
        self.assertIn('Ayşe Yılmaz', all_text)
        self.assertIn('Öğretmen', all_text)
        self.assertIn('PERSONEL LİSTESİ', all_text)

    def test_export_search_filters_rows(self):
        res = self.client.get(
            '/personel/api/export/?format=csv&q=Mehmet', **self._ctx_headers(),
        )
        self.assertEqual(res.status_code, 200)
        content = res.content.decode('utf-8-sig')
        self.assertIn('Mehmet Demir', content)
        self.assertNotIn('Ayşe Yılmaz', content)

    def test_export_requires_sube_context(self):
        res = self.client.get(
            '/personel/api/export/?format=csv', HTTP_X_KURUM_ID=str(self.kurum.id),
        )
        self.assertEqual(res.status_code, 400)
