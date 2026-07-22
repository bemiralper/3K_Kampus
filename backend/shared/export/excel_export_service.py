"""
ExcelExportService — 3K Kampüs kurumsal Excel (.xlsx) dışa aktarma servisi.

Tüm liste ekranları (öğrenci, personel, veli, öğretmen, finans, yoklama,
görüşme, kütüphane, deneme, sınıf vb.) bu servisi kullanarak kurumsal
kimliğe uygun, okunaklı, filtrelenebilir ve yazdırılabilir Excel raporları
üretmelidir. Bkz. `ExportStyleManager` (stil/format kuralları) ve
`CsvExportService` (CSV karşılığı).
"""
from __future__ import annotations

import io
from typing import Any, Sequence

from django.http import HttpResponse

from shared.export import style_manager as sm
from shared.export.style_manager import ExportColumn, ExportStat, ReportMeta


class ExcelExportService:
    """Kurumsal şablonla tek bir sayfalık liste/rapor Excel dosyası üretir."""

    HEADER_BLOCK_ROWS_NO_LOGO = 7  # yazılım adı, rapor adı + 5 bilgi satırı

    @classmethod
    def export(
        cls,
        rows: Sequence[dict[str, Any]],
        columns: Sequence[ExportColumn | dict],
        *,
        meta: ReportMeta | dict[str, Any],
        stats: Sequence[ExportStat | dict] | None = None,
        orientation: str = "landscape",
        sheet_name: str | None = None,
        filename: str | None = None,
    ) -> HttpResponse:
        wb = cls.build_workbook(rows, columns, meta=meta, stats=stats, orientation=orientation, sheet_name=sheet_name)
        buf = io.BytesIO()
        wb.save(buf)
        title = filename or (meta.report_title if isinstance(meta, ReportMeta) else meta.get("report_title"))
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{sm.safe_filename(title)}.xlsx"'
        return response

    @classmethod
    def build_workbook(
        cls,
        rows: Sequence[dict[str, Any]],
        columns: Sequence[ExportColumn | dict],
        *,
        meta: ReportMeta | dict[str, Any],
        stats: Sequence[ExportStat | dict] | None = None,
        orientation: str = "landscape",
        sheet_name: str | None = None,
    ):
        from openpyxl import Workbook

        if isinstance(meta, dict):
            meta = ReportMeta(**meta)
        cols = sm.normalize_columns(columns)
        stat_list = sm.normalize_stats(stats)

        wb = Workbook()
        ws = wb.active
        ws.title = sm.safe_sheet_title(sheet_name or meta.report_title)

        header_row = cls._write_letterhead(ws, meta, num_cols=len(cols))
        stats_end_row = cls._write_stats(ws, stat_list, start_row=header_row, num_cols=len(cols))
        table_header_row = stats_end_row
        last_row, last_col, col_widths, col_needs_wrap = cls._write_table(
            ws, rows, cols, header_row=table_header_row,
        )
        cls._apply_page_setup(ws, orientation=orientation, header_row=table_header_row, last_row=last_row,
                               last_col=last_col, report_title=meta.report_title)
        cls._apply_column_widths(ws, cols, col_widths, col_needs_wrap)
        ws.freeze_panes = ws.cell(row=table_header_row + 1, column=1).coordinate
        return wb

    # ------------------------------------------------------------------
    # Letterhead (kurumsal başlık bloğu)
    # ------------------------------------------------------------------
    @classmethod
    def write_letterhead(cls, ws, meta: ReportMeta | dict[str, Any], *, num_cols: int = 6) -> int:
        """Kurumsal başlık bloğunu (logo, yazılım adı, rapor adı, kurum/şube/tarih/oluşturan)
        verilen worksheet'in üstüne yazar ve bir sonraki boş satırı döner.

        Tek tablolu export'lar dışında (örn. çok bölümlü gün sonu raporları) da
        kurumsal görünüm tutarlılığı için kullanılabilir.
        """
        if isinstance(meta, dict):
            meta = ReportMeta(**meta)
        return cls._write_letterhead(ws, meta, num_cols=num_cols)

    @classmethod
    def apply_page_setup(cls, ws, *, orientation: str = "landscape", header_row: int = 1, last_row: int = 1,
                          last_col: int = 1, report_title: str = ""):
        """Sayfa ayarlarını (A4, yön, sığdır, tekrarlayan başlık, üst/alt bilgi) uygular."""
        cls._apply_page_setup(ws, orientation=orientation, header_row=header_row, last_row=last_row,
                               last_col=last_col, report_title=report_title)

    @classmethod
    def _write_letterhead(cls, ws, meta: ReportMeta, *, num_cols: int) -> int:
        from openpyxl.styles import Alignment, Font
        from openpyxl.drawing.image import Image as XLImage

        total_cols = max(num_cols, 6)
        text_start_col = 1
        logo_path = sm.resolve_logo_path()
        if logo_path is not None:
            try:
                img = XLImage(str(logo_path))
                img.width = 56
                img.height = 56
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 30
                ws.row_dimensions[2].height = 30
                text_start_col = 2
            except Exception:
                text_start_col = 1

        row = 1
        ws.merge_cells(start_row=row, start_column=text_start_col, end_row=row, end_column=max(total_cols, text_start_col))
        c = ws.cell(row=row, column=text_start_col, value=sm.tr_upper(sm.SOFTWARE_FULL_NAME))
        c.font = Font(name=sm.FONT_NAME, size=9, bold=True, color="8C99A6")
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        ws.merge_cells(start_row=row, start_column=text_start_col, end_row=row, end_column=max(total_cols, text_start_col))
        c = ws.cell(row=row, column=text_start_col, value=sm.tr_upper(meta.report_title or "RAPOR"))
        c.font = Font(name=sm.FONT_NAME, size=16, bold=True, color=sm.BRAND_PRIMARY_HEX)
        c.alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        info_pairs = [
            ("Kurum", meta.kurum_ad or "—"),
            ("Şube", meta.sube_ad or "—"),
            ("Eğitim Yılı", meta.egitim_yili or "—"),
            ("Oluşturma Tarihi", sm.format_datetime_display(meta.resolved_generated_at())),
            ("Oluşturan", meta.generated_by or "—"),
        ]
        for label, value in info_pairs:
            label_cell = ws.cell(row=row, column=1, value=f"{label} :")
            label_cell.font = Font(name=sm.FONT_NAME, size=10, bold=True, color="475569")
            label_cell.alignment = Alignment(horizontal="left")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max(total_cols, 2))
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.font = Font(name=sm.FONT_NAME, size=10, color="1E293B")
            value_cell.alignment = Alignment(horizontal="left")
            row += 1

        row += 1  # boş satır
        return row

    # ------------------------------------------------------------------
    # İstatistik özet kutuları
    # ------------------------------------------------------------------
    @classmethod
    def _write_stats(cls, ws, stats: list[ExportStat], *, start_row: int, num_cols: int) -> int:
        if not stats:
            return start_row
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        total_cols = max(num_cols, len(stats))
        base_span = total_cols // len(stats)
        remainder = total_cols - base_span * len(stats)

        thin = Side(style="thin", color=sm.BORDER_COLOR_HEX)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        label_row = start_row
        value_row = start_row + 1

        col = 1
        for idx, stat in enumerate(stats):
            span = base_span + (1 if idx < remainder else 0)
            span = max(span, 1)
            end_col = col + span - 1

            ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=end_col)
            lc = ws.cell(row=label_row, column=col, value=sm.tr_upper(stat.label))
            lc.font = Font(name=sm.FONT_NAME, size=8, bold=True, color="64748B")
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.fill = PatternFill("solid", fgColor=sm.STAT_LABEL_FILL_HEX)
            lc.border = border
            for cc in range(col, end_col + 1):
                ws.cell(row=label_row, column=cc).border = border
                ws.cell(row=label_row, column=cc).fill = PatternFill("solid", fgColor=sm.STAT_LABEL_FILL_HEX)

            display_value = stat.value
            if stat.type == "currency":
                display_value = sm.format_currency_display(stat.value, stat.currency)
            elif stat.type in ("integer", "decimal", "percent"):
                display_value = sm.format_cell_display(stat.value, ExportColumn(key="_", label="_", type=stat.type,
                                                                                 currency=stat.currency))

            ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row, end_column=end_col)
            vc = ws.cell(row=value_row, column=col, value=display_value)
            vc.font = Font(name=sm.FONT_NAME, size=13, bold=True, color=sm.BRAND_PRIMARY_HEX)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.fill = PatternFill("solid", fgColor=sm.STAT_VALUE_FILL_HEX)
            vc.border = border
            for cc in range(col, end_col + 1):
                ws.cell(row=value_row, column=cc).border = border
                ws.cell(row=value_row, column=cc).fill = PatternFill("solid", fgColor=sm.STAT_VALUE_FILL_HEX)

            col = end_col + 1

        ws.row_dimensions[label_row].height = 16
        ws.row_dimensions[value_row].height = 22
        return value_row + 2  # bir boş satır bırak

    # ------------------------------------------------------------------
    # Veri tablosu
    # ------------------------------------------------------------------
    @classmethod
    def _write_table(cls, ws, rows: Sequence[dict[str, Any]], cols: list[ExportColumn], *, header_row: int,
                     auto_filter: bool = True):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin = Side(style="thin", color=sm.BORDER_COLOR_HEX)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor=sm.BRAND_PRIMARY_HEX)
        header_font = Font(name=sm.FONT_NAME, size=11, bold=True, color=sm.BRAND_TEXT_ON_PRIMARY_HEX)
        zebra_fill = PatternFill("solid", fgColor=sm.ZEBRA_FILL_HEX)

        for col_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col.label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[header_row].height = 22

        display_values_per_col: list[list[str]] = [[] for _ in cols]
        data_row = header_row + 1
        for i, row in enumerate(rows):
            row_texts: list[tuple[str, int]] = []
            is_zebra = (i % 2 == 1)
            for col_idx, col in enumerate(cols, start=1):
                raw = row.get(col.key)
                excel_value = sm.excel_cell_value(raw, col)
                cell = ws.cell(row=data_row, column=col_idx, value=excel_value)
                numfmt = sm.excel_number_format(col)
                if numfmt:
                    cell.number_format = numfmt
                cell.font = Font(name=sm.FONT_NAME, size=sm.FONT_SIZE)
                align = col.alignment()
                wrap = col.wrap if col.wrap is not None else (col.type == "text")
                cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
                cell.border = border
                if is_zebra:
                    cell.fill = zebra_fill

                display_text = sm.format_cell_display(raw, col)
                display_values_per_col[col_idx - 1].append(display_text)
                row_texts.append((display_text, col.width or sm.MAX_COLUMN_WIDTH))
            data_row += 1

        last_row = data_row - 1
        col_widths: list[int] = []
        col_needs_wrap: list[bool] = []
        for col_idx, col in enumerate(cols):
            if col.width:
                col_widths.append(col.width)
                col_needs_wrap.append(False)
                continue
            width, needs_wrap = sm.compute_column_width(col.label, display_values_per_col[col_idx])
            col_widths.append(width)
            col_needs_wrap.append(needs_wrap or col.type == "text" and width >= sm.MAX_COLUMN_WIDTH)

        # Wrap gerektiren satırların yüksekliğini ikinci geçişte ayarla.
        if any(col_needs_wrap):
            for r_idx, row in enumerate(rows):
                sheet_row = header_row + 1 + r_idx
                pairs = []
                for col_idx, col in enumerate(cols):
                    if not col_needs_wrap[col_idx]:
                        continue
                    text = display_values_per_col[col_idx][r_idx] if r_idx < len(display_values_per_col[col_idx]) else ""
                    pairs.append((text, col_widths[col_idx]))
                if pairs:
                    height = sm.estimate_wrapped_row_height(pairs)
                    if height > 18:
                        ws.row_dimensions[sheet_row].height = height

        from openpyxl.utils import get_column_letter
        last_col_letter = get_column_letter(max(len(cols), 1))
        if auto_filter:
            ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{max(last_row, header_row)}"

        return last_row, len(cols), col_widths, col_needs_wrap

    # ------------------------------------------------------------------
    # Sayfa ayarları / yazdırma
    # ------------------------------------------------------------------
    @classmethod
    def _apply_page_setup(cls, ws, *, orientation: str, header_row: int, last_row: int, last_col: int,
                           report_title: str = ""):
        from openpyxl.utils import get_column_letter

        ori = "landscape" if (orientation or "landscape").lower() in ("landscape", "yatay", "horizontal") else "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ori
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = f"{header_row}:{header_row}"
        last_col_letter = get_column_letter(max(last_col, 1))
        ws.print_area = f"A1:{last_col_letter}{max(last_row, header_row)}"

        ws.oddHeader.center.text = sm.tr_upper(report_title or "")
        ws.oddHeader.center.size = 9
        ws.oddFooter.left.text = sm.SOFTWARE_URL
        ws.oddFooter.left.size = 8
        ws.oddFooter.right.text = "Sayfa &P / &N"
        ws.oddFooter.right.size = 8
        ws.evenHeader.center.text = sm.tr_upper(report_title or "")
        ws.evenFooter.left.text = sm.SOFTWARE_URL
        ws.evenFooter.right.text = "Sayfa &P / &N"

    @classmethod
    def _apply_column_widths(cls, ws, cols: list[ExportColumn], col_widths: list[int], col_needs_wrap: list[bool]):
        from openpyxl.utils import get_column_letter

        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
