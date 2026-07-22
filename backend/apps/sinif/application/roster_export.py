"""Sınıf öğrenci listesi (roster) dışa aktarma — dönem yerleşimi bazlı."""
from __future__ import annotations

from typing import Any

from apps.academic.domain.student_class_placement import StudentClassPlacement
from apps.academic.services.active_academic_year import get_active_academic_year
from apps.ogrenci.domain.models import OgrenciKayit
from apps.ogrenci.interfaces.list_helpers import (
    EXPORT_COLUMN_TYPES as OGRENCI_EXPORT_COLUMN_TYPES,
    EXPORT_COLUMNS as OGRENCI_EXPORT_COLUMNS,
    build_kalem_ozet_map,
    format_export_row,
    get_varsayilan_veli,
    serialize_kayit_row,
)
from apps.sinif.application.placement_helpers import placement_counts_for_term
from apps.sinif.domain.models import Sinif

MAX_ROSTER_EXPORT_ROWS = 10000

# Öğrenci listesi export sütunları + sınıf listesine özel alanlar
ROSTER_ONLY_COLUMNS = {
    'sira': 'Sıra',
    'alan': 'Alan',
}

ROSTER_EXPORT_COLUMNS = {
    **ROSTER_ONLY_COLUMNS,
    **OGRENCI_EXPORT_COLUMNS,
}

ROSTER_COLUMN_TYPES = {
    'sira': 'integer',
    **OGRENCI_EXPORT_COLUMN_TYPES,
}

DEFAULT_ROSTER_KEYS = [
    'sira',
    'tam_ad',
    'okul_no',
    'alan',
    'telefon',
    'veli_ad_soyad',
    'veli_tc_kimlik_no',
    'veli_telefon',
]

# Tek tabloda sınıf sütunları (geriye dönük / düz liste)
FULL_ROSTER_KEYS = list(ROSTER_EXPORT_COLUMNS.keys())


def parse_roster_column_keys(raw: str | None) -> list[str]:
    if not raw:
        return list(DEFAULT_ROSTER_KEYS)
    keys = [k.strip() for k in raw.split(',') if k.strip()]
    return [k for k in keys if k in ROSTER_EXPORT_COLUMNS] or list(DEFAULT_ROSTER_KEYS)


def parse_sinif_ids_param(raw: str | None) -> list[int]:
    if not raw:
        return []
    ids: list[int] = []
    for part in raw.split(','):
        part = part.strip()
        if not part:
            continue
        try:
            ids.append(int(part))
        except ValueError:
            continue
    return ids


def resolve_siniflar_for_scope(
    *,
    kurum_id: int,
    sube_id: int,
    scope: str,
    egitim_yili_id: int | None = None,
    sinif_id: int | None = None,
    sinif_seviyesi_id: int | None = None,
    sinif_ids: list[int] | None = None,
) -> list[Sinif]:
    qs = Sinif.objects.filter(
        kurum_id=kurum_id,
        sube_id=sube_id,
        aktif_mi=True,
    ).select_related('sinif_seviyesi', 'sube', 'egitim_yili')

    if egitim_yili_id:
        qs = qs.filter(egitim_yili_id=egitim_yili_id)

    if scope == 'sinif':
        if not sinif_id:
            raise ValueError('sinif_id zorunludur')
        qs = qs.filter(id=sinif_id)
    elif scope == 'seviye':
        if not sinif_seviyesi_id:
            raise ValueError('sinif_seviyesi_id zorunludur')
        qs = qs.filter(sinif_seviyesi_id=sinif_seviyesi_id)
    elif scope == 'custom':
        if not sinif_ids:
            raise ValueError('sinif_ids zorunludur')
        qs = qs.filter(id__in=sinif_ids)
    elif scope != 'all':
        raise ValueError('Geçersiz scope')

    return list(qs.order_by('sinif_seviyesi__sira', 'sinif_seviyesi__ad', 'ad'))


def build_roster_groups(siniflar: list[Sinif], term_id: int) -> list[dict[str, Any]]:
    """Sınıf bazlı gruplar — yalnızca dönem yerleşimi olan öğrenciler."""
    active_year = get_active_academic_year()
    groups: list[dict[str, Any]] = []

    for sinif in siniflar:
        placements = (
            StudentClassPlacement.objects.filter(
                academic_year=active_year,
                term_id=term_id,
                classroom_id=sinif.id,
                is_active=True,
            )
            .select_related('student')
            .order_by('student__ad', 'student__soyad')
        )

        student_ids = [p.student_id for p in placements]
        kayit_map: dict[int, OgrenciKayit] = {}
        if student_ids:
            for kayit in OgrenciKayit.objects.filter(
                ogrenci_id__in=student_ids,
                egitim_yili_id=sinif.egitim_yili_id,
                aktif_mi=True,
            ).select_related(
                'alan',
                'school',
                'sube',
                'egitim_yili',
                'sinif',
                'sinif__sinif_seviyesi',
                'sinif_seviyesi',
                'ogrenci',
            ):
                kayit_map[kayit.ogrenci_id] = kayit

        kayit_list = list(kayit_map.values())
        kalem_ozet_map = build_kalem_ozet_map(kayit_list) if kayit_list else {}

        rows: list[dict[str, Any]] = []
        for idx, placement in enumerate(placements, start=1):
            student = placement.student
            kayit = kayit_map.get(student.id)
            if kayit:
                row = serialize_kayit_row(
                    kayit,
                    include_egitim_yili=True,
                    kalem_ozet=kalem_ozet_map.get(kayit.id, ''),
                )
            else:
                veli = get_varsayilan_veli(student)
                row = {
                    'tam_ad': student.tam_ad,
                    'ad': student.ad,
                    'soyad': student.soyad,
                    'tc_kimlik_no': student.tc_kimlik_no or '',
                    'okul_no': '',
                    'telefon': student.telefon or '',
                    'email': student.email or '',
                    'cinsiyet': student.cinsiyet or '',
                    'dogum_tarihi': '',
                    'veli_ad_soyad': (
                        f'{veli.ad} {veli.soyad}'.strip() if veli else (student.veli_ad_soyad or '')
                    ),
                    'veli_tc_kimlik_no': (veli.tc_kimlik_no or '') if veli else '',
                    'veli_telefon': (
                        (veli.telefon or student.veli_telefon or '') if veli else (student.veli_telefon or '')
                    ),
                    'veli_yakinlik_display': '',
                    'sube_ad': sinif.sube.ad if sinif.sube_id and sinif.sube else '',
                    'kayit_tarihi': '',
                    'giris_turu_display': '',
                    'aktif_mi': student.aktif_mi,
                    'egitim_yili': '',
                    'kalem_ozet': '',
                    'geldigi_okul': '',
                }

            row['sira'] = idx
            row['sinif_ad'] = sinif.ad
            row['sinif_seviyesi'] = sinif.sinif_seviyesi.ad if sinif.sinif_seviyesi else ''
            row['alan'] = kayit.alan.ad if kayit and kayit.alan_id and kayit.alan else ''
            rows.append(row)

        groups.append({
            'sinif_id': sinif.id,
            'sinif_ad': sinif.ad,
            'sinif_seviyesi': sinif.sinif_seviyesi.ad if sinif.sinif_seviyesi else '',
            'ogrenci_sayisi': len(rows),
            'rows': rows,
        })

    return groups


def columns_for_grouped_export(column_keys: list[str]) -> list[str]:
    """Sınıf başlığı ayrı olduğu için tablo sütunlarından sınıf/seviye çıkar."""
    keys = [k for k in column_keys if k in ROSTER_EXPORT_COLUMNS]
    per_class = [k for k in keys if k not in ('sinif_ad', 'sinif_seviyesi')]
    return per_class or list(DEFAULT_ROSTER_KEYS)


def group_section_title(group: dict[str, Any]) -> str:
    seviye = group.get('sinif_seviyesi') or ''
    count = group.get('ogrenci_sayisi', 0)
    if seviye:
        return f"{group['sinif_ad']} — {seviye} ({count} öğrenci)"
    return f"{group['sinif_ad']} ({count} öğrenci)"


def _group_rows_for_export(group: dict[str, Any], column_keys: list[str]) -> list[dict[str, object]]:
    keys = columns_for_grouped_export(column_keys)
    return [format_export_row(row, keys) for row in group.get('rows', [])]


def flatten_roster_groups(groups: list[dict[str, Any]], column_keys: list[str]) -> list[dict[str, Any]]:
    flat: list[dict[str, Any]] = []
    for group in groups:
        for row in group['rows']:
            flat.append({k: row.get(k, '') for k in column_keys})
            if len(flat) >= MAX_ROSTER_EXPORT_ROWS:
                return flat
    return flat


def roster_export_cell(row: dict[str, Any], key: str) -> object:
    return row.get(key, '')


def build_roster_export_rows(groups: list[dict[str, Any]], column_keys: list[str]) -> list[dict[str, object]]:
    keys = [k for k in column_keys if k in ROSTER_EXPORT_COLUMNS]
    return [
        {k: roster_export_cell(row, k) for k in keys}
        for row in flatten_roster_groups(groups, keys)
    ]


def build_roster_export_columns(column_keys: list[str]):
    from shared.export.style_manager import ExportColumn

    return [
        ExportColumn(
            key=k,
            label=ROSTER_EXPORT_COLUMNS[k],
            type=ROSTER_COLUMN_TYPES.get(k, 'text'),
        )
        for k in column_keys
        if k in ROSTER_EXPORT_COLUMNS
    ]


def build_roster_export_meta(
    request,
    *,
    kurum_id: int | None,
    sube_id: int | None,
    egitim_yili,
    term_name: str = '',
    scope_label: str = '',
    report_title: str = 'SINIF ÖĞRENCİ LİSTESİ',
):
    from apps.sinif.application.export import build_export_meta

    meta = build_export_meta(
        request,
        kurum_id=kurum_id,
        sube_id=sube_id,
        egitim_yili=egitim_yili,
        report_title=report_title,
    )
    if term_name:
        meta.extra['donem'] = term_name
    if scope_label:
        meta.extra['filter_ozet'] = scope_label
    return meta


def build_roster_export_stats(groups: list[dict[str, Any]]):
    from shared.export.style_manager import ExportStat

    toplam_sinif = len(groups)
    dolu_sinif = sum(1 for g in groups if g['ogrenci_sayisi'] > 0)
    toplam_ogrenci = sum(g['ogrenci_sayisi'] for g in groups)
    return [
        ExportStat(label='Toplam Sınıf', value=toplam_sinif, type='integer'),
        ExportStat(label='Öğrencili Sınıf', value=dolu_sinif, type='integer'),
        ExportStat(label='Toplam Öğrenci', value=toplam_ogrenci, type='integer'),
    ]


def scope_display_label(scope: str, *, sinif_ad: str = '', seviye_ad: str = '') -> str:
    if scope == 'sinif' and sinif_ad:
        return f'Sınıf: {sinif_ad}'
    if scope == 'seviye' and seviye_ad:
        return f'Seviye: {seviye_ad}'
    if scope == 'custom':
        return 'Seçili sınıflar'
    return 'Tüm sınıflar'


def mevcutluk_map_for_siniflar(siniflar: list[Sinif], term_id: int | None) -> dict[int, int]:
    if not term_id or not siniflar:
        return {}
    return placement_counts_for_term(term_id, [s.id for s in siniflar])


def build_grouped_csv_text(
    groups: list[dict[str, Any]],
    column_keys: list[str],
    meta,
) -> str:
    import csv
    import io

    from shared.export.csv_export_service import CsvExportService

    per_class_keys = columns_for_grouped_export(column_keys)
    cols = build_roster_export_columns(per_class_keys)
    buf = io.StringIO()
    buf.write('\ufeff')
    writer = csv.writer(buf, delimiter=';', lineterminator='\r\n')
    CsvExportService.write_letterhead_rows(writer, meta)
    if meta.extra.get('donem'):
        writer.writerow([f"Dönem: {meta.extra['donem']}"])
    if meta.extra.get('filter_ozet'):
        writer.writerow([f"Kapsam: {meta.extra['filter_ozet']}"])
    writer.writerow([])

    for group in groups:
        writer.writerow([group_section_title(group)])
        writer.writerow([c.label for c in cols])
        rows = _group_rows_for_export(group, column_keys)
        if rows:
            for row in rows:
                writer.writerow([
                    CsvExportService._csv_cell(row.get(c.key), c, decimal_comma=True)
                    for c in cols
                ])
        else:
            writer.writerow(['(Bu sınıfta yerleşmiş öğrenci yok)'])
        writer.writerow([])

    return buf.getvalue()


def export_grouped_csv_response(
    groups: list[dict[str, Any]],
    column_keys: list[str],
    meta,
    *,
    filename: str = 'sinif_ogrenci_listesi',
):
    from django.http import HttpResponse
    from shared.export import style_manager as sm

    content = build_grouped_csv_text(groups, column_keys, meta)
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{sm.safe_filename(filename)}.csv"'
    return response


def export_grouped_xlsx_response(
    groups: list[dict[str, Any]],
    column_keys: list[str],
    meta,
    stats,
    *,
    filename: str = 'sinif_ogrenci_listesi',
):
    import io

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from shared.export import style_manager as sm
    from shared.export.excel_export_service import ExcelExportService

    per_class_keys = columns_for_grouped_export(column_keys)
    cols = build_roster_export_columns(per_class_keys)

    wb = Workbook()
    ws = wb.active
    ws.title = sm.safe_sheet_title('Sınıf Listeleri')

    header_row = ExcelExportService._write_letterhead(ws, meta, num_cols=len(cols))
    current_row = ExcelExportService._write_stats(ws, stats, start_row=header_row, num_cols=len(cols))

    merged_col_widths = [c.width or 12 for c in cols]
    merged_needs_wrap = [False] * len(cols)
    last_row = current_row

    for group in groups:
        title = group_section_title(group)
        title_cell = ws.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(name=sm.FONT_NAME, size=12, bold=True, color=sm.BRAND_PRIMARY_HEX)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        if len(cols) > 1:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(cols),
            )
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        rows = _group_rows_for_export(group, column_keys)
        if rows:
            last_row, _, col_widths, col_needs_wrap = ExcelExportService._write_table(
                ws, rows, cols, header_row=current_row, auto_filter=False,
            )
            current_row = last_row + 2
            for i, w in enumerate(col_widths):
                merged_col_widths[i] = max(merged_col_widths[i], w)
                merged_needs_wrap[i] = merged_needs_wrap[i] or col_needs_wrap[i]
        else:
            empty_cell = ws.cell(row=current_row, column=1, value='(Bu sınıfta yerleşmiş öğrenci yok)')
            empty_cell.font = Font(name=sm.FONT_NAME, size=sm.FONT_SIZE, italic=True)
            current_row += 2

    ExcelExportService._apply_page_setup(
        ws,
        orientation='landscape',
        header_row=header_row,
        last_row=max(last_row, current_row - 1),
        last_col=len(cols),
        report_title=meta.report_title,
    )
    ExcelExportService._apply_column_widths(ws, cols, merged_col_widths, merged_needs_wrap)
    ws.auto_filter.ref = None

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{sm.safe_filename(filename)}.xlsx"'
    return response
