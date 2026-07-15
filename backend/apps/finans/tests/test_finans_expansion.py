"""
Finans modülü genişletme testleri — Faz 0–4.
"""
import io
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from apps.egitim_yili.domain.models import EgitimYili

from apps.finans.application.export.export_service import ExportService
from apps.finans.application.export.gun_sonu_detay_export_service import GunSonuDetayExportService
from apps.finans.application.gun_sonu_detay_report_service import GunSonuDetayReportService
from apps.finans.application.overdue_messaging import build_overdue_context
from apps.finans.application.period.period_service import PeriodService
from apps.finans.application.reports.report_service import ReportService
from apps.finans.constants.payment_types import OdemeYontemiTipi
from apps.finans.domain.payment_method import OdemeYontemi
from apps.finans.domain.financial_account import MaliHesap
from apps.kurum.domain.models import Kurum
from apps.odeme_takip.application.services.tahsilat_service import TahsilatService
from apps.odeme_takip.domain.cek_senet import CekSenetDetay, CekSenetDurum
from apps.odeme_takip.domain.enums import SozlesmeDurum, TaksitDurum, TahsilatDurum, TahsilatTuru
from apps.odeme_takip.domain.models import Sozlesme, Taksit, Tahsilat
from apps.ogrenci.domain.models import Ogrenci, OgrenciVeli
from apps.roller.models import Permission, Role, RolePermission, UserRole
from apps.sube.domain.models import Sube

User = get_user_model()


def _assign_perms(user, *codes):
    role, _ = Role.objects.get_or_create(
        code='finans_exp_test',
        defaults={'name': 'Finans Exp Test', 'level': 100, 'is_system_role': True},
    )
    for code in codes:
        perm, _ = Permission.objects.get_or_create(
            code=code,
            defaults={'name': code, 'module': code.split('.')[0], 'permission_type': 'write'},
        )
        RolePermission.objects.get_or_create(role=role, permission=perm)
    UserRole.objects.update_or_create(user=user, defaults={'role': role})


class FinansExpansionTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.kurum = Kurum.objects.create(ad='Fin Exp Kurum', kod='FEXP')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='FEXP')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.user = User.objects.create_user(username='finexp', password='test')
        _assign_perms(
            self.user,
            'finans.read',
            'finans.manage',
            'communication.write',
        )
        self.client.force_authenticate(user=self.user)

        self.ogrenci = Ogrenci.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            ad='Ayşe',
            soyad='Test',
            aktif_mi=True,
        )
        self.veli = OgrenciVeli.objects.create(
            ogrenci=self.ogrenci,
            veli_turu='anne',
            ad='Veli',
            soyad='Test',
            telefon='05321112233',
            sms_bildirimleri=['odeme', 'duyuru'],
        )
        self.mali_hesap = MaliHesap.objects.create(
            sube=self.sube,
            ad='Merkez Kasa',
        )
        self.odeme_yontemi = OdemeYontemi.objects.create(
            mali_hesap=self.mali_hesap,
            kurum=self.kurum,
            ad='Nakit',
            tip=OdemeYontemiTipi.NAKIT,
            komisyon_orani=Decimal('0'),
        )
        self.cek_yontemi = OdemeYontemi.objects.create(
            mali_hesap=self.mali_hesap,
            kurum=self.kurum,
            ad='Çek',
            tip=OdemeYontemiTipi.CEK,
            komisyon_orani=Decimal('0'),
        )
        self.sozlesme = Sozlesme.objects.create(
            sozlesme_no='SZ-FEXP-001',
            ogrenci=self.ogrenci,
            egitim_yili=self.ey,
            kurum=self.kurum,
            sube=self.sube,
            veli=self.veli,
            baslangic_tarihi=timezone.localdate(),
            bitis_tarihi=timezone.localdate() + timedelta(days=365),
            brut_tutar=10000,
            net_tutar=10000,
            durum=SozlesmeDurum.AKTIF,
        )
        self.overdue_taksit = Taksit.objects.create(
            sozlesme=self.sozlesme,
            taksit_no=1,
            vade_tarihi=timezone.localdate() - timedelta(days=7),
            tutar=5000,
            odenen_tutar=0,
            kalan_tutar=5000,
            durum=TaksitDurum.BEKLEMEDE,
        )
        self.future_taksit = Taksit.objects.create(
            sozlesme=self.sozlesme,
            taksit_no=2,
            vade_tarihi=timezone.localdate() + timedelta(days=30),
            tutar=5000,
            kalan_tutar=5000,
            durum=TaksitDurum.BEKLEMEDE,
        )

    def test_odeme_yontemi_tipi_cek_senet(self):
        self.assertIn(OdemeYontemiTipi.CEK, OdemeYontemiTipi.get_values())
        self.assertIn(OdemeYontemiTipi.SENET, OdemeYontemiTipi.get_values())

    def test_overdue_payments_api(self):
        response = self.client.get(
            '/finans/api/overdue-payments/',
            {'kurum_id': self.kurum.id, 'sube_id': self.sube.id},
        )
        self.assertEqual(response.status_code, 200)
        ozet = response.data['ozet']
        self.assertEqual(ozet['toplam_taksit_sayisi'], 1)
        self.assertEqual(ozet['toplam_kalan_tutar'], 5000)
        self.assertEqual(ozet['kisi_sayisi'], 1)
        self.assertIn('ortalama_gecikme_gun', ozet)
        self.assertEqual(len(response.data['results']), 1)
        self.assertIn('page', response.data)
        self.assertIn('total_pages', response.data)
        item = response.data['results'][0]
        self.assertEqual(item['ogrenci_id'], self.ogrenci.id)
        self.assertEqual(item['veli_adi'], self.veli.tam_ad)
        self.assertEqual(item['veli_telefon'], self.veli.telefon)
        self.assertEqual(item['toplam_gecikmis_tutar'], 5000)
        self.assertIn('already_sent_24h', item)

    def test_overdue_payments_kurum_from_header(self):
        response = self.client.get(
            '/finans/api/overdue-payments/',
            HTTP_X_KURUM_ID=str(self.kurum.id),
            HTTP_X_SUBE_ID=str(self.sube.id),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['ozet']['toplam_taksit_sayisi'], 1)

    def test_overdue_payments_min_tutar_filter(self):
        response = self.client.get(
            '/finans/api/overdue-payments/',
            {'kurum_id': self.kurum.id, 'sube_id': self.sube.id, 'min_tutar': 6000},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 0)

    def test_overdue_payments_csv_export(self):
        response = self.client.get(
            '/finans/api/overdue-payments/',
            {'kurum_id': self.kurum.id, 'sube_id': self.sube.id, 'format': 'csv'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])
        content = response.content.decode('utf-8-sig')
        self.assertTrue(content.startswith('\ufeff') or 'Sözleşme No' in content)
        self.assertIn(';', content)

    def test_overdue_payments_csv_export_respects_columns(self):
        response = self.client.get(
            '/finans/api/overdue-payments/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'format': 'csv',
                'columns': 'ogrenci_adi,kalan_tutar',
            },
        )
        self.assertEqual(response.status_code, 200)
        content = response.content.decode('utf-8-sig')
        self.assertIn('Öğrenci', content)
        self.assertIn('Kalan Borç', content)
        self.assertNotIn('Sözleşme No', content)
        self.assertNotIn('Veli', content)

    def test_period_summary_alinan(self):
        Tahsilat.objects.create(
            sozlesme=self.sozlesme,
            taksit=self.overdue_taksit,
            odeme_yontemi=self.odeme_yontemi,
            tutar=2000,
            tahsilat_tarihi=timezone.localdate(),
            durum=TahsilatDurum.AKTIF,
            tahsilat_turu=TahsilatTuru.NORMAL,
        )
        baslangic = (timezone.localdate() - timedelta(days=1)).isoformat()
        bitis = (timezone.localdate() + timedelta(days=1)).isoformat()
        response = self.client.get(
            '/finans/api/period-summary/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'baslangic': baslangic,
                'bitis': bitis,
                'mode': 'alinan',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['mode'], 'alinan')
        ozet = response.data['ozet']
        self.assertGreaterEqual(ozet['toplam_tutar'], 2000)
        self.assertIn('yontem_dagilimi', ozet)
        self.assertIn('kaynak_kirilimi', ozet)
        self.assertIn('grafik', ozet)

    def test_period_summary_beklenen(self):
        baslangic = timezone.localdate().isoformat()
        bitis = (timezone.localdate() + timedelta(days=60)).isoformat()
        data = PeriodService.period_summary(
            kurum_id=self.kurum.id,
            baslangic=date.fromisoformat(baslangic),
            bitis=date.fromisoformat(bitis),
            mode='beklenen',
        )
        self.assertEqual(data['mode'], 'beklenen')
        self.assertGreaterEqual(data['ozet']['toplam_tutar'], 5000)
        self.assertIn('kaynak_kirilimi', data['ozet'])

    def test_period_details_api(self):
        Tahsilat.objects.create(
            sozlesme=self.sozlesme,
            odeme_yontemi=self.odeme_yontemi,
            tutar=1500,
            tahsilat_tarihi=timezone.localdate(),
            durum=TahsilatDurum.AKTIF,
        )
        response = self.client.get(
            '/finans/api/period-details/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'baslangic': timezone.localdate().isoformat(),
                'bitis': timezone.localdate().isoformat(),
                'kaynak': 'sozlesme',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.data['count'], 1)
        self.assertIn('total_pages', response.data)
        self.assertIn('mode', response.data)
        item = response.data['results'][0]
        self.assertIn('kisi_adi', item)
        self.assertIn('kaynak_label', item)

    def test_overdue_reminder_preview(self):
        response = self.client.post(
            '/finans/api/overdue-reminders/preview/',
            {
                'kurum_id': self.kurum.id,
                'taksit_ids': [self.overdue_taksit.id],
                'template': 'Sayın {veli_ad}, {ogrenci_ad}',
            },
            format='json',
            HTTP_X_SUBE_ID=str(self.sube.id),
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['sendable_count'], 1)
        self.assertIn('recipients', response.data)
        self.assertIn('template', response.data)
        recipient = response.data['recipients'][0]
        self.assertIn('rendered_body', recipient)
        self.assertEqual(recipient['veli_id'], self.veli.id)
        self.assertIn('taksit_ids', recipient)
        self.assertIn(self.overdue_taksit.id, recipient['taksit_ids'])
        self.assertIn('already_sent_24h', recipient)
        self.assertIn('available_veliler', recipient)

    def test_build_overdue_context_variables(self):
        ctx = build_overdue_context(self.overdue_taksit, toplam_gecikmis=5000)
        self.assertIn('veli_ad', ctx)
        self.assertIn('gecikme_gunu', ctx)
        self.assertIn('toplam_gecikmis_tutar', ctx)

    def test_tahsilat_cek_senet_detay(self):
        service = TahsilatService()
        tahsilat, errors = service.create({
            'sozlesme_id': self.sozlesme.id,
            'taksit_id': self.overdue_taksit.id,
            'odeme_yontemi_id': self.cek_yontemi.id,
            'tutar': 1000,
            'tahsilat_tarihi': timezone.localdate().isoformat(),
            'cek_senet_detay': {
                'cek_senet_no': 'CHK-EXP-001',
                'banka_adi': 'Ziraat',
                'vade_tarihi': (timezone.localdate() + timedelta(days=14)).isoformat(),
                'durum': CekSenetDurum.PORTFOYDE,
            },
        }, user=self.user)
        self.assertIsNone(errors)
        self.assertTrue(hasattr(tahsilat, 'cek_senet_detay'))
        self.assertEqual(tahsilat.cek_senet_detay.cek_senet_no, 'CHK-EXP-001')

    def test_cek_senet_report(self):
        CekSenetDetay.objects.create(
            taksit=self.overdue_taksit,
            cek_senet_no='CHK-001',
            banka_adi='Ziraat',
            vade_tarihi=timezone.localdate() + timedelta(days=10),
            durum=CekSenetDurum.PORTFOYDE,
        )
        response = self.client.get(
            '/finans/api/reports/cek-senet-listesi/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'baslangic': timezone.localdate().isoformat(),
                'bitis': (timezone.localdate() + timedelta(days=30)).isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(len(response.data['rows']), 1)
        self.assertIn('columns', response.data)
        self.assertGreater(len(response.data['columns']), 0)

    def test_report_json_includes_columns(self):
        response = self.client.get(
            '/finans/api/reports/gunluk-satis/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'baslangic': timezone.localdate().isoformat(),
                'bitis': timezone.localdate().isoformat(),
                'format': 'json',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('columns', response.data)
        self.assertEqual(response.data['columns'][0]['key'], 'tarih')
        self.assertIn('rows', response.data)
        self.assertIn('title', response.data)

    def test_gunluk_satis_report(self):
        Tahsilat.objects.create(
            sozlesme=self.sozlesme,
            odeme_yontemi=self.odeme_yontemi,
            tutar=1000,
            tahsilat_tarihi=timezone.localdate(),
            durum=TahsilatDurum.AKTIF,
        )
        report = ReportService.run(
            'gunluk-satis',
            kurum_id=self.kurum.id,
            params={
                'baslangic': timezone.localdate().isoformat(),
                'bitis': timezone.localdate().isoformat(),
            },
        )
        self.assertGreaterEqual(report['toplam'], 1000)

    def test_tahsilat_analiz_report(self):
        response = self.client.get(
            '/finans/api/reports/tahsilat-analiz/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'baslangic': (timezone.localdate() - timedelta(days=30)).isoformat(),
                'bitis': timezone.localdate().isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('rows', response.data)

    def test_export_service_csv_bom_semicolon(self):
        rows = [{'ad': 'Test', 'tutar': 100}]
        columns = [{'key': 'ad', 'label': 'Ad'}, {'key': 'tutar', 'label': 'Tutar'}]
        response = ExportService.build('csv', rows, columns, title='Test')
        text = response.content.decode('utf-8-sig')
        self.assertIn('Ad;Tutar', text)
        self.assertIn('Test', text)
        self.assertTrue(response.content.startswith(b'\xef\xbb\xbf'))

    def test_export_service_xlsx(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest('openpyxl not installed')
        rows = [{'ad': 'X', 'tutar': 50}]
        columns = [{'key': 'ad', 'label': 'Ad'}, {'key': 'tutar', 'label': 'Tutar'}]
        response = ExportService.build('xlsx', rows, columns, title='XLSX Test')
        self.assertIn('spreadsheetml', response['Content-Type'])

    def test_export_service_xlsx_landscape_orientation(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest('openpyxl not installed')
        rows = [{'ad': 'X', 'tutar': 50}]
        columns = [{'key': 'ad', 'label': 'Ad'}, {'key': 'tutar', 'label': 'Tutar'}]
        response = ExportService.build(
            'xlsx', rows, columns, title='Landscape Test', orientation='landscape',
        )
        wb = load_workbook(io.BytesIO(response.content))
        self.assertEqual(wb.active.page_setup.orientation, 'landscape')

    def test_export_service_pdf(self):
        rows = [{'ad': 'PDF', 'tutar': 1}]
        columns = [{'key': 'ad', 'label': 'Ad'}, {'key': 'tutar', 'label': 'Tutar'}]
        response = ExportService.build('pdf', rows, columns, title='PDF Test')
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_dashboard_overview(self):
        Tahsilat.objects.create(
            sozlesme=self.sozlesme,
            odeme_yontemi=self.odeme_yontemi,
            tutar=500,
            tahsilat_tarihi=timezone.localdate(),
            durum=TahsilatDurum.AKTIF,
        )
        response = self.client.get(
            '/finans/api/dashboard/overview/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'egitim_yili_id': self.ey.id,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('ozet_kartlar', response.data)
        self.assertIn('bugunku_islemler', response.data)
        self.assertIn('gunluk_gelir_gider_net', response.data)
        self.assertGreaterEqual(response.data['ozet_kartlar']['bugun_alinan'], 500)

    def test_dashboard_overview_gider_null_egitim_yili(self):
        """Gider kayıtları egitim_yili boş olsa bile kurum bağlamında görünmeli."""
        from apps.finans.domain.gider_kaydi import GiderKaydi
        from apps.finans.domain.gider_kategorisi import GiderKategorisi
        from apps.finans.domain.gider_odeme import GiderOdeme
        from apps.finans.constants.gider_types import GiderDurum, OdemeDurum
        from apps.finans.domain.cari_hesap import CariHesap

        cari = CariHesap.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            unvan='Dash Gider Cari',
            hesap_turu='tedarikci',
        )
        kat = GiderKategorisi.objects.create(
            kurum=self.kurum, sube=self.sube, ad='Dash Kat',
        )
        gider = GiderKaydi.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            cari_hesap=cari,
            gider_kategorisi=kat,
            fatura_tarihi=timezone.localdate(),
            vade_tarihi=timezone.localdate(),
            brut_tutar=1000,
            kdv_orani=0,
            kdv_tutar=0,
            net_tutar=1000,
            durum=GiderDurum.ONAYLANDI,
        )
        GiderOdeme.objects.create(
            gider_kaydi=gider,
            tutar=1000,
            odeme_tarihi=timezone.localdate(),
            durum=OdemeDurum.TAMAMLANDI,
        )
        response = self.client.get(
            '/finans/api/dashboard/overview/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'egitim_yili_id': self.ey.id,
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.data['ozet_kartlar']['bu_ay_gider'], 1000)
        self.assertGreaterEqual(len(response.data['son_giderler']), 1)
        self.assertTrue(
            any(r.get('kaynak') == 'gider' for r in response.data['bugunku_islemler']),
        )

    def test_dashboard_bugun_counts_kayit_gunu(self):
        """İşlem tarihi dün, kayıt bugün — bugün alınan kartında görünmeli."""
        from datetime import timedelta

        yesterday = timezone.localdate() - timedelta(days=1)
        Tahsilat.objects.create(
            sozlesme=self.sozlesme,
            odeme_yontemi=self.odeme_yontemi,
            tutar=2500,
            tahsilat_tarihi=yesterday,
            durum=TahsilatDurum.AKTIF,
        )
        response = self.client.get(
            '/finans/api/dashboard/overview/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'egitim_yili_id': self.ey.id,
                'referans_tarih': timezone.localdate().isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.data['ozet_kartlar']['bugun_alinan'], 2500)

    def test_dashboard_bugun_gider_partial_taksit_odeme(self):
        """Kısmi taksit ödemesi bugün kaydedildiğinde kart ve grafik tutarlı olmalı."""
        from datetime import timedelta
        from apps.finans.domain.gider_kaydi import GiderKaydi
        from apps.finans.domain.gider_kategorisi import GiderKategorisi
        from apps.finans.domain.gider_taksit import GiderTaksit
        from apps.finans.domain.gider_odeme import GiderOdeme
        from apps.finans.constants.gider_types import GiderDurum, GiderTaksitDurum, OdemeDurum
        from apps.finans.domain.cari_hesap import CariHesap

        bugun = timezone.localdate()
        dun = bugun - timedelta(days=1)

        cari = CariHesap.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            unvan='Taksit Gider Cari',
            hesap_turu='tedarikci',
        )
        kat = GiderKategorisi.objects.create(
            kurum=self.kurum, sube=self.sube, ad='Taksit Kat',
        )
        gider = GiderKaydi.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            cari_hesap=cari,
            gider_kategorisi=kat,
            fatura_tarihi=bugun,
            vade_tarihi=bugun,
            brut_tutar=3000,
            kdv_orani=0,
            kdv_tutar=0,
            net_tutar=3000,
            durum=GiderDurum.ONAYLANDI,
        )
        taksit = GiderTaksit.objects.create(
            gider_kaydi=gider,
            taksit_no=1,
            vade_tarihi=bugun,
            tutar=1500,
            odenen_tutar=1500,
            durum=GiderTaksitDurum.ODENDI,
        )
        GiderTaksit.objects.create(
            gider_kaydi=gider,
            taksit_no=2,
            vade_tarihi=bugun + timedelta(days=30),
            tutar=1500,
            odenen_tutar=0,
            durum=GiderTaksitDurum.BEKLEMEDE,
        )
        # Ödeme tarihi dün seçilmiş olsa bile kayıt bugün — bugün gider sayılmalı
        GiderOdeme.objects.create(
            gider_kaydi=gider,
            gider_taksit=taksit,
            tutar=1500,
            odeme_tarihi=dun,
            durum=OdemeDurum.TAMAMLANDI,
        )

        response = self.client.get(
            '/finans/api/dashboard/overview/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'egitim_yili_id': self.ey.id,
                'referans_tarih': bugun.isoformat(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(response.data['ozet_kartlar']['bugun_gider'], 1500)

        chart_today = next(
            (d for d in response.data['gunluk_gelir_gider_net'] if d['tarih'] == bugun.isoformat()),
            None,
        )
        self.assertIsNotNone(chart_today)
        self.assertGreaterEqual(chart_today['gider'], 1500)

        gider_kayit_rows = [
            r for r in response.data['bugunku_islemler']
            if r.get('kaynak') == 'gider_kayit' and r.get('gider_id') == gider.id
        ]
        self.assertEqual(len(gider_kayit_rows), 0, 'Ödeme yapılmış gider kaydı çift görünmemeli')

        odeme_rows = [
            r for r in response.data['bugunku_islemler']
            if r.get('kaynak') == 'gider' and r.get('gider_id') == gider.id
        ]
        self.assertEqual(len(odeme_rows), 1)
        self.assertEqual(odeme_rows[0]['tutar'], 1500)

    def test_tam_odeme_sonrasi_acik_taksit_satirlari_kapanir(self):
        """Taksitsiz tam ödeme sonrası taksit satırları gider durumu ile hizalanmalı."""
        from apps.finans.domain.cari_hesap import CariHesap
        from apps.finans.domain.gider_kategorisi import GiderKategorisi
        from apps.finans.domain.gider_kaydi import GiderKaydi
        from apps.finans.domain.gider_taksit import GiderTaksit
        from apps.finans.domain.gider_odeme import GiderOdeme
        from apps.finans.constants.gider_types import GiderDurum, GiderTaksitDurum, OdemeDurum
        from apps.finans.application.gider_service import GiderService
        from apps.finans.application.selectors.gider_selector import GiderSelector

        cari = CariHesap.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            unvan='Sync Test Cari',
            hesap_turu='tedarikci',
        )
        kat = GiderKategorisi.objects.create(kurum=self.kurum, sube=self.sube, ad='Sync Kat')
        gider = GiderKaydi.objects.create(
            kurum=self.kurum,
            sube=self.sube,
            cari_hesap=cari,
            gider_kategorisi=kat,
            fatura_tarihi=timezone.localdate(),
            vade_tarihi=timezone.localdate(),
            brut_tutar=5000,
            kdv_orani=0,
            kdv_tutar=0,
            net_tutar=5000,
            durum=GiderDurum.ODENDI,
            odenen_toplam=5000,
        )
        taksit = GiderTaksit.objects.create(
            gider_kaydi=gider,
            taksit_no=1,
            vade_tarihi=timezone.localdate(),
            tutar=5000,
            odenen_tutar=0,
            durum=GiderTaksitDurum.BEKLEMEDE,
        )
        GiderOdeme.objects.create(
            gider_kaydi=gider,
            tutar=5000,
            odeme_tarihi=timezone.localdate(),
            odeme_yontemi=self.odeme_yontemi,
            mali_hesap=self.mali_hesap,
            durum=OdemeDurum.TAMAMLANDI,
        )

        GiderService().repair_inconsistent_taksit_rows(self.kurum.id, sube_id=self.sube.id)
        taksit.refresh_from_db()
        self.assertEqual(taksit.durum, GiderTaksitDurum.ODENDI)
        self.assertEqual(taksit.odenen_tutar, Decimal('5000'))

        listed = list(GiderSelector().yaklasan_vadeler(self.kurum.id, gun=14, sube_id=self.sube.id))
        self.assertFalse(any(row.id == taksit.id for row in listed))

    def test_gun_sonu_detay_api(self):
        Tahsilat.objects.create(
            sozlesme=self.sozlesme,
            odeme_yontemi=self.odeme_yontemi,
            tutar=1500,
            tahsilat_tarihi=timezone.localdate(),
            durum=TahsilatDurum.AKTIF,
            islem_yapan=self.user,
        )
        response = self.client.get(
            '/finans/api/gun-sonu/',
            {
                'kurum_id': self.kurum.id,
                'sube_id': self.sube.id,
                'gun': timezone.localdate().isoformat(),
                'rapor': 'detay',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn('detay_rapor', response.data)
        detay = response.data['detay_rapor']
        for key in (
            'kapak', 'yonetici_ozeti', 'uyarilar', 'gunluk_finans_ozeti', 'tahsilat_ozeti',
            'tahsilat_listesi', 'gelir_hareketleri', 'gider_hareketleri',
            'cari_hareketleri', 'ogrenci_hareketleri', 'iptal_islemleri', 'iade_islemleri',
            'odeme_turu_dagilimi', 'kategori_gelirler', 'kategori_giderler',
            'personel_performans', 'kullanici_islem_detayi', 'kasa_hareketleri',
            'banka_hareketleri', 'pos_hareketleri', 'kasa_ozeti', 'grafikler', 'sistem',
        ):
            self.assertIn(key, detay)
        self.assertEqual(len(detay['tahsilat_listesi']), 1)
        self.assertEqual(detay['tahsilat_listesi'][0]['tutar'], 1500)
        self.assertIn('sozlesme_no', detay['tahsilat_listesi'][0])
        self.assertIn('nakit_tahsilatlar', detay['kasa_ozeti'])
        self.assertNotIn('filtreler', detay['sistem'])

    def test_gun_sonu_detay_xlsx_export(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest('openpyxl not installed')

        report = GunSonuDetayReportService().build_detay_rapor(
            self.kurum.id,
            timezone.localdate(),
            self.sube.id,
            hazirlayan='Test User',
        )
        response = GunSonuDetayExportService.build('xlsx', report, orientation='landscape')
        self.assertIn('spreadsheetml', response['Content-Type'])
        wb = load_workbook(io.BytesIO(response.content))
        expected_sheets = {
            'Özet', 'Tahsilatlar', 'Gelirler', 'Giderler', 'Cari Hareketleri',
            'İptaller', 'İadeler', 'Kullanıcı Özeti', 'Ödeme Türleri',
            'Kategori Analizi', 'Kasa Özeti',
            'Kasa Hareketleri', 'Banka Hareketleri', 'POS Hareketleri',
            'Personel Performans',
        }
        self.assertEqual(set(wb.sheetnames), expected_sheets)
        self.assertEqual(wb['Özet'].page_setup.orientation, 'landscape')

    def test_gun_sonu_detay_csv_export(self):
        report = GunSonuDetayReportService().build_detay_rapor(
            self.kurum.id,
            timezone.localdate(),
            self.sube.id,
        )
        response = GunSonuDetayExportService.build('csv', report)
        self.assertIn('text/csv', response['Content-Type'])
        self.assertIn('.csv', response['Content-Disposition'])
        text = response.content.decode('utf-8-sig')
        self.assertIn('GÜN SONU DETAY RAPORU', text)
        self.assertIn('Özet', text)
        self.assertIn('Tahsilatlar', text)
        self.assertIn('Kasa Özeti', text)
        self.assertTrue(response.content.startswith(b'\xef\xbb\xbf'))


class DashboardMaliHesapBloklariTest(TestCase):
    """Dashboard kasa/banka widget'ı canlı BakiyeHareketi bakiyesini kullanmalı."""

    def setUp(self):
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.hareket_types import HareketKaynagi, HareketYonu
        from apps.finans.domain.bakiye_hareketi import BakiyeHareketi
        from apps.finans.domain.donem_bakiye import DonemBakiye

        self.kurum = Kurum.objects.create(ad='Dash Mali Kurum', kod='DMK')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='DMK')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.kasa = MaliHesap.objects.create(
            sube=self.sube, ad='Nakit', tip=MaliHesapTipi.KASA,
        )
        self.banka_eski = MaliHesap.objects.create(
            sube=self.sube, ad='Ziraat Bankası (İ.K)', tip=MaliHesapTipi.BANKA, banka='ziraat',
        )
        self.banka_yeni = MaliHesap.objects.create(
            sube=self.sube, ad='Vakıfbank', tip=MaliHesapTipi.BANKA, banka='vakifbank',
        )
        today = timezone.localdate()
        BakiyeHareketi.objects.create(
            mali_hesap=self.kasa,
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.ey,
            yon=HareketYonu.GIRIS,
            tutar=249_031,
            kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0,
            bakiye_sonrasi=249_031,
            islem_tarihi=today,
        )
        BakiyeHareketi.objects.create(
            mali_hesap=self.banka_yeni,
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.ey,
            yon=HareketYonu.GIRIS,
            tutar=252_000,
            kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0,
            bakiye_sonrasi=252_000,
            islem_tarihi=today,
        )
        # Dönem satırı var ama canlı bakiye 0 — eski dashboard burada 78.500 gösteriyordu
        DonemBakiye.objects.create(
            mali_hesap=self.banka_eski,
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.ey,
            donem_basi_bakiye=78_500,
            toplam_gelir=0,
            toplam_gider=0,
            donem_sonu_bakiye=78_500,
        )
        # Vakıfbank için dönem satırı yok — eski dashboard bu hesabı hiç listelemiyordu

    def test_dashboard_uses_live_balances_not_stale_donem_rows(self):
        from apps.finans.application.dashboard_overview_service import _mali_hesap_bloklari

        kasa, banka, kasa_toplam, banka_toplam = _mali_hesap_bloklari(
            self.kurum.id, self.sube.id, self.ey.id,
        )
        kasa_map = {h['mali_hesap_ad']: h['donem_sonu_bakiye'] for h in kasa}
        banka_map = {h['mali_hesap_ad']: h['donem_sonu_bakiye'] for h in banka}

        self.assertEqual(kasa_map['Nakit'], 249_031)
        self.assertEqual(kasa_toplam, 249_031)
        self.assertEqual(banka_map['Ziraat Bankası (İ.K)'], 0)
        self.assertEqual(banka_map['Vakıfbank'], 252_000)
        self.assertEqual(banka_toplam, 252_000)
        self.assertIn('Vakıfbank', banka_map)


class DonemBakiyeSenkronizeTest(TestCase):
    """DonemBakiye okuma öncesi hareketlerden self-heal senkronizasyonu."""

    def setUp(self):
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.hareket_types import HareketKaynagi, HareketYonu
        from apps.finans.domain.bakiye_hareketi import BakiyeHareketi
        from apps.finans.domain.donem_bakiye import DonemBakiye

        self.kurum = Kurum.objects.create(ad='Donem Sync Kurum', kod='DSK')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='DSK')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.kasa = MaliHesap.objects.create(
            sube=self.sube, ad='Nakit', tip=MaliHesapTipi.KASA,
        )
        self.banka_eski = MaliHesap.objects.create(
            sube=self.sube, ad='Ziraat Bankası (İ.K)', tip=MaliHesapTipi.BANKA, banka='ziraat',
        )
        self.banka_yeni = MaliHesap.objects.create(
            sube=self.sube, ad='Vakıfbank', tip=MaliHesapTipi.BANKA, banka='vakifbank',
        )
        today = timezone.localdate()
        BakiyeHareketi.objects.create(
            mali_hesap=self.banka_eski,
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.ey,
            yon=HareketYonu.CIKIS,
            tutar=78_500,
            kaynak=HareketKaynagi.GIDER,
            bakiye_oncesi=78_500,
            bakiye_sonrasi=0,
            islem_tarihi=today,
        )
        BakiyeHareketi.objects.create(
            mali_hesap=self.banka_yeni,
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.ey,
            yon=HareketYonu.GIRIS,
            tutar=252_000,
            kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0,
            bakiye_sonrasi=252_000,
            islem_tarihi=today,
        )
        DonemBakiye.objects.create(
            mali_hesap=self.banka_eski,
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.ey,
            donem_basi_bakiye=78_500,
            toplam_gelir=0,
            toplam_gider=0,
            donem_sonu_bakiye=78_500,
        )

    def test_get_sube_ozet_senkronize_stale_and_missing_accounts(self):
        from apps.finans.application.selectors.donem_bakiye_selector import DonemBakiyeSelector

        ozet = DonemBakiyeSelector().get_sube_ozet(self.sube.id, self.ey.id)
        by_ad = {h['mali_hesap_ad']: h for h in ozet['hesaplar']}

        self.assertEqual(by_ad['Ziraat Bankası (İ.K)']['donem_sonu_bakiye'], 0)
        self.assertEqual(by_ad['Ziraat Bankası (İ.K)']['toplam_gider'], 78_500)
        self.assertIn('Vakıfbank', by_ad)
        self.assertEqual(by_ad['Vakıfbank']['donem_sonu_bakiye'], 252_000)
        self.assertIn('Nakit', by_ad)


class DashboardPosHesaplariTest(TestCase):
    """Finans dashboard'daki 'Kasa & Banka' widget'ı POS hesaplarını da göstermeli.

    Önceki sürümde _mali_hesap_bloklari sadece KASA+BANKA tiplerini döndürüyordu;
    POS tipindeki hesaplar (gerçek parası olsa da) widget'tan tamamen düşüyordu.
    """

    def setUp(self):
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.hareket_types import HareketKaynagi, HareketYonu
        from apps.finans.domain.bakiye_hareketi import BakiyeHareketi

        self.kurum = Kurum.objects.create(ad='Pos Dash Kurum', kod='PDK')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='PDK')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.pos = MaliHesap.objects.create(
            sube=self.sube, ad='Vakıfbank POS', tip=MaliHesapTipi.POS, banka='vakifbank',
        )
        BakiyeHareketi.objects.create(
            mali_hesap=self.pos, kurum=self.kurum, sube=self.sube, egitim_yili=self.ey,
            yon=HareketYonu.GIRIS, tutar=252_000, kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0, bakiye_sonrasi=252_000, islem_tarihi=timezone.localdate(),
        )

    def test_pos_hesap_bloklari_returns_live_balance(self):
        from apps.finans.application.dashboard_overview_service import _pos_hesap_bloklari

        pos, pos_toplam = _pos_hesap_bloklari(self.kurum.id, self.sube.id)
        self.assertEqual(pos_toplam, 252_000)
        self.assertEqual(len(pos), 1)
        self.assertEqual(pos[0]['mali_hesap_ad'], 'Vakıfbank POS')
        self.assertEqual(pos[0]['donem_sonu_bakiye'], 252_000)


class AdminDashboardKasaDagilimiTest(TestCase):
    """Admin dashboard 'Kasa Dağılımı' grafiği KPI kartıyla AYNI canlı kaynağı kullanmalı."""

    def setUp(self):
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.hareket_types import HareketKaynagi, HareketYonu
        from apps.finans.domain.bakiye_hareketi import BakiyeHareketi
        from apps.finans.domain.donem_bakiye import DonemBakiye

        self.kurum = Kurum.objects.create(ad='Admin Dash Kurum', kod='ADK')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='ADK')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.kasa = MaliHesap.objects.create(sube=self.sube, ad='Nakit', tip=MaliHesapTipi.KASA)
        self.banka = MaliHesap.objects.create(
            sube=self.sube, ad='Vakıfbank', tip=MaliHesapTipi.BANKA, banka='vakifbank',
        )
        self.pos = MaliHesap.objects.create(
            sube=self.sube, ad='Garanti POS', tip=MaliHesapTipi.POS, banka='garanti',
        )
        today = timezone.localdate()
        BakiyeHareketi.objects.create(
            mali_hesap=self.kasa, kurum=self.kurum, sube=self.sube, egitim_yili=self.ey,
            yon=HareketYonu.GIRIS, tutar=10_000, kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0, bakiye_sonrasi=10_000, islem_tarihi=today,
        )
        BakiyeHareketi.objects.create(
            mali_hesap=self.banka, kurum=self.kurum, sube=self.sube, egitim_yili=self.ey,
            yon=HareketYonu.GIRIS, tutar=50_000, kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0, bakiye_sonrasi=50_000, islem_tarihi=today,
        )
        BakiyeHareketi.objects.create(
            mali_hesap=self.pos, kurum=self.kurum, sube=self.sube, egitim_yili=self.ey,
            yon=HareketYonu.GIRIS, tutar=7_500, kaynak=HareketKaynagi.TAHSILAT,
            bakiye_oncesi=0, bakiye_sonrasi=7_500, islem_tarihi=today,
        )
        # Eski/stale dönem satırı — KPI ile aynı canlı kaynağa geçtiğimizi doğrular.
        DonemBakiye.objects.create(
            mali_hesap=self.banka, kurum=self.kurum, sube=self.sube, egitim_yili=self.ey,
            donem_basi_bakiye=999_999, toplam_gelir=0, toplam_gider=0, donem_sonu_bakiye=999_999,
        )

    def test_kasa_dagilimi_matches_live_kpi_and_includes_pos(self):
        from apps.dashboard.application.admin_dashboard_service import AdminDashboardService
        from apps.finans.application.dashboard_overview_service import _mali_hesap_bloklari

        kasa, banka, kasa_toplam, banka_toplam = _mali_hesap_bloklari(
            self.kurum.id, self.sube.id, self.ey.id,
        )
        dagilim = AdminDashboardService._kasa_dagilimi(
            kasa, banka, self.kurum.id, self.sube.id, self.ey.id,
        )
        by_label = {d['label']: d['value'] for d in dagilim}

        # Grafikteki Nakit + Banka KPI'daki kasa_banka_toplam ile birebir eşleşmeli
        # (eski sürümde DonemBakiye'nin stale 999.999 değeri görünüyordu).
        self.assertEqual(by_label['Nakit'], 10_000)
        self.assertEqual(by_label['Nakit'], kasa_toplam)
        self.assertEqual(by_label['Banka Hesapları'], 50_000)
        self.assertEqual(by_label['Banka Hesapları'], banka_toplam)
        # POS artık Mali Hesaplar ağacındaki gibi canlı bakiyeden görünür.
        self.assertEqual(by_label['POS Hesapları'], 7_500)

    def test_kasa_banka_kpi_includes_pos_toplam(self):
        """Ana dashboard 'Kasa + Banka' KPI'sı Kasa+Banka+POS/Diğer toplamını yansıtmalı."""
        from apps.dashboard.application.admin_dashboard_service import AdminDashboardService

        data = AdminDashboardService.build(
            kurum_id=self.kurum.id, sube_id=self.sube.id, egitim_yili_id=self.ey.id,
        )
        self.assertEqual(data['genel']['kasa_toplam'], 10_000)
        self.assertEqual(data['genel']['banka_toplam'], 50_000)
        self.assertEqual(data['genel']['pos_toplam'], 7_500)
        self.assertEqual(data['genel']['kasa_banka_toplam'], 10_000 + 50_000 + 7_500)
        self.assertEqual(data['finans']['kpis']['kasa_banka'], data['genel']['kasa_banka_toplam'])


class GelirGiderRaporCompletenessTest(TestCase):
    """Raporlama → Gelir-Gider: tüm gelir kaynakları + nakit bazlı gider."""

    def setUp(self):
        from apps.finans.application.cari_hareket_service import CariHareketService
        from apps.finans.application.gelir_service import GelirService
        from apps.finans.application.gelir_tahsilat_service import GelirTahsilatService
        from apps.finans.application.gider_odeme_service import GiderOdemeService
        from apps.finans.application.gider_service import GiderService
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.cari_types import CariHareketTuru, CariHareketYonu
        from apps.finans.domain.cari_hesap import CariHesap
        from apps.finans.domain.gelir_kategorisi import GelirKategorisi
        from apps.finans.domain.gider_kategorisi import GiderKategorisi

        self.client = APIClient()
        self.kurum = Kurum.objects.create(ad='Rapor Tam Kurum', kod='RTK')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='RTK')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.user = User.objects.create_user(username='rapor_tam', password='test')
        _assign_perms(self.user, 'finans.read')
        self.client.force_authenticate(user=self.user)
        self.today = timezone.localdate()

        mali_hesap = MaliHesap.objects.create(sube=self.sube, ad='Kasa', tip=MaliHesapTipi.KASA)
        yontem = OdemeYontemi.objects.create(
            mali_hesap=mali_hesap, kurum=self.kurum, ad='Nakit',
            tip=OdemeYontemiTipi.NAKIT, komisyon_orani=Decimal('0'),
        )
        gelir_kat = GelirKategorisi.objects.create(kurum=self.kurum, sube=self.sube, ad='Diğer Gelir')
        gider_kat = GiderKategorisi.objects.create(kurum=self.kurum, sube=self.sube, ad='Kira')
        musteri = CariHesap.objects.create(
            kurum=self.kurum, sube=self.sube, unvan='Serbest Müşteri', hesap_turu='musteri',
        )
        tedarikci = CariHesap.objects.create(
            kurum=self.kurum, sube=self.sube, unvan='Kira Tedarikçi', hesap_turu='tedarikci',
        )

        # 1) Sözleşme dışı "serbest gelir" — önceki sürümde rapora HİÇ girmiyordu.
        gelir, err = GelirService().create({
            'kurum_id': self.kurum.id, 'sube_id': self.sube.id, 'egitim_yili_id': self.ey.id,
            'cari_hesap_id': musteri.id, 'gelir_kategorisi_id': gelir_kat.id,
            'fatura_tarihi': self.today, 'vade_tarihi': self.today,
            'brut_tutar': Decimal('5000'), 'kdv_orani': 0, 'olusturan': self.user,
        })
        self.assertIsNone(err, err)
        _, err = GelirTahsilatService().tahsilat_yap({
            'gelir_kaydi_id': gelir.id, 'tutar': Decimal('2000'),
            'tahsilat_tarihi': self.today, 'mali_hesap_id': mali_hesap.id,
            'odeme_yontemi_id': yontem.id, 'islem_yapan': self.user,
        })
        self.assertIsNone(err, err)
        self.serbest_gelir_tutari = 2000

        # 2) Bağımsız (sözleşme/gelir kaydına bağlı olmayan) cari tahsilatı.
        CariHareketService().hareket_olustur(
            cari_hesap_id=musteri.id, kurum_id=self.kurum.id, sube_id=self.sube.id,
            egitim_yili_id=self.ey.id, tutar=Decimal('1500'),
            yon=CariHareketYonu.ALACAK, islem_turu=CariHareketTuru.TAHSILAT,
            islem_tarihi=self.today, aciklama='Bağımsız tahsilat',
        )
        self.cari_gelir_tutari = 1500

        # 3) Gider: 8000 tutarında fatura, sadece 3000'i ÖDENDİ.
        # Eski rapor GiderKaydi.net_tutar (fatura tarihi) kullanıyordu → 8000 gösterirdi.
        gider, err = GiderService().create({
            'kurum_id': self.kurum.id, 'sube_id': self.sube.id, 'egitim_yili_id': self.ey.id,
            'cari_hesap_id': tedarikci.id, 'gider_kategorisi_id': gider_kat.id,
            'fatura_tarihi': self.today, 'vade_tarihi': self.today,
            'brut_tutar': Decimal('8000'), 'kdv_orani': 0, 'olusturan': self.user,
        })
        self.assertIsNone(err, err)
        _, err = GiderOdemeService().odeme_yap({
            'gider_kaydi_id': gider.id, 'tutar': Decimal('3000'),
            'odeme_tarihi': self.today, 'mali_hesap_id': mali_hesap.id,
            'odeme_yontemi_id': yontem.id, 'islem_yapan': self.user,
        })
        self.assertIsNone(err, err)
        self.odenen_gider_tutari = 3000
        self.tahakkuk_gider_tutari = 8000

    def _bu_ay_satiri(self, body):
        bu_ay = self.today.strftime('%Y-%m')
        for a in body['aylik']:
            if a['ay'] == bu_ay:
                return a
        self.fail('Bu ayın satırı bulunamadı')

    def test_gelir_includes_serbest_gelir_and_cari_tahsilat(self):
        res = self.client.get(
            '/finans/api/raporlar/gelir-gider/',
            {'kurum_id': self.kurum.id, 'egitim_yili_id': self.ey.id},
            HTTP_X_SUBE_ID=str(self.sube.id),
        )
        self.assertEqual(res.status_code, 200, res.content[:300])
        body = res.json()
        satir = self._bu_ay_satiri(body)
        self.assertEqual(
            satir['gelir'], self.serbest_gelir_tutari + self.cari_gelir_tutari,
        )

    def test_gider_uses_cash_basis_not_accrual(self):
        res = self.client.get(
            '/finans/api/raporlar/gelir-gider/',
            {'kurum_id': self.kurum.id, 'egitim_yili_id': self.ey.id},
            HTTP_X_SUBE_ID=str(self.sube.id),
        )
        self.assertEqual(res.status_code, 200, res.content[:300])
        body = res.json()
        satir = self._bu_ay_satiri(body)
        self.assertEqual(satir['gider'], self.odenen_gider_tutari)
        self.assertNotEqual(satir['gider'], self.tahakkuk_gider_tutari)


class GunSonuNetFinansalSonucCariTest(TestCase):
    """Gün Sonu Özet Raporu — net sonuç, Gelir-Gider raporuyla aynı gelir tanımını kullanmalı."""

    def setUp(self):
        from apps.finans.application.cari_hareket_service import CariHareketService
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.cari_types import CariHareketTuru, CariHareketYonu
        from apps.finans.domain.cari_hesap import CariHesap

        self.kurum = Kurum.objects.create(ad='Gün Sonu Cari Kurum', kod='GSC')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='GSC')
        self.today = timezone.localdate()

        mali_hesap = MaliHesap.objects.create(sube=self.sube, ad='Kasa', tip=MaliHesapTipi.KASA)
        OdemeYontemi.objects.create(
            mali_hesap=mali_hesap, kurum=self.kurum, ad='Nakit',
            tip=OdemeYontemiTipi.NAKIT, komisyon_orani=Decimal('0'),
        )
        musteri = CariHesap.objects.create(
            kurum=self.kurum, sube=self.sube, unvan='Serbest Müşteri', hesap_turu='musteri',
        )

        # Sözleşmeye/gelir kaydına bağlı olmayan bağımsız cari tahsilatı — "Toplam
        # Alınan" KPI'sına dahildi ama önceki sürümde net sonuca hiç girmiyordu.
        CariHareketService().hareket_olustur(
            cari_hesap_id=musteri.id, kurum_id=self.kurum.id, sube_id=self.sube.id,
            egitim_yili_id=None, tutar=Decimal('1500'),
            yon=CariHareketYonu.ALACAK, islem_turu=CariHareketTuru.TAHSILAT,
            islem_tarihi=self.today, aciklama='Bağımsız tahsilat',
        )
        self.cari_tahsilat_tutari = 1500

    def test_net_gunluk_finansal_sonuc_includes_bagimsiz_cari_tahsilat(self):
        from apps.finans.application.gun_sonu_report_service import GunSonuReportService

        rapor = GunSonuReportService().build_ozet_rapor(self.kurum.id, self.today, self.sube.id)
        gunluk = rapor['ozet_rapor']['gunluk_ozet']

        self.assertEqual(gunluk['toplam_alinan'], self.cari_tahsilat_tutari)
        # Önceki sürümde bu alan 0 kalırdı çünkü sadece sözleşme + gelir tahsilatı
        # sayıyordu; artık "Toplam Alınan" ile aynı kapsamda olmalı.
        self.assertEqual(gunluk['net_gunluk_finansal_sonuc'], self.cari_tahsilat_tutari)


class SonBakiyeChainCorruptionTest(TestCase):
    """
    son_bakiye() zincir bozulmasına karşı dayanıklı olmalı.

    Canlıda görülen gerçek senaryo: toplu içe aktarma sırasında her BakiyeHareketi
    satırı servis kilidi (select_for_update + son_bakiye okuma) DIŞINDA doğrudan
    repository.create ile yazılmış; bu yüzden her satırın bakiye_oncesi'si 0'dan
    başlamış, kümülatif zincir kopmuş. Sonuç: hesabın son satırı bakiye_sonrasi=0
    gösterirken gerçek toplam giriş−çıkış 78.500 idi (Mali Hesaplar ekranı 0,
    Hareketler sekmesindeki Toplam Giriş/Çıkış/Net kartı 78.500 gösteriyordu).
    """

    def setUp(self):
        from apps.finans.constants.account_types import MaliHesapTipi
        from apps.finans.constants.hareket_types import HareketKaynagi, HareketYonu
        from apps.finans.domain.bakiye_hareketi import BakiyeHareketi

        self.kurum = Kurum.objects.create(ad='Zincir Kurum', kod='ZNC')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='ZNC')
        self.ey = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.hesap = MaliHesap.objects.create(sube=self.sube, ad='Ziraat', tip=MaliHesapTipi.BANKA)

        def _row(gun, yon, tutar, bakiye_oncesi, bakiye_sonrasi):
            return BakiyeHareketi.objects.create(
                mali_hesap=self.hesap, kurum=self.kurum, sube=self.sube, egitim_yili=self.ey,
                yon=yon, tutar=tutar, kaynak=HareketKaynagi.TAHSILAT,
                bakiye_oncesi=bakiye_oncesi, bakiye_sonrasi=bakiye_sonrasi,
                islem_tarihi=date(2026, 5, 4) + timedelta(days=gun),
            )

        # Her satır kendi bakiye_oncesi'ni 0 kabul ediyor — gerçek toplu
        # aktarım hatasının aynısı (bkz. yukarıdaki docstring).
        _row(0, HareketYonu.GIRIS, 5_000, 0, 5_000)
        _row(45, HareketYonu.GIRIS, 19_500, 0, 19_500)
        _row(51, HareketYonu.GIRIS, 25_000, 0, 25_000)
        _row(53, HareketYonu.GIRIS, 10_000, 5_000, 15_000)
        _row(57, HareketYonu.GIRIS, 19_000, 0, 19_000)
        _row(69, HareketYonu.GIRIS, 5_000, 0, 5_000)
        _row(69, HareketYonu.CIKIS, 5_000, 5_000, 0)  # zincirdeki son satır

    def test_son_bakiye_sum_bazli_dogru_toplami_verir(self):
        from apps.finans.application.selectors.bakiye_hareketi_selector import BakiyeHareketiSelector

        selector = BakiyeHareketiSelector()
        # Zincir (son satırın bakiye_sonrasi'si) 0 derdi; gerçek net 78.500.
        self.assertEqual(selector.get_son_bakiye(self.hesap.id), 78_500)
        ozet = selector.get_ozet(self.hesap.id, self.ey.id)
        self.assertEqual(ozet['net'], 78_500)
        # Mali Hesaplar ağacı ile Hareketler sekmesi artık aynı sayıyı göstermeli.
        self.assertEqual(selector.get_son_bakiye(self.hesap.id), ozet['net'])

    def test_yeni_hareket_dogru_bakiye_oncesi_ile_zinciri_onarir(self):
        from apps.finans.application.bakiye_hareketi_service import BakiyeHareketiService

        yeni = BakiyeHareketiService().hareket_olustur(
            mali_hesap_id=self.hesap.id, kurum_id=self.kurum.id, sube_id=self.sube.id,
            egitim_yili_id=self.ey.id, yon='giris', tutar=1_000,
            kaynak='tahsilat', islem_tarihi=date(2026, 7, 20),
        )
        # Bozuk zincirden (0) değil, gerçek toplamdan (78.500) devam etmeli.
        self.assertEqual(yeni.bakiye_oncesi, 78_500)
        self.assertEqual(yeni.bakiye_sonrasi, 79_500)
