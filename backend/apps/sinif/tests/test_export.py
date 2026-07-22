"""Sınıf/Şube listesi dışa aktarma — CSV/Excel/JSON."""
from django.contrib.auth import get_user_model
from django.test import Client, TestCase

from apps.egitim_yili.domain.models import EgitimYili
from apps.kurum.domain.models import Kurum
from apps.oda.domain.models import Oda
from apps.sinif.domain.models import Sinif
from apps.sube.domain.models import Sube

User = get_user_model()


class SinifExportTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.kurum = Kurum.objects.create(ad='Export Test Kurum', kod='EXPT')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez Şube', kod='EXPT-M')
        self.sube_b = Sube.objects.create(kurum=self.kurum, ad='Diğer Şube', kod='EXPT-D')
        self.user = User.objects.create_user(username='exporttest', password='test')
        self.client.force_login(self.user)

        self.egitim_yili = EgitimYili.objects.create(baslangic_yil=2025, bitis_yil=2026, aktif_mi=True)
        self.oda = Oda.objects.create(kurum=self.kurum, sube=self.sube, ad='A101', kapasite=30)

        self.sinif_a = Sinif.objects.create(
            kurum=self.kurum, sube=self.sube, egitim_yili=self.egitim_yili,
            ad='9-A', kod='9A', kapasite=30, oda=self.oda, aktif_mi=True,
        )
        self.sinif_b = Sinif.objects.create(
            kurum=self.kurum, sube=self.sube, egitim_yili=self.egitim_yili,
            ad='9-B', kod='9B', kapasite=25, aktif_mi=True,
        )
        # Başka şubede bir sınıf — izolasyon kontrolü için
        Sinif.objects.create(
            kurum=self.kurum, sube=self.sube_b, egitim_yili=self.egitim_yili,
            ad='10-A', kod='10A', kapasite=20, aktif_mi=True,
        )

    def _headers(self, sube=None):
        headers = {'HTTP_X_KURUM_ID': str(self.kurum.id)}
        if sube is not None:
            headers['HTTP_X_SUBE_ID'] = str(sube.id)
        return headers

    def test_export_requires_sube(self):
        res = self.client.get('/siniflar/api/export/', **self._headers())
        self.assertEqual(res.status_code, 400)

    def test_export_json_scoped_to_active_sube(self):
        res = self.client.get(
            '/siniflar/api/export/?format=json',
            **self._headers(self.sube),
        )
        self.assertEqual(res.status_code, 200)
        body = res.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['total'], 2)
        names = {row['ad'] for row in body['rows']}
        self.assertEqual(names, {'9-A', '9-B'})
        self.assertIn('ogrenci_sayisi', body['columns'])

    def test_export_xlsx_returns_valid_workbook(self):
        res = self.client.get(
            '/siniflar/api/export/?format=xlsx',
            **self._headers(self.sube),
        )
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            res['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', res['Content-Disposition'])

        import io
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        all_text = ' '.join(
            str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
        )
        self.assertIn('9-A', all_text)
        self.assertIn('9-B', all_text)
        self.assertIn('Export Test Kurum', all_text)
        self.assertIn('Merkez Şube', all_text)

    def test_export_csv_returns_utf8_bom_content(self):
        res = self.client.get(
            '/siniflar/api/export/?format=csv',
            **self._headers(self.sube),
        )
        self.assertEqual(res.status_code, 200)
        self.assertIn('text/csv', res['Content-Type'])
        content = res.content.decode('utf-8-sig')
        self.assertIn('9-A', content)
        self.assertIn('9-B', content)

    def test_export_filters_by_aktif(self):
        Sinif.objects.create(
            kurum=self.kurum, sube=self.sube, egitim_yili=self.egitim_yili,
            ad='9-C', kod='9C', kapasite=20, aktif_mi=False,
        )
        res = self.client.get(
            '/siniflar/api/export/?format=json&aktif=false',
            **self._headers(self.sube),
        )
        self.assertEqual(res.status_code, 200)
        body = res.json()
        names = {row['ad'] for row in body['rows']}
        self.assertEqual(names, {'9-C'})
