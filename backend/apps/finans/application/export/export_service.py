"""
Finans dışa aktarma servisi — CSV, XLSX, PDF (Faz 4).
"""
from __future__ import annotations

import csv
import io
from typing import Any

from django.http import HttpResponse
from django.utils import timezone


class ExportService:
    """Tablo verisini json/csv/xlsx/pdf formatlarına dönüştürür."""

    SUPPORTED_FORMATS = ('json', 'csv', 'xlsx', 'pdf')

    @classmethod
    def build(
        cls,
        export_format: str,
        rows: list[dict[str, Any]],
        columns: list[dict[str, str]],
        *,
        title: str = 'Rapor',
        filters_meta: dict[str, Any] | None = None,
        orientation: str = 'portrait',
    ):
        fmt = (export_format or 'json').lower()
        if fmt not in cls.SUPPORTED_FORMATS:
            raise ValueError(f'Desteklenmeyen format: {export_format}')

        ori = cls._normalize_orientation(orientation)

        if fmt == 'json':
            return cls._build_json(rows, columns, title=title, filters_meta=filters_meta)
        if fmt == 'csv':
            return cls._build_csv(rows, columns, title=title, filters_meta=filters_meta)
        if fmt == 'xlsx':
            return cls._build_xlsx(rows, columns, title=title, filters_meta=filters_meta, orientation=ori)
        return cls._build_pdf(rows, columns, title=title, filters_meta=filters_meta, orientation=ori)

    @staticmethod
    def _normalize_orientation(value: str | None) -> str:
        v = (value or 'portrait').lower().strip()
        if v in ('landscape', 'yatay', 'horizontal'):
            return 'landscape'
        return 'portrait'

    @classmethod
    def _column_keys(cls, columns: list[dict[str, str]]) -> tuple[list[str], list[str]]:
        keys = [c['key'] for c in columns]
        labels = [c.get('label', c['key']) for c in columns]
        return keys, labels

    @classmethod
    def _row_values(cls, row: dict, keys: list[str]) -> list[Any]:
        return [row.get(k, '') for k in keys]

    @classmethod
    def _build_json(cls, rows, columns, *, title, filters_meta):
        keys, labels = cls._column_keys(columns)
        return {
            'title': title,
            'filters': filters_meta or {},
            'columns': [{'key': k, 'label': l} for k, l in zip(keys, labels)],
            'rows': [{k: row.get(k, '') for k in keys} for row in rows],
            'count': len(rows),
        }

    _CSV_META_SKIP_KEYS = frozenset({
        'kurum_ad', 'sube_ad', 'sube', 'kurum_id', 'sube_id',
        'toplam', 'toplam_tutar', 'adet', 'count', 'toplam_kalan',
    })

    @classmethod
    def _write_csv_branding(cls, writer, *, title: str, filters_meta: dict[str, Any] | None = None) -> None:
        """CSV üst bilgi satırları — logo yalnızca PDF/Excel'de gösterilir."""
        from apps.finans.application.export.report_html_template import _format_filter_label

        meta = filters_meta or {}
        kurum_ad = meta.get('kurum_ad') or ''
        sube_ad = meta.get('sube_ad') or meta.get('sube') or ''
        if kurum_ad:
            writer.writerow([kurum_ad])
        if sube_ad:
            writer.writerow([f'Şube: {sube_ad}'])
        writer.writerow([title])
        now = timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')
        writer.writerow([f'Oluşturulma: {now}'])
        for fk, fv in meta.items():
            if fk in cls._CSV_META_SKIP_KEYS or fv in (None, ''):
                continue
            writer.writerow([f'{_format_filter_label(fk)}: {fv}'])
        writer.writerow([])

    @classmethod
    def _build_csv(cls, rows, columns, *, title, filters_meta=None):
        keys, labels = cls._column_keys(columns)
        buf = io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf, delimiter=';')
        cls._write_csv_branding(writer, title=title, filters_meta=filters_meta)
        writer.writerow(labels)
        for row in rows:
            writer.writerow([cls._format_cell(row.get(k, '')) for k in keys])
        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.csv"'
        return response

    @classmethod
    def _build_xlsx(cls, rows, columns, *, title, filters_meta, orientation='portrait'):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError as exc:
            raise RuntimeError('openpyxl yüklü değil.') from exc

        keys, labels = cls._column_keys(columns)
        wb = Workbook()
        ws = wb.active
        ws.title = (title or 'Rapor')[:31]

        row_idx = 1
        if filters_meta and filters_meta.get('kurum_ad'):
            ws.cell(row=row_idx, column=1, value=str(filters_meta['kurum_ad'])).font = Font(bold=True, size=12)
            row_idx += 1
        sube_label = (filters_meta or {}).get('sube_ad') or (filters_meta or {}).get('sube')
        if sube_label:
            ws.cell(row=row_idx, column=1, value=f'Şube: {sube_label}')
            row_idx += 1
        ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True)
        row_idx += 1
        if filters_meta:
            for fk, fv in filters_meta.items():
                if fk in cls._CSV_META_SKIP_KEYS or fv in (None, ''):
                    continue
                from apps.finans.application.export.report_html_template import _format_filter_label
                ws.cell(row=row_idx, column=1, value=f'{_format_filter_label(fk)}: {fv}')
                row_idx += 1
            row_idx += 1

        for col_idx, label in enumerate(labels, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.font = Font(bold=True)
        row_idx += 1

        for row in rows:
            for col_idx, key in enumerate(keys, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cls._format_cell(row.get(key, '')))
            row_idx += 1

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = (
            'landscape' if orientation == 'landscape' else 'portrait'
        )
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        buf = io.BytesIO()
        wb.save(buf)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.xlsx"'
        return response

    @classmethod
    def _build_pdf(cls, rows, columns, *, title, filters_meta, orientation='portrait'):
        from apps.communication.application.html_to_pdf import render_html_to_pdf

        meta = filters_meta or {}
        report_kind = meta.get('report_kind')
        kurum_ad = str(meta["kurum_ad"]) if meta.get("kurum_ad") else None

        if report_kind in ('cari_bakiye', 'cari_ozet', 'cari_ekstre'):
            from apps.finans.application.export.cari_report_html import (
                build_cari_report_html,
                cari_report_footer_template,
            )

            html_doc = build_cari_report_html(
                title=title or "Cari Hesap Bakiye Raporu",
                columns=columns,
                rows=rows,
                filters_meta=meta,
                kurum_ad=kurum_ad,
                orientation=orientation,
            )
            footer = cari_report_footer_template(meta)
            try:
                pdf_bytes = render_html_to_pdf(
                    html_doc,
                    landscape=(orientation == 'landscape'),
                    footer_template=footer,
                )
            except RuntimeError:
                pdf_bytes = render_html_to_pdf(
                    html_doc,
                    landscape=(orientation == 'landscape'),
                )
        else:
            from apps.finans.application.export.report_html_template import build_finans_report_html

            keys, _ = cls._column_keys(columns)
            summary = {}
            if filters_meta:
                for k in ("toplam", "toplam_tutar", "adet", "count", "toplam_kalan"):
                    if k in (filters_meta or {}):
                        summary[k] = filters_meta[k]

            html_doc = build_finans_report_html(
                title=title or "Finans Raporu",
                columns=columns,
                rows=rows,
                filters_meta={k: v for k, v in (filters_meta or {}).items() if k not in ("kurum_ad",)},
                kurum_ad=kurum_ad,
                summary=summary or None,
                orientation=orientation,
            )
            try:
                pdf_bytes = render_html_to_pdf(
                    html_doc,
                    landscape=(orientation == 'landscape'),
                )
            except RuntimeError:
                from apps.communication.application.pdf_render_service import PdfRenderService

                keys, labels = cls._column_keys(columns)
                lines = [title or "Rapor", ""]
                if filters_meta:
                    for fk, fv in filters_meta.items():
                        if fk != "kurum_ad":
                            lines.append(f"{fk}: {fv}")
                    lines.append("")
                header = " | ".join(labels)
                lines.append(header)
                lines.append("-" * min(len(header), 120))
                for row in rows[:500]:
                    lines.append(" | ".join(str(cls._format_cell(row.get(k, ""))) for k in keys))
                pdf_bytes = PdfRenderService.render_simple_text_pdf(title or "Rapor", "\n".join(lines))

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{cls._safe_filename(title)}.pdf"'
        return response

    @staticmethod
    def _format_cell(value: Any) -> str:
        if value is None:
            return ''
        if isinstance(value, float):
            return f'{value:.2f}'
        return str(value)

    @staticmethod
    def _safe_filename(name: str) -> str:
        safe = ''.join(c if c.isalnum() or c in '-_' else '_' for c in (name or 'rapor'))
        return safe[:80] or 'rapor'
