"""
Gün Sonu Detay Raporu — PDF / çok sayfalı Excel / tek CSV export.
"""
from __future__ import annotations

import csv
import io

from django.http import HttpResponse

from apps.communication.application.html_to_pdf import render_html_to_pdf
from apps.finans.application.export.export_service import ExportService
from apps.finans.application.export.gun_sonu_detay_report_html import build_gun_sonu_detay_html


class GunSonuDetayExportService:
    SUPPORTED_FORMATS = ('pdf', 'csv', 'xlsx')

    @classmethod
    def build(cls, export_format: str, report: dict, *, orientation: str = 'landscape'):
        fmt = (export_format or 'pdf').lower()
        if fmt not in cls.SUPPORTED_FORMATS:
            raise ValueError(f'Desteklenmeyen format: {export_format}')

        ori = ExportService._normalize_orientation(orientation)
        detay = report.get('detay_rapor') or {}
        meta = detay.get('meta') or {}
        title = f"Gun_Sonu_Detay_{meta.get('tarih_iso', 'rapor')}"

        if fmt == 'pdf':
            return cls._build_pdf(report, title, orientation=ori)
        if fmt == 'csv':
            return cls._build_csv(report, title)
        return cls._build_xlsx(report, title, orientation=ori)

    @classmethod
    def _build_pdf(cls, report: dict, title: str, *, orientation: str) -> HttpResponse:
        html_doc = build_gun_sonu_detay_html(report, orientation=orientation)
        pdf_bytes = render_html_to_pdf(html_doc, landscape=(orientation == 'landscape'))
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.pdf"'
        return response

    @classmethod
    def _build_csv(cls, report: dict, title: str) -> HttpResponse:
        """Tek CSV dosyası — bölüm başlıkları ve marka üst bilgisi (logo PDF/Excel'de)."""
        detay = report.get('detay_rapor') or {}
        meta = detay.get('meta') or {}

        buf = io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf, delimiter=';')

        kurum = meta.get('kurum_ad') or meta.get('marka') or ''
        if kurum:
            writer.writerow([kurum])
        writer.writerow([meta.get('baslik', 'GÜN SONU DETAY RAPORU')])
        writer.writerow([f"Tarih: {meta.get('tarih', '')}"])
        writer.writerow([f"Şube: {meta.get('sube', '')}"])
        writer.writerow([f"Hazırlayan: {meta.get('hazirlayan', '')}"])
        writer.writerow([f"Oluşturulma: {meta.get('olusturulma', '')}"])
        writer.writerow([])

        def write_section(section_title: str, headers: list[str], rows: list[list]):
            writer.writerow([section_title])
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            writer.writerow([])

        write_section('Özet', ['Kalem', 'Tutar'], cls._ozet_rows(detay))
        write_section(
            'Tahsilatlar',
            ['Saat', 'Sözleşme No', 'Makbuz', 'Öğrenci', 'Veli', 'Taksit', 'Dönem', 'Ödeme Türü', 'Tutar', 'Personel', 'Açıklama'],
            [
                [r['saat'], r.get('sozlesme_no', ''), r['makbuz'], r['ogrenci'], r['veli'],
                 r.get('taksit_no', ''), r.get('odeme_donemi', ''), r['odeme_turu'], r['tutar'], r['personel'], r.get('aciklama', '')]
                for r in detay.get('tahsilat_listesi') or []
            ],
        )
        write_section(
            'Gelirler',
            ['Saat', 'Gelir Kodu', 'Kategori', 'Açıklama', 'Tutar', 'Personel'],
            [
                [r['saat'], r['gelir_kodu'], r['kategori'], r['aciklama'], r['tutar'], r['personel']]
                for r in detay.get('gelir_hareketleri') or []
            ],
        )
        write_section(
            'Giderler',
            ['Saat', 'Gider Kodu', 'Kategori', 'Açıklama', 'Tutar', 'Personel'],
            [
                [r['saat'], r['gider_kodu'], r['kategori'], r['aciklama'], r['tutar'], r['personel']]
                for r in detay.get('gider_hareketleri') or []
            ],
        )
        write_section(
            'Cari Hareketleri',
            ['Cari', 'Borç', 'Alacak', 'Bakiye'],
            [
                [r['cari'], r['borc'], r['alacak'], r['bakiye']]
                for r in detay.get('cari_hareketleri') or []
            ],
        )
        write_section(
            'İptaller',
            ['Saat', 'İşlem No', 'Tür', 'Sebep', 'Kullanıcı', 'Tutar'],
            [
                [r['saat'], r['islem_no'], r['tur'], r['sebep'], r['kullanici'], r.get('tutar', '')]
                for r in detay.get('iptal_islemleri') or []
            ],
        )
        write_section(
            'İadeler',
            ['Saat', 'Öğrenci', 'Tutar', 'Açıklama'],
            [
                [r['saat'], r['ogrenci'], r['tutar'], r['aciklama']]
                for r in detay.get('iade_islemleri') or []
            ],
        )
        write_section(
            'Ödeme Türleri',
            ['Ödeme Türü', 'Adet', 'Tutar'],
            [
                [r['label'], r.get('adet', ''), r['tutar']]
                for r in (detay.get('odeme_turu_dagilimi') or {}).get('ozet') or []
            ],
        )
        write_section(
            'Kategori Analizi',
            ['Tür', 'Kategori', 'Tutar'],
            [['Gelir', r['kategori'], r['tutar']] for r in detay.get('kategori_gelirler') or [] if not r.get('is_total')]
            + [['Gider', r['kategori'], r['tutar']] for r in detay.get('kategori_giderler') or [] if not r.get('is_total')],
        )
        write_section(
            'Kullanıcı Detay',
            ['Personel', 'Saat', 'Tür', 'Açıklama', 'Tutar'],
            cls._kullanici_flat_rows(detay),
        )
        write_section(
            'Kasa Özeti',
            ['Kalem', 'Tutar'],
            cls._kasa_rows(detay.get('kasa_ozeti') or {}),
        )

        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.csv"'
        return response

    @classmethod
    def _build_xlsx(cls, report: dict, title: str, *, orientation: str) -> HttpResponse:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError('openpyxl yüklü değil.') from exc

        detay = report.get('detay_rapor') or {}
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='1F3C88')
        total_font = Font(bold=True)

        def write_sheet(name: str, headers: list[str], rows: list[list], *, money_cols: set[int] | None = None):
            ws = wb.create_sheet(name[:31])
            money_cols = money_cols or set()
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for ri, row in enumerate(rows, 2):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    if ci in money_cols and isinstance(val, (int, float)):
                        cell.number_format = '#,##0 ₺'
                        cell.alignment = Alignment(horizontal='right')
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 16
            ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 1)}'
            ws.page_setup.orientation = 'landscape' if orientation == 'landscape' else 'portrait'
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            return ws

        write_sheet('Özet', ['Kalem', 'Tutar (TL)'], cls._ozet_rows(detay), money_cols={2})
        write_sheet(
            'Tahsilatlar',
            ['Saat', 'Sözleşme', 'Makbuz', 'Öğrenci', 'Veli', 'Taksit', 'Dönem', 'Ödeme', 'Tutar', 'Personel'],
            [[r['saat'], r.get('sozlesme_no', ''), r['makbuz'], r['ogrenci'], r['veli'],
              r.get('taksit_no', ''), r.get('odeme_donemi', ''), r['odeme_turu'], r['tutar'], r['personel']]
             for r in detay.get('tahsilat_listesi') or []],
            money_cols={9},
        )
        write_sheet(
            'Gelirler',
            ['Saat', 'Gelir Kodu', 'Kategori', 'Açıklama', 'Tutar', 'Personel'],
            [[r['saat'], r['gelir_kodu'], r['kategori'], r['aciklama'], r['tutar'], r['personel']]
             for r in detay.get('gelir_hareketleri') or []],
            money_cols={5},
        )
        write_sheet(
            'Giderler',
            ['Saat', 'Gider Kodu', 'Kategori', 'Açıklama', 'Tutar', 'Personel'],
            [[r['saat'], r['gider_kodu'], r['kategori'], r['aciklama'], r['tutar'], r['personel']]
             for r in detay.get('gider_hareketleri') or []],
            money_cols={5},
        )
        write_sheet(
            'Cari Hareketleri',
            ['Cari', 'Borç', 'Alacak', 'Bakiye'],
            [[r['cari'], r['borc'], r['alacak'], r['bakiye']] for r in detay.get('cari_hareketleri') or []],
            money_cols={2, 3, 4},
        )
        write_sheet(
            'İptaller',
            ['Saat', 'İşlem No', 'Tür', 'Sebep', 'Kullanıcı', 'Tutar'],
            [[r['saat'], r['islem_no'], r['tur'], r['sebep'], r['kullanici'], r.get('tutar', 0)]
             for r in detay.get('iptal_islemleri') or []],
            money_cols={6},
        )
        write_sheet(
            'İadeler',
            ['Saat', 'Öğrenci', 'Tutar', 'Açıklama'],
            [[r['saat'], r['ogrenci'], r['tutar'], r['aciklama']] for r in detay.get('iade_islemleri') or []],
            money_cols={3},
        )
        write_sheet(
            'Kullanıcı Özeti',
            ['Personel', 'Saat', 'Tür', 'Açıklama', 'Tutar'],
            cls._kullanici_flat_rows(detay),
            money_cols={5},
        )
        write_sheet(
            'Ödeme Türleri',
            ['Ödeme Türü', 'Adet', 'Tutar'],
            [[r['label'], r.get('adet'), r['tutar']] for r in (detay.get('odeme_turu_dagilimi') or {}).get('ozet') or []],
            money_cols={3},
        )

        kat_rows = (
            [['Gelir', r['kategori'], r['tutar']] for r in detay.get('kategori_gelirler') or [] if not r.get('is_total')]
            + [['Gider', r['kategori'], r['tutar']] for r in detay.get('kategori_giderler') or [] if not r.get('is_total')]
        )
        ws_kat = write_sheet('Kategori Analizi', ['Tür', 'Kategori', 'Tutar'], kat_rows, money_cols={3})
        if kat_rows:
            total_row = len(kat_rows) + 2
            ws_kat.cell(row=total_row, column=2, value='GENEL TOPLAM').font = total_font
            ws_kat.cell(row=total_row, column=3, value=sum(r[2] for r in kat_rows)).font = total_font

        write_sheet('Kasa Özeti', ['Kalem', 'Tutar'], cls._kasa_rows(detay.get('kasa_ozeti') or {}), money_cols={2})

        write_sheet(
            'Kasa Hareketleri',
            ['Saat', 'Kasa', 'Yön', 'Kaynak', 'Tutar', 'Açıklama', 'Personel'],
            [[r['saat'], r['kasa'], r['yon'], r['kaynak'], r['tutar'], r.get('aciklama', ''), r['personel']]
             for r in detay.get('kasa_hareketleri') or []],
            money_cols={5},
        )

        banka = detay.get('banka_hareketleri') or {}
        write_sheet(
            'Banka Hareketleri',
            ['Saat', 'Banka', 'Yön', 'Tür', 'Tutar', 'Açıklama'],
            [[r['saat'], r['banka'], r['yon'], r['tur'], r['tutar'], r.get('aciklama', '')]
             for r in banka.get('detay') or []],
            money_cols={5},
        )

        write_sheet(
            'POS Hareketleri',
            ['POS Cihazı', 'Banka', 'Kart Türü', 'Tutar', 'İşlem Sayısı'],
            [[r['pos_cihazi'], r['banka'], r['kart_turu'], r['tutar'], r['islem_sayisi']]
             for r in detay.get('pos_hareketleri') or []],
            money_cols={4},
        )

        write_sheet(
            'Personel Performans',
            ['Personel', 'Tahsilat Adet', 'Tahsilat Tutarı', 'Gelir', 'Gider', 'İade', 'İptal', 'Toplam İşlem'],
            [[r['personel'], r['tahsilat_sayisi'], r['tahsilat_tutari'], r['gelir_sayisi'],
              r['gider_sayisi'], r['iade_sayisi'], r['iptal_sayisi'], r['toplam_islem']]
             for r in detay.get('personel_performans') or []],
            money_cols={3},
        )

        buf = io.BytesIO()
        wb.save(buf)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.xlsx"'
        return response

    @staticmethod
    def _ozet_rows(detay: dict) -> list[list]:
        ozet = detay.get('ozet') or {}
        finans = detay.get('gunluk_finans_ozeti') or {}
        rows = [
            ['Toplam Tahsilat (Sözleşme)', ozet.get('toplam_tahsilat', 0)],
            ['Toplam Alınan (Dashboard uyumlu)', ozet.get('toplam_alinan', 0)],
            ['Toplam Gelir', ozet.get('toplam_gelir', 0)],
            ['Toplam Gider', ozet.get('toplam_gider', 0)],
            ['Toplam İade', ozet.get('toplam_iade', 0)],
            ['Net Nakit (Kasa)', ozet.get('net_nakit_girisi', 0)],
        ]
        if finans:
            rows.append(['Net Günlük Finansal Sonuç', finans.get('net_gunluk_finansal_sonuc', 0)])
        return rows

    @staticmethod
    def _kasa_rows(kasa: dict) -> list[list]:
        return [
            ['Açılış Kasası', kasa.get('acilis_kasa', 0)],
            ['Nakit Tahsilatlar', kasa.get('nakit_tahsilatlar', 0)],
            ['Nakit Gelirler', kasa.get('nakit_gelirler', 0)],
            ['Nakit Giderler', kasa.get('nakit_giderler', 0)],
            ['Kasaya Para Girişi', kasa.get('kasaya_para_girisi', 0)],
            ['Kasadan Para Çıkışı', kasa.get('kasadan_para_cikisi', 0)],
            ['Bankaya Aktarım', kasa.get('bankaya_aktarim', 0)],
            ['Bankadan Kasaya Aktarım', kasa.get('bankadan_kasaya_aktarim', 0)],
            ['Beklenen Kasa', kasa.get('beklenen_kasa', 0)],
            ['Sayılan Kasa', kasa.get('sayilan_kasa') if kasa.get('sayilan_kasa') is not None else ''],
            ['Kasa Farkı', kasa.get('kasa_farki', 0)],
        ]

    @staticmethod
    def _kullanici_flat_rows(detay: dict) -> list[list]:
        rows = []
        for block in detay.get('kullanici_islem_detayi') or []:
            for i in block.get('islemler') or []:
                rows.append([block['personel'], i['saat'], i['tur'], i['aciklama'], i['tutar']])
        return rows

    @staticmethod
    def _safe_filename(name: str) -> str:
        safe = ''.join(c if c.isalnum() or c in '-_' else '_' for c in (name or 'rapor'))
        return safe[:80] or 'rapor'
