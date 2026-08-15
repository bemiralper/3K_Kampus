"""
Sıralama listesi Excel/CSV dışa aktarma — /exams/{pk}/analysis/rankings/?format=xlsx|csv
"""
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from rest_framework.test import APIClient

from apps.coaching.olcme_degerlendirme.models import (
    Exam, ExamSection, ExamSession, StudentAnswer, StudentSectionScore,
)
from apps.egitim_yili.domain.models import EgitimYili
from apps.kurum.domain.models import Kurum
from apps.ogrenci.domain.models import Ogrenci
from apps.sube.domain.models import Sube

User = get_user_model()


class OlcmeRankingsExportAPITest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.kurum = Kurum.objects.create(ad='Sıralama Export Kurum', kod='SEXP')
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez Şube', kod='SEXP-A')
        self.egitim_yili = EgitimYili.objects.create(
            baslangic_yil=2025, bitis_yil=2026, aktif_mi=True,
        )
        self.user = User.objects.create_user(username='olcmeexport', password='test')
        self.client.force_authenticate(user=self.user)

        self.exam = Exam.objects.create(
            name='Deneme Sınavı Export Testi',
            exam_type='DENEME',
            kurum=self.kurum,
            sube=self.sube,
            egitim_yili=self.egitim_yili,
        )
        self.turkce = ExamSection.objects.create(
            exam=self.exam, name='Türkçe', order=1, question_start=1, question_end=40,
        )
        self.matematik = ExamSection.objects.create(
            exam=self.exam, name='Matematik', order=2, question_start=41, question_end=80,
        )

        self.session = ExamSession.objects.create(
            exam=self.exam, status=ExamSession.Status.COMPLETED, original_filename='test.dat',
        )

        self.ogrenciler = []
        for i in range(3):
            ogrenci = Ogrenci.objects.create(
                kurum=self.kurum, sube=self.sube, ad=f'Öğrenci{i}', soyad='Test',
            )
            self.ogrenciler.append(ogrenci)
            correct = 30 - i * 5
            wrong = 5 + i
            empty = 40 - correct - wrong
            net = Decimal(correct) - Decimal(wrong) / Decimal(4)
            answer = StudentAnswer.objects.create(
                session=self.session,
                student=ogrenci,
                raw_student_id=str(1000 + i),
                raw_student_name=f'Öğrenci{i} Test',
                total_correct=correct * 2,
                total_wrong=wrong * 2,
                total_empty=empty * 2,
                total_net=net * 2,
            )
            for section in (self.turkce, self.matematik):
                StudentSectionScore.objects.create(
                    student_answer=answer, section=section,
                    correct=correct, wrong=wrong, empty=empty, net=net,
                )

        self.url = f'/api/coaching/olcme-degerlendirme/exams/{self.exam.id}/analysis/rankings/'
        self.headers = {
            'HTTP_X_KURUM_ID': str(self.kurum.id),
            'HTTP_X_SUBE_ID': str(self.sube.id),
        }

    def test_rankings_json_default(self):
        res = self.client.get(self.url, **self.headers)
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertEqual(len(data['rankings']), 3)
        self.assertIn('total_correct', data['rankings'][0])

    def test_rankings_export_xlsx(self):
        from io import BytesIO
        from openpyxl import load_workbook

        res = self.client.get(self.url, {'format': 'xlsx'}, **self.headers)
        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            res['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('Content-Disposition', res)
        self.assertGreater(len(res.content), 0)

        wb = load_workbook(BytesIO(res.content))
        self.assertIn('Sıralama', wb.sheetnames)
        self.assertIn('İstatistik Grafikleri', wb.sheetnames)
        texts = []
        for row in wb['Sıralama'].iter_rows(max_row=25, max_col=20, values_only=True):
            texts.extend(str(v) for v in row if v is not None)
        self.assertTrue(any('Tahmini Sıralama Yılı' in t for t in texts))
        self.assertTrue(any(t == 'D' for t in texts))
        self.assertTrue(any(t == 'Net' for t in texts))
        self.assertTrue(any('T.Net' in t for t in texts))
        chart_texts = []
        for row in wb['İstatistik Grafikleri'].iter_rows(max_row=6, max_col=12, values_only=True):
            chart_texts.extend(str(v) for v in row if v is not None)
        self.assertTrue(any('Tahmini sıralama yılı' in t for t in chart_texts))
        self.assertTrue(any('İstatistik Grafikleri' in t for t in chart_texts))

    def test_rankings_export_csv(self):
        res = self.client.get(self.url, {'format': 'csv'}, **self.headers)
        self.assertEqual(res.status_code, 200)
        self.assertIn('text/csv', res['Content-Type'])
        body = res.content.decode('utf-8-sig')
        self.assertIn('Öğrenci0 Test', body)
        self.assertIn('Sıra', body)
        self.assertIn('Tahmini Sıralama Yılı', body)
        self.assertIn('İSTATİSTİK GRAFİKLERİ', body)
        self.assertIn('Türkçe (40) D', body)
        self.assertIn('Türkçe (40) Net', body)
        self.assertIn('T.Net', body)
        self.assertIn('Tah.Sıra', body)
        self.assertIn('Kurs Ortalaması', body)

    def test_rankings_export_csv_uses_requested_year(self):
        res = self.client.get(self.url, {'format': 'csv', 'ranking_year': 2024}, **self.headers)
        self.assertEqual(res.status_code, 200)
        body = res.content.decode('utf-8-sig')
        self.assertIn('Tahmini Sıralama Yılı: 2024', body)
        self.assertIn('Tah.Sıra (2024)', body)


class OlcmeRankingsExportBuilderTest(SimpleTestCase):
    def test_pdf_like_columns_include_dyb_and_year(self):
        from apps.coaching.application.olcme_rankings_export import (
            build_export_columns,
            build_export_rows,
        )

        sections = [
            {'id': 1, 'name': 'Türkçe', 'is_sub_section': False, 'parent_id': None, 'question_count': 40},
            {'id': 2, 'name': 'Matematik', 'is_sub_section': False, 'parent_id': None, 'question_count': 40},
        ]
        cols = build_export_columns(is_ayt=False, sections=sections, ranking_year=2025)
        labels = [c.label for c in cols]
        self.assertIn('Türkçe (40) D', labels)
        self.assertIn('Türkçe (40) Y', labels)
        self.assertIn('Türkçe (40) Net', labels)
        self.assertIn('T.Net', labels)
        self.assertIn('Puan', labels)
        self.assertIn('Tah.Sıra (2025)', labels)

        rows = build_export_rows(
            [{
                'kurum_ici_sira': 1,
                'student_name': 'Ada Test',
                'sinif': '12-A',
                'alan': 'SAYISAL',
                'toplam_net': 60,
                'puan': 400,
                'kurum_ici_yuzdelik': 90,
                'tahmini_siralama': 12000,
                'yuzdelik_dilim': 8,
                'section_nets': {
                    '1': {'correct': 30, 'wrong': 4, 'net': 29.0, 'empty': 6},
                    '2': {'correct': 28, 'wrong': 8, 'net': 26.0, 'empty': 4},
                },
            }],
            is_ayt=False,
            sections=sections,
            include_avg_row=False,
        )
        self.assertEqual(rows[0]['d_1'], 30)
        self.assertEqual(rows[0]['y_1'], 4)
        self.assertEqual(rows[0]['net_1'], 29.0)
