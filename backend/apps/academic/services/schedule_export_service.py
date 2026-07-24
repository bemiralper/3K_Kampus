"""Sınıf ders programı toplu dışa aktarma — grid matrisi + kurumsal Excel/CSV."""
from __future__ import annotations

import csv
import io
from typing import Any, Optional

from django.http import HttpResponse

from apps.academic.domain.program_grid_cell import CellStatus, ProgramGridCell
from apps.academic.domain.schedule_version import ScheduleVersion
from apps.academic.domain.timeslot import TimeSlot
from apps.academic.domain.weekly_day import WeeklyDay
from apps.sinif.domain.models import Sinif
from apps.term.domain.models import Term


class ScheduleExportError(Exception):
    def __init__(self, message: str, field: Optional[str] = None):
        self.message = message
        self.field = field
        super().__init__(message)


TEACHER_DISPLAY_MODES = ('full', 'initials', 'hidden')
COLOR_BY_MODES = ('ders', 'ogretmen', 'none')


def _u32(n: int) -> int:
    return n & 0xFFFFFFFF


def _i32(n: int) -> int:
    n = _u32(n)
    return n - 0x100000000 if n >= 0x80000000 else n


def _imul32(a: int, b: int) -> int:
    return _i32(_u32(a) * _u32(b))


def _schedule_hash_id(cid: int) -> int:
    """frontend/lib/schedule-color.ts hashId ile birebir."""
    x = _imul32(cid ^ 0x9E3779B9, 0x85EBCA6B)
    x = _i32(x ^ (_u32(x) >> 13))
    x = _imul32(x, 0xC2B2AE35)
    x = _i32(x ^ (_u32(x) >> 16))
    return abs(x)


def _hsl_to_rgb_hex(h: int, s: int, l: int) -> str:
    """HSL (0–360, 0–100, 0–100) → RRGGBB (openpyxl)."""
    hf = h / 360.0
    sf = s / 100.0
    lf = l / 100.0
    a = sf * min(lf, 1 - lf)

    def f(n: float) -> int:
        k = (n + hf * 12) % 12
        color = lf - a * max(min(k - 3, 9 - k, 1), -1)
        return max(0, min(255, round(255 * color)))

    return f'{f(0):02X}{f(8):02X}{f(4):02X}'


def schedule_cell_fill_hex(entity_id: int | None) -> str | None:
    """Ekrandaki pastel hücre rengi (bg) — ders/öğretmen id."""
    if not entity_id or entity_id <= 0:
        return None
    h = _schedule_hash_id(entity_id) % 360
    s = 48 + (_schedule_hash_id(entity_id + 7) % 18)
    l = 88 + (_schedule_hash_id(entity_id + 13) % 6)
    return _hsl_to_rgb_hex(h, s, l)


def format_teacher_name(full_name: str, mode: str = 'full') -> str:
    """Öğretmen adı: full | initials (A. Y.) | hidden."""
    name = (full_name or '').strip()
    if not name or mode == 'hidden':
        return ''
    if mode == 'initials':
        parts = [p for p in name.replace('.', ' ').split() if p]
        if not parts:
            return ''
        return ' '.join(f'{p[0].upper()}.' for p in parts)
    return name


def apply_teacher_display(payload: dict[str, Any], mode: str = 'full') -> dict[str, Any]:
    """Payload hücrelerindeki öğretmen metnini seçilen moda göre düzenle."""
    if mode not in TEACHER_DISPLAY_MODES:
        mode = 'full'
    for group in payload.get('groups') or []:
        for row in group.get('rows') or []:
            new_cells = []
            for cell in row.get('cells') or []:
                if not cell:
                    new_cells.append(None)
                    continue
                teacher = format_teacher_name(cell.get('teacher') or '', mode)
                lesson = cell.get('lesson') or ''
                label = lesson
                if teacher:
                    label = f'{lesson}\n{teacher}'.strip()
                new_cells.append({
                    **cell,
                    'teacher': teacher,
                    'label': label,
                })
            row['cells'] = new_cells
    payload['teacher_display'] = mode
    return payload


def _slots_for_version(version: ScheduleVersion, day_ids: list[int]) -> list:
    days_qs = WeeklyDay.objects.filter(id__in=day_ids)
    template_ids = set(
        days_qs.filter(schedule_template_id__isnull=False)
        .values_list('schedule_template_id', flat=True)
    )
    if version.schedule_template_id:
        template_ids.add(version.schedule_template_id)
    if not template_ids:
        return []
    return list(
        TimeSlot.objects.filter(
            schedule_template_id__in=template_ids,
            slot_type='LESSON',
            is_active=True,
        ).order_by('order', 'id')
    )


def build_classroom_schedule_payload(
    *,
    term_id: int,
    version_id: Optional[int],
    classroom_ids: list[int],
    sube_id: int,
) -> dict[str, Any]:
    try:
        term = Term.objects.select_related('kurum', 'sube', 'egitim_yili').get(
            pk=term_id, sube_id=sube_id,
        )
    except Term.DoesNotExist as exc:
        raise ScheduleExportError('Dönem bulunamadı.', 'term_id') from exc

    if not version_id:
        raise ScheduleExportError(
            'Program versiyonu seçilmedi. Ders Programı ekranında bir versiyon seçip tekrar deneyin.',
            'version_id',
        )

    try:
        version = ScheduleVersion.objects.select_related(
            'weekly_cycle', 'schedule_template',
        ).get(pk=version_id, term_id=term_id)
    except ScheduleVersion.DoesNotExist as exc:
        raise ScheduleExportError('Program versiyonu bulunamadı.', 'version_id') from exc

    days = list(
        WeeklyDay.objects.filter(
            weekly_cycle=version.weekly_cycle,
            is_active=True,
        ).order_by('order')
    )
    slots = _slots_for_version(version, [d.id for d in days])

    classrooms = list(
        Sinif.objects.filter(
            id__in=classroom_ids,
            sube_id=sube_id,
            aktif_mi=True,
        ).order_by('ad')
    )
    if not classrooms:
        raise ScheduleExportError('Dışa aktarılacak sınıf yok.', 'classroom_ids')

    days_meta = [
        {
            'id': d.id,
            'name': d.name,
            'short_name': getattr(d, 'day_name_short', None) or d.name[:3],
            'order': d.order,
        }
        for d in days
    ]
    slots_meta = [
        {
            'id': s.id,
            'name': s.name,
            'start': s.start_time.strftime('%H:%M') if s.start_time else '',
            'end': s.end_time.strftime('%H:%M') if s.end_time else '',
            'order': s.order,
        }
        for s in slots
    ]

    groups = []
    for sinif in classrooms:
        cells = ProgramGridCell.objects.filter(
            schedule_version=version,
            sinif_id=sinif.id,
            is_active=True,
            status=CellStatus.FILLED,
        ).select_related(
            'ders', 'ogretmen', 'weekly_day', 'timeslot',
            'class_lesson_plan', 'class_lesson_plan__ders',
        )

        from apps.egitim_tanimlari.display import resolve_ders_display_name

        cell_map: dict[str, dict[str, Any]] = {}
        for c in cells:
            key = f'{c.weekly_day_id}:{c.timeslot_id}'
            lesson_name = resolve_ders_display_name(
                ders=c.ders if c.ders_id else None,
                plan=getattr(c, 'class_lesson_plan', None),
            )
            teacher_name = (
                f'{c.ogretmen.ad} {c.ogretmen.soyad}'.strip()
                if c.ogretmen_id else ''
            )
            cell_map[key] = {
                'lesson': lesson_name,
                'lesson_id': c.ders_id,
                'teacher': teacher_name,
                'teacher_id': c.ogretmen_id,
                'label': (
                    f'{lesson_name}'
                    + (f'\n{teacher_name}' if teacher_name else '')
                ).strip(),
            }

        rows = []
        for slot in slots:
            row = {
                'slot_id': slot.id,
                'slot_name': slot.name,
                'slot_time': (
                    f"{slot.start_time.strftime('%H:%M') if slot.start_time else ''}"
                    f"–{slot.end_time.strftime('%H:%M') if slot.end_time else ''}"
                ),
                'cells': [],
            }
            for day in days:
                key = f'{day.id}:{slot.id}'
                row['cells'].append(cell_map.get(key))
            rows.append(row)

        groups.append({
            'classroom_id': sinif.id,
            'classroom_name': sinif.ad,
            'rows': rows,
            'filled_count': len(cell_map),
        })

    egitim_yili = ''
    if term.egitim_yili_id:
        egitim_yili = str(term.egitim_yili)

    return {
        'term': {'id': term.id, 'name': term.name},
        'version': {
            'id': version.id,
            'name': version.name,
            'is_locked': version.is_locked,
        },
        'kurum_ad': term.kurum.ad if term.kurum_id else '',
        'sube_ad': term.sube.ad if term.sube_id else '',
        'egitim_yili': egitim_yili,
        'days': days_meta,
        'slots': slots_meta,
        'groups': groups,
    }


def _report_meta(payload: dict[str, Any]):
    from shared.export.style_manager import ReportMeta

    return ReportMeta(
        report_title='DERS PROGRAMI',
        kurum_ad=payload.get('kurum_ad') or '',
        sube_ad=payload.get('sube_ad') or '',
        egitim_yili=payload.get('egitim_yili') or '',
        extra={
            'Dönem': payload['term']['name'],
            'Versiyon': payload['version']['name'],
        },
    )


def export_schedule_csv(payload: dict[str, Any], *, filename: str) -> HttpResponse:
    """Kurumsal letterhead + sınıf blokları (öğrenci listesi CSV kalıbı)."""
    from shared.export.csv_export_service import CsvExportService
    from shared.export import style_manager as sm

    days = payload['days']
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';', lineterminator='\n')
    CsvExportService.write_letterhead_rows(writer, _report_meta(payload))
    writer.writerow([f"Dönem: {payload['term']['name']}"])
    writer.writerow([f"Versiyon: {payload['version']['name']}"])
    writer.writerow([])

    for group in payload['groups']:
        writer.writerow([group['classroom_name']])
        writer.writerow(['Saat'] + [d['short_name'] or d['name'] for d in days])
        for row in group['rows']:
            cells = []
            for cell in row['cells']:
                if not cell:
                    cells.append('')
                else:
                    text = cell.get('lesson') or ''
                    if cell.get('teacher'):
                        text = f"{text} ({cell['teacher']})"
                    cells.append(text)
            writer.writerow([f"{row['slot_name']} {row.get('slot_time') or ''}".strip()] + cells)
        writer.writerow([])

    content = '\ufeff' + buf.getvalue()
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{sm.safe_filename(filename)}.csv"'
    return response


def export_schedule_xlsx(
    payload: dict[str, Any],
    *,
    filename: str,
    layout: str = 'stacked',
    color_by: str = 'ders',
) -> HttpResponse:
    """Kurumsal logo + letterhead ile haftalık grid Excel (öğrenci listesi altyapısı)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from shared.export import style_manager as sm
    from shared.export.excel_export_service import ExcelExportService
    from shared.export.style_manager import ExportStat

    if color_by not in COLOR_BY_MODES:
        color_by = 'ders'

    days = payload['days']
    num_cols = max(2, 1 + len(days))
    meta = _report_meta(payload)
    stats = [
        ExportStat(label='Sınıf sayısı', value=len(payload['groups']), type='integer'),
        ExportStat(label='Dönem', value=payload['term']['name'], type='text'),
        ExportStat(label='Versiyon', value=payload['version']['name'], type='text'),
    ]

    thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    header_fill = PatternFill('solid', fgColor='E8F1F8')
    header_font = Font(name=sm.FONT_NAME, size=11, bold=True, color=sm.BRAND_PRIMARY_HEX)
    cell_font = Font(name=sm.FONT_NAME, size=10, color='0F172A')
    title_font = Font(name=sm.FONT_NAME, size=13, bold=True, color=sm.BRAND_PRIMARY_HEX)
    slot_fill = PatternFill('solid', fgColor='F8FAFC')
    fill_cache: dict[int, PatternFill] = {}

    def fill_for_cell(cell_data: dict[str, Any] | None) -> PatternFill | None:
        if color_by == 'none' or not cell_data:
            return None
        eid = (
            cell_data.get('teacher_id')
            if color_by == 'ogretmen'
            else cell_data.get('lesson_id')
        )
        try:
            eid_int = int(eid) if eid is not None else 0
        except (TypeError, ValueError):
            return None
        hex_bg = schedule_cell_fill_hex(eid_int)
        if not hex_bg:
            return None
        if eid_int not in fill_cache:
            fill_cache[eid_int] = PatternFill('solid', fgColor=hex_bg)
        return fill_cache[eid_int]

    def write_class_grid(ws, group, start_row: int) -> int:
        title_cell = ws.cell(row=start_row, column=1, value=group['classroom_name'])
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        if num_cols > 1:
            ws.merge_cells(
                start_row=start_row, start_column=1,
                end_row=start_row, end_column=num_cols,
            )
        ws.row_dimensions[start_row].height = 22
        r = start_row + 1

        headers = ['Saat'] + [d['short_name'] or d['name'] for d in days]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 20
        r += 1

        for row in group['rows']:
            slot_label = row['slot_name']
            if row.get('slot_time'):
                slot_label = f"{row['slot_name']}\n{row['slot_time']}"
            c0 = ws.cell(row=r, column=1, value=slot_label)
            c0.font = cell_font
            c0.fill = slot_fill
            c0.border = thin
            c0.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            for col, cell_data in enumerate(row['cells'], 2):
                text = ''
                if cell_data:
                    text = cell_data['lesson']
                    if cell_data.get('teacher'):
                        text = f"{text}\n{cell_data['teacher']}"
                c = ws.cell(row=r, column=col, value=text)
                c.font = cell_font
                c.border = thin
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                fill = fill_for_cell(cell_data)
                if fill is not None:
                    c.fill = fill
            has_teacher = any(
                (cd or {}).get('teacher') for cd in row['cells'] if cd
            )
            ws.row_dimensions[r].height = 44 if has_teacher else 28
            r += 1

        for col in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20 if col > 1 else 15
        return r + 1

    wb = Workbook()

    if layout == 'per_class_sheet':
        first = True
        for group in payload['groups']:
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet()
            ws.title = sm.safe_sheet_title(group['classroom_name'])[:31]
            header_row = ExcelExportService._write_letterhead(ws, meta, num_cols=num_cols)
            current = ExcelExportService._write_stats(
                ws, stats, start_row=header_row, num_cols=num_cols,
            )
            end_row = write_class_grid(ws, group, current)
            ExcelExportService._apply_page_setup(
                ws,
                orientation='landscape',
                header_row=header_row,
                last_row=max(end_row, header_row + 1),
                last_col=num_cols,
                report_title=meta.report_title,
            )
    else:
        ws = wb.active
        ws.title = sm.safe_sheet_title('Ders Programları')
        header_row = ExcelExportService._write_letterhead(ws, meta, num_cols=num_cols)
        current = ExcelExportService._write_stats(
            ws, stats, start_row=header_row, num_cols=num_cols,
        )
        for group in payload['groups']:
            current = write_class_grid(ws, group, current)
        ExcelExportService._apply_page_setup(
            ws,
            orientation='landscape',
            header_row=header_row,
            last_row=max(current - 1, header_row),
            last_col=num_cols,
            report_title=meta.report_title,
        )

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{sm.safe_filename(filename)}.xlsx"'
    return response
