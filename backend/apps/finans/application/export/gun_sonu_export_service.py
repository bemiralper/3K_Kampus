"""
Gün Sonu Özet Raporu — PDF / CSV / Excel dışa aktarma.
"""
from __future__ import annotations

import csv
import io
from typing import Any

from django.http import HttpResponse

from apps.communication.application.html_to_pdf import render_html_to_pdf
from apps.finans.application.export.gun_sonu_report_html import build_gun_sonu_ozet_html


class GunSonuExportService:
    SUPPORTED_FORMATS = ('pdf', 'csv', 'xlsx')

    @classmethod
    def build(cls, export_format: str, report: dict, *, orientation: str = 'portrait') -> HttpResponse | dict:
        fmt = (export_format or 'pdf').lower()
        if fmt not in cls.SUPPORTED_FORMATS:
            raise ValueError(f'Desteklenmeyen format: {export_format}')

        from apps.finans.application.export.export_service import ExportService
        ori = ExportService._normalize_orientation(orientation)

        ozet = report.get('ozet_rapor') or {}
        meta = ozet.get('meta') or {}
        title = f"Gun_Sonu_Ozet_{meta.get('tarih_iso', 'rapor')}"

        if fmt == 'pdf':
            return cls._build_pdf(report, title, orientation=ori)
        if fmt == 'csv':
            return cls._build_csv(report, title)
        return cls._build_xlsx(report, title, orientation=ori)

    @classmethod
    def render_pdf_bytes(cls, report: dict, *, orientation: str = 'portrait') -> bytes:
        from apps.finans.application.export.export_service import ExportService
        ori = ExportService._normalize_orientation(orientation)
        html_doc = build_gun_sonu_ozet_html(report, orientation=ori)
        try:
            return render_html_to_pdf(html_doc, landscape=(ori == 'landscape'))
        except RuntimeError:
            from apps.communication.application.pdf_render_service import PdfRenderService

            ozet = report.get('ozet_rapor') or {}
            meta = ozet.get('meta') or {}
            lines = [
                meta.get('baslik', 'Gün Sonu Raporu'),
                f"Tarih: {meta.get('tarih', '')}",
                f"Şube: {meta.get('sube', '')}",
                '',
            ]
            for k, v in (ozet.get('gunluk_ozet') or {}).items():
                lines.append(f'{k}: {v}')
            return PdfRenderService.render_simple_text_pdf(
                meta.get('baslik', 'Gün Sonu Raporu'),
                '\n'.join(lines),
            )

    @classmethod
    def _build_pdf(cls, report: dict, title: str, *, orientation: str = 'portrait') -> HttpResponse:
        pdf_bytes = cls.render_pdf_bytes(report, orientation=orientation)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.pdf"'
        return response

    @classmethod
    def _build_csv(cls, report: dict, title: str) -> HttpResponse:
        ozet = report.get('ozet_rapor') or {}
        meta = ozet.get('meta') or {}
        gunluk = ozet.get('gunluk_ozet') or {}

        buf = io.StringIO()
        buf.write('\ufeff')
        writer = csv.writer(buf, delimiter=';')

        writer.writerow([meta.get('kurum_ad') or meta.get('marka') or ''])
        writer.writerow([meta.get('baslik', '')])
        writer.writerow([f"Tarih: {meta.get('tarih', '')}"])
        writer.writerow([f"Şube: {meta.get('sube', '')}"])
        writer.writerow([f"Hazırlayan: {meta.get('hazirlayan', '')}"])
        writer.writerow([f"Oluşturulma: {meta.get('olusturulma', '')}"])
        writer.writerow([])

        writer.writerow(['A. Günlük Özet'])
        writer.writerow(['Bilgi', 'Tutar'])
        for label, key in [
            ('Toplam Tahsilat', 'toplam_tahsilat'),
            ('Toplam İade', 'toplam_iade'),
            ('Toplam Gelir', 'toplam_gelir'),
            ('Toplam Gider', 'toplam_gider'),
            ('Net Nakit Girişi', 'net_nakit_girisi'),
        ]:
            writer.writerow([label, gunluk.get(key, 0)])
        writer.writerow([])

        writer.writerow(['B. Tahsilat Dağılımı'])
        writer.writerow(['Ödeme Türü', 'Tutar'])
        for row in ozet.get('tahsilat_dagilimi') or []:
            writer.writerow([row.get('label'), row.get('tutar', 0)])
        writer.writerow([])

        writer.writerow(['C. İşlem Sayıları'])
        writer.writerow(['İşlem', 'Adet'])
        islem = ozet.get('islem_sayilari') or {}
        for label, key in [
            ('Tahsilat', 'tahsilat'),
            ('Gelir Kaydı', 'gelir_kaydi'),
            ('Gider Kaydı', 'gider_kaydi'),
            ('İade', 'iade'),
            ('İptal', 'iptal'),
        ]:
            writer.writerow([label, islem.get(key, 0)])
        writer.writerow([])

        writer.writerow(['Kullanıcı Bazlı İşlem Özeti'])
        writer.writerow(['Personel', 'Tahsilat', 'Gelir', 'Gider'])
        for row in ozet.get('kullanici_ozeti') or []:
            writer.writerow([
                row.get('personel'),
                row.get('tahsilat', 0),
                row.get('gelir', 0),
                row.get('gider', 0),
            ])
        writer.writerow([])

        writer.writerow(['G. Notlar'])
        writer.writerow([ozet.get('notlar') or ''])

        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.csv"'
        return response

    @classmethod
    def _build_xlsx(cls, report: dict, title: str, *, orientation: str = 'portrait') -> HttpResponse:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError as exc:
            raise RuntimeError('openpyxl yüklü değil.') from exc

        ozet = report.get('ozet_rapor') or {}
        meta = ozet.get('meta') or {}
        gunluk = ozet.get('gunluk_ozet') or {}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Özet'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1F3C88')
        title_font = Font(bold=True, size=14, color='1F3C88')
        thin = Side(style='thin', color='E2E8F0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row = 1
        ws.cell(row=row, column=1, value=meta.get('kurum_ad') or meta.get('marka') or '').font = title_font
        row += 1
        ws.cell(row=row, column=1, value=meta.get('baslik', '')).font = Font(bold=True, size=12)
        row += 1
        for label in ('tarih', 'sube', 'hazirlayan', 'olusturulma'):
            ws.cell(row=row, column=1, value=f"{label.title()}: {meta.get(label, '')}")
            row += 1
        row += 1

        def write_section(sheet_row: int, section_title: str, headers: list[str], data_rows: list[list[Any]]):
            sheet_row += 1
            ws.cell(row=sheet_row, column=1, value=section_title).font = Font(bold=True, color='1F3C88')
            sheet_row += 1
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=sheet_row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            sheet_row += 1
            for data in data_rows:
                for col, val in enumerate(data, start=1):
                    cell = ws.cell(row=sheet_row, column=col, value=val)
                    cell.border = border
                    if col > 1:
                        cell.alignment = Alignment(horizontal='right')
                sheet_row += 1
            return sheet_row

        row = write_section(
            row, 'A. Günlük Özet', ['Bilgi', 'Tutar (TL)'],
            [
                ['Toplam Tahsilat', gunluk.get('toplam_tahsilat', 0)],
                ['Toplam İade', gunluk.get('toplam_iade', 0)],
                ['Toplam Gelir', gunluk.get('toplam_gelir', 0)],
                ['Toplam Gider', gunluk.get('toplam_gider', 0)],
                ['Net Nakit Girişi', gunluk.get('net_nakit_girisi', 0)],
            ],
        )

        dagilim_rows = [[r.get('label'), r.get('tutar', 0)] for r in ozet.get('tahsilat_dagilimi') or []]
        row = write_section(row, 'B. Tahsilat Dağılımı', ['Ödeme Türü', 'Tutar (TL)'], dagilim_rows)

        islem = ozet.get('islem_sayilari') or {}
        row = write_section(
            row, 'C. İşlem Sayıları', ['İşlem', 'Adet'],
            [
                ['Tahsilat', islem.get('tahsilat', 0)],
                ['Gelir Kaydı', islem.get('gelir_kaydi', 0)],
                ['Gider Kaydı', islem.get('gider_kaydi', 0)],
                ['İade', islem.get('iade', 0)],
                ['İptal', islem.get('iptal', 0)],
            ],
        )

        kullanici_rows = [
            [k.get('personel'), k.get('tahsilat', 0), k.get('gelir', 0), k.get('gider', 0)]
            for k in ozet.get('kullanici_ozeti') or []
        ]
        row = write_section(
            row, 'Kullanıcı Bazlı İşlem Özeti', ['Personel', 'Tahsilat', 'Gelir', 'Gider'],
            kullanici_rows,
        )

        row += 1
        ws.cell(row=row, column=1, value='G. Notlar').font = Font(bold=True, color='1F3C88')
        row += 1
        ws.cell(row=row, column=1, value=ozet.get('notlar') or '')

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14

        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape' if orientation == 'landscape' else 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        buf = io.BytesIO()
        wb.save(buf)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{cls._safe_filename(title)}.xlsx"'
        return response

    @staticmethod
    def _safe_filename(name: str) -> str:
        safe = ''.join(c if c.isalnum() or c in '-_' else '_' for c in (name or 'rapor'))
        return safe[:80] or 'rapor'
