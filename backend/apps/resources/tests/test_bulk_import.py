"""Toplu kitap Excel şablonu ve tür eşleştirme."""
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.egitim_tanimlari.models import Ders, SinifSeviyesi
from apps.kurum.domain.models import Kurum
from apps.resources.application.bulk_import import (
    BulkBookImportService,
    build_excel_template,
)
from apps.resources.models import BookType, ResourceBook
from apps.sube.domain.models import Sube

User = get_user_model()


class BulkBookImportTemplateTest(TestCase):
    def setUp(self):
        self.kurum = Kurum.objects.create(ad='Bulk Kurum', kod='BLK', aktif_mi=True)
        self.sube = Sube.objects.create(kurum=self.kurum, ad='Merkez', kod='MRK', aktif_mi=True)
        self.book_type = BookType.objects.create(kod='ABCC', ad='ABC Tür')
        self.ders = Ders.objects.create(
            ad='Matematik', kod='MAT', kurum=self.kurum, sube=self.sube, aktif_mi=True,
        )
        self.sinif = SinifSeviyesi.objects.create(
            ad='11. Sınıf', kod='S11', sira=11, kurum=self.kurum, sube=self.sube, aktif_mi=True,
        )

    def test_template_has_dropdowns_and_lists(self):
        data = build_excel_template(kurum_id=self.kurum.id, sube_id=self.sube.id)
        wb = load_workbook(BytesIO(data))
        self.assertIn('Kitaplar', wb.sheetnames)
        self.assertIn('Listeler', wb.sheetnames)
        self.assertIn('Talimat', wb.sheetnames)
        lists = wb['Listeler']
        labels = [lists.cell(r, 1).value for r in range(2, 10) if lists.cell(r, 1).value]
        self.assertTrue(any('ABCC' in (x or '') for x in labels))
        kitaplar = wb['Kitaplar']
        self.assertGreaterEqual(len(kitaplar.data_validations.dataValidation), 3)

    def test_resolve_book_type_from_label_and_code(self):
        svc = BulkBookImportService(kurum_id=self.kurum.id, sube_id=self.sube.id)
        self.assertEqual(svc._resolve_book_type('ABCC — ABC Tür').id, self.book_type.id)
        self.assertEqual(svc._resolve_book_type('abcc').id, self.book_type.id)
        self.assertEqual(svc._resolve_book_type('ABC Tür').id, self.book_type.id)

    def test_import_with_dropdown_label(self):
        svc = BulkBookImportService(kurum_id=self.kurum.id, sube_id=self.sube.id)
        result = svc.import_rows([{
            'ad': 'Test Kitap',
            'kod': 'TK1',
            'book_type': 'ABCC — ABC Tür',
            'ders': 'MAT — Matematik',
            'sinif': 'S11 — 11. Sınıf',
            'zorluk_min': '0',
            'zorluk_max': '4',
        }])
        self.assertEqual(result.eklenen, 1, result.hatalar)
        book = ResourceBook.objects.get(kod='TK1')
        self.assertEqual(book.book_type_id, self.book_type.id)
        self.assertEqual(book.zorluk_min, 0)

    def test_edited_row_not_skipped_when_aciklama_mentions_ornek(self):
        """Eski şablonda Açıklama'da 'örnek satır' kalsa bile gerçek ad işlenmeli."""
        svc = BulkBookImportService(kurum_id=self.kurum.id, sube_id=self.sube.id)
        result = svc.import_rows([{
            'ad': 'ABC Matematik',
            'kod': 'ABCC1',
            'book_type': 'ABCC',
            'ders': 'MAT',
            'sinif': 'S11',
            'aciklama': 'Örnek satır — yüklemeden önce silin',
        }])
        self.assertEqual(result.eklenen, 1, result.hatalar)

    def test_empty_import_returns_helpful_error(self):
        svc = BulkBookImportService(kurum_id=self.kurum.id, sube_id=self.sube.id)
        result = svc.import_rows([])
        self.assertEqual(result.eklenen, 0)
        self.assertEqual(result.hatali, 1)
        self.assertTrue(any('İşlenecek satır' in h['neden'] for h in result.hatalar))

    def test_import_template_api_requires_sube(self):
        user = User.objects.create_user(
            username='bulkadmin', email='b@test.com', password='x', is_staff=True,
        )
        client = APIClient()
        client.force_authenticate(user=user)
        res = client.get('/api/resources/books/import-template/')
        self.assertEqual(res.status_code, 400)
        res_ok = client.get(
            '/api/resources/books/import-template/',
            HTTP_X_KURUM_ID=str(self.kurum.id),
            HTTP_X_SUBE_ID=str(self.sube.id),
        )
        self.assertEqual(res_ok.status_code, 200)
        self.assertIn(
            'spreadsheetml',
            res_ok.get('Content-Type', ''),
        )
