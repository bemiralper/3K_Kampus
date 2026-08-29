"""
Okulizyon kazanım kataloğu — Excel → Subject / Topic / Outcome / SubOutcome.

Hiyerarşi:
  Ünite (2 parça)     → Topic
  Konu (3 parça)      → Outcome      (YKS deneme kodu: 21.1.3)
  Kazanım (4 parça)   → SubOutcome   (resmi MEB: 9.9.1.2)
  Alt kazanım (5)     → SubOutcome
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from django.db import transaction

from apps.coaching.olcme_degerlendirme.models.curriculum import (
    Outcome,
    SubOutcome,
    Subject,
    Topic,
)
from apps.coaching.olcme_degerlendirme.models.exam import ExamSection

XLSX_NAME = 'Okulizyon-Kazanimlar-20230812.xlsx'

DERS_SUBJECTS: dict[str, dict] = {
    'Türkçe': {
        'code': 'TURKCE', 'name': 'Türkçe', 'display_name': 'Türkçe', 'order': 1,
    },
    'Türk Dili ve Edebiyatı': {
        'code': 'TDE', 'name': 'Türk Dili ve Edebiyatı',
        'display_name': 'Türk Dili ve Edebiyatı', 'order': 2,
    },
    'Matematik': {
        'code': 'MATEMATIK', 'name': 'Matematik', 'display_name': 'Matematik', 'order': 3,
    },
    'Geometri': {
        'code': 'GEOMETRI', 'name': 'Geometri', 'display_name': 'Geometri', 'order': 4,
    },
    'Fen': {
        'code': 'FEN', 'name': 'Fen Bilimleri', 'display_name': 'Fen Bilimleri', 'order': 5,
    },
    'Fizik': {
        'code': 'FIZIK', 'name': 'Fizik', 'display_name': 'Fizik', 'order': 6,
    },
    'Kimya': {
        'code': 'KIMYA', 'name': 'Kimya', 'display_name': 'Kimya', 'order': 7,
    },
    'Biyoloji': {
        'code': 'BIYOLOJI', 'name': 'Biyoloji', 'display_name': 'Biyoloji', 'order': 8,
    },
    'Tarih': {
        'code': 'TARIH', 'name': 'Tarih', 'display_name': 'Tarih', 'order': 9,
    },
    'Coğrafya': {
        'code': 'COGRAFYA', 'name': 'Coğrafya', 'display_name': 'Coğrafya', 'order': 10,
    },
    'Felsefe': {
        'code': 'FELSEFE', 'name': 'Felsefe', 'display_name': 'Felsefe', 'order': 11,
    },
    'Din Kültürü ve Ahlak Bilgisi': {
        'code': 'DKAB', 'name': 'Din Kültürü ve Ahlak Bilgisi',
        'display_name': 'Din Kültürü ve Ahlak Bilgisi', 'order': 12,
    },
    'T.C.İnkılap Tarihi ve Atatürkçülük': {
        'code': 'INKILAP', 'name': 'T.C. İnkılap Tarihi ve Atatürkçülük',
        'display_name': 'T.C. İnkılap Tarihi ve Atatürkçülük', 'order': 13,
    },
    'İngilizce': {
        'code': 'INGILIZCE', 'name': 'İngilizce', 'display_name': 'İngilizce', 'order': 14,
    },
    'Sosyal': {
        'code': 'SOSYAL', 'name': 'Sosyal Bilgiler',
        'display_name': 'Sosyal Bilgiler', 'order': 15,
    },
    'Hayat Bilgisi': {
        'code': 'HAYAT_BILGISI', 'name': 'Hayat Bilgisi',
        'display_name': 'Hayat Bilgisi', 'order': 16,
    },
}

# Eski sınav-türü kodları → tek kanonik ders
SUBJECT_ALIASES: dict[str, str] = {
    'TURKCE_TYT': 'TURKCE',
    'TURKCE_LGS': 'TURKCE',
    'TDE_AYT': 'TDE',
    'MAT_TYT': 'MATEMATIK',
    'MAT_AYT': 'MATEMATIK',
    'MAT_LGS': 'MATEMATIK',
    'GEO_TYT': 'GEOMETRI',
    'GEO_AYT': 'GEOMETRI',
    'FEN_LGS': 'FEN',
    'FIZ_TYT': 'FIZIK',
    'FIZ_AYT': 'FIZIK',
    'KIM_TYT': 'KIMYA',
    'KIM_AYT': 'KIMYA',
    'BIO_TYT': 'BIYOLOJI',
    'BIO_AYT': 'BIYOLOJI',
    'TARIH_TYT': 'TARIH',
    'TARIH1_AYT': 'TARIH',
    'TARIH2_AYT': 'TARIH',
    'COGRAFYA_TYT': 'COGRAFYA',
    'COG1_AYT': 'COGRAFYA',
    'COG2_AYT': 'COGRAFYA',
    'FELSEFE_TYT': 'FELSEFE',
    'FELSEFE_AYT': 'FELSEFE',
    'DINKUL_TYT': 'DKAB',
    'DKAB_AYT': 'DKAB',
    'DINKUL_LGS': 'DKAB',
    'INKILAP_LGS': 'INKILAP',
    'YABDIL_LGS': 'INGILIZCE',
}

# Sınav bölüm adı → (subject_code, görünen ad, exam_type_filter)
SECTION_SUBJECT_MAP: dict[str, dict[str, tuple[str, str, str]]] = {
    'YKS_TYT': {
        'Türkçe':      ('TURKCE', 'Türkçe', 'ALL'),
        'Tarih':       ('TARIH', 'Tarih', 'ALL'),
        'Coğrafya':    ('COGRAFYA', 'Coğrafya', 'ALL'),
        'Felsefe':     ('FELSEFE', 'Felsefe', 'ALL'),
        'Din Kültürü': ('DKAB', 'Din Kültürü ve Ahlak Bilgisi', 'ALL'),
        'Matematik':   ('MATEMATIK', 'Matematik', 'ALL'),
        'Geometri':    ('GEOMETRI', 'Geometri', 'ALL'),
        'Fizik':       ('FIZIK', 'Fizik', 'ALL'),
        'Kimya':       ('KIMYA', 'Kimya', 'ALL'),
        'Biyoloji':    ('BIYOLOJI', 'Biyoloji', 'ALL'),
    },
    'YKS_AYT': {
        'Türk Dili ve Edebiyatı':      ('TDE', 'Türk Dili ve Edebiyatı', 'ALL'),
        'Tarih-1':                     ('TARIH', 'Tarih', 'ALL'),
        'Coğrafya-1':                  ('COGRAFYA', 'Coğrafya', 'ALL'),
        'Tarih-2':                     ('TARIH', 'Tarih', 'ALL'),
        'Coğrafya-2':                  ('COGRAFYA', 'Coğrafya', 'ALL'),
        'Felsefe Grubu':               ('FELSEFE', 'Felsefe', 'ALL'),
        'Din Kültürü ve Ahlak Bilgisi': ('DKAB', 'Din Kültürü ve Ahlak Bilgisi', 'ALL'),
        'Matematik':                   ('MATEMATIK', 'Matematik', 'ALL'),
        'Geometri':                    ('GEOMETRI', 'Geometri', 'ALL'),
        'Fizik':                       ('FIZIK', 'Fizik', 'ALL'),
        'Kimya':                       ('KIMYA', 'Kimya', 'ALL'),
        'Biyoloji':                    ('BIYOLOJI', 'Biyoloji', 'ALL'),
    },
    'LGS': {
        'Türkçe':         ('TURKCE', 'Türkçe', 'ALL'),
        'İnkılap Tarihi': ('INKILAP', 'T.C. İnkılap Tarihi ve Atatürkçülük', 'ALL'),
        'Din Kültürü':    ('DKAB', 'Din Kültürü ve Ahlak Bilgisi', 'ALL'),
        'Yabancı Dil':    ('INGILIZCE', 'İngilizce', 'ALL'),
        'Matematik':      ('MATEMATIK', 'Matematik', 'ALL'),
        'Fen Bilimleri':  ('FEN', 'Fen Bilimleri', 'ALL'),
    },
    'DENEME': {
        'Türkçe':      ('TURKCE', 'Türkçe', 'ALL'),
        'Tarih':       ('TARIH', 'Tarih', 'ALL'),
        'Coğrafya':    ('COGRAFYA', 'Coğrafya', 'ALL'),
        'Felsefe':     ('FELSEFE', 'Felsefe', 'ALL'),
        'Din Kültürü': ('DKAB', 'Din Kültürü ve Ahlak Bilgisi', 'ALL'),
        'Matematik':   ('MATEMATIK', 'Matematik', 'ALL'),
        'Geometri':    ('GEOMETRI', 'Geometri', 'ALL'),
        'Fizik':       ('FIZIK', 'Fizik', 'ALL'),
        'Kimya':       ('KIMYA', 'Kimya', 'ALL'),
        'Biyoloji':    ('BIYOLOJI', 'Biyoloji', 'ALL'),
    },
}


def default_xlsx_path() -> Path | None:
    here = Path(__file__).resolve()
    packaged = here.parents[1] / 'data' / XLSX_NAME
    repo_root = here.parents[5] / XLSX_NAME
    for candidate in (packaged, repo_root, Path.cwd() / XLSX_NAME):
        if candidate.is_file():
            return candidate
    return None


def _nz(value) -> bool:
    return value not in (None, '', 0, '0')


def classify_level(unite, konu, kazanim, alt) -> str:
    if not _nz(konu) and not _nz(kazanim) and not _nz(alt):
        return 'unite'
    if not _nz(kazanim) and not _nz(alt):
        return 'konu'
    if not _nz(alt):
        return 'kazanim'
    return 'alt'


def sinif_label(sinif: str) -> str:
    s = str(sinif or '').strip()
    if s.startswith('SHG (22)'):
        return 'SHG22'
    if s.startswith('SHG (21)'):
        return 'SHG21'
    if s.startswith('SHG'):
        return 'SHG'
    if s.isdigit():
        return f'{s}. sınıf'
    return s or '?'


def parent_code(code: str) -> str:
    parts = [p for p in str(code).split('.') if p]
    if len(parts) <= 1:
        return code
    return '.'.join(parts[:-1])


def parse_excel_rows(path: str | Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header or str(header[0] or '').strip() != 'Ders':
            raise ValueError('Excel başlığı beklenen formatta değil (Ders, Kazanım Kodu, …).')
        parsed = []
        for raw in rows_iter:
            if not raw or all(c is None or str(c).strip() == '' for c in raw):
                continue
            ders = str(raw[0] or '').strip()
            kod = str(raw[1] or '').strip()
            if not ders or not kod:
                continue
            parsed.append({
                'ders': ders,
                'kod': kod,
                'sinif': str(raw[2] or '').strip(),
                'unite': raw[3],
                'konu': raw[4],
                'kazanim': raw[5],
                'alt': raw[6],
                'metin': str(raw[7] or '').strip(),
            })
        return parsed
    finally:
        wb.close()


def rows_to_topics(ders_rows: list[dict]) -> list[dict]:
    """Bir dersin satırlarını Topic → Outcome → SubOutcome ağacına çevir."""
    topics: list[dict] = []
    current_topic: dict | None = None
    current_outcome: dict | None = None

    def ensure_topic(row: dict, code: str, name: str) -> dict:
        nonlocal current_topic, current_outcome
        current_topic = {
            'code': code,
            'name': name[:200],
            'outcomes': [],
        }
        topics.append(current_topic)
        current_outcome = None
        return current_topic

    for row in ders_rows:
        level = classify_level(row['unite'], row['konu'], row['kazanim'], row['alt'])
        code = row['kod']
        text = row['metin'] or code
        prefix = sinif_label(row['sinif'])

        if level == 'unite':
            ensure_topic(row, code, f'{prefix} · {text}')
            continue

        if current_topic is None:
            ensure_topic(row, parent_code(code), f'{prefix} · {parent_code(code)}')

        if level == 'konu':
            current_outcome = {
                'code': code,
                'text': text,
                'sub_outcomes': [],
            }
            current_topic['outcomes'].append(current_outcome)
            continue

        if current_outcome is None:
            current_outcome = {
                'code': parent_code(code) if level == 'alt' else code,
                'text': text if level == 'kazanim' else parent_code(code),
                'sub_outcomes': [],
            }
            current_topic['outcomes'].append(current_outcome)
            if level == 'kazanim':
                continue

        current_outcome['sub_outcomes'].append({
            'code': code,
            'text': text,
        })

    return topics


def persist_catalog(rows: list[dict], *, replace: bool = True) -> dict:
    """Satırları Subject ağaçlarına yazar. replace=True ise mevcut konu ağacı silinir."""
    by_ders: dict[str, list[dict]] = defaultdict(list)
    unknown = set()
    for row in rows:
        if row['ders'] not in DERS_SUBJECTS:
            unknown.add(row['ders'])
            continue
        by_ders[row['ders']].append(row)

    stats = {
        'subjects': 0,
        'topics': 0,
        'outcomes': 0,
        'sub_outcomes': 0,
        'unknown_ders': sorted(unknown),
        'per_subject': {},
    }

    with transaction.atomic():
        for ders, ders_rows in by_ders.items():
            spec = DERS_SUBJECTS[ders]
            subject, _created = Subject.objects.update_or_create(
                code=spec['code'],
                defaults={
                    'name': spec['name'],
                    'display_name': spec['display_name'],
                    'exam_type_filter': Subject.ExamTypeFilter.ALL,
                    'order': spec['order'],
                },
            )
            if replace:
                subject.topics.all().delete()

            topics_data = rows_to_topics(ders_rows)
            t_count = o_count = s_count = 0
            for t_idx, t_data in enumerate(topics_data):
                topic = Topic.objects.create(
                    subject=subject,
                    code=t_data['code'],
                    name=t_data['name'],
                    order=t_idx,
                )
                t_count += 1
                for o_idx, o_data in enumerate(t_data['outcomes']):
                    outcome = Outcome.objects.create(
                        topic=topic,
                        code=o_data['code'],
                        text=o_data['text'],
                        order=o_idx,
                    )
                    o_count += 1
                    for s_idx, s_data in enumerate(o_data['sub_outcomes']):
                        SubOutcome.objects.create(
                            outcome=outcome,
                            code=s_data['code'],
                            text=s_data['text'],
                            order=s_idx,
                        )
                        s_count += 1

            stats['subjects'] += 1
            stats['topics'] += t_count
            stats['outcomes'] += o_count
            stats['sub_outcomes'] += s_count
            stats['per_subject'][spec['code']] = {
                'topics': t_count,
                'outcomes': o_count,
                'sub_outcomes': s_count,
                'rows': len(ders_rows),
            }

        relinked = relink_alias_subjects()
        stats['relinked_sections'] = relinked

    return stats


def relink_alias_subjects() -> int:
    """Eski MAT_TYT vb. derslere bağlı sınav bölümlerini kanonik derse taşır."""
    moved = 0
    for alias, canonical in SUBJECT_ALIASES.items():
        old = Subject.objects.filter(code=alias).first()
        new = Subject.objects.filter(code=canonical).first()
        if not old or not new or old.id == new.id:
            continue
        moved += ExamSection.objects.filter(subject=old).update(subject=new)
        old.delete()
    return moved
