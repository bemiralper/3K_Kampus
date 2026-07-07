"""Toplu okul içe aktarma — şube kapsamlı, bulk insert."""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import BinaryIO

from apps.okul.application.service import normalize_okul_ad
from apps.okul.models import Okul

MAX_AD_LENGTH = 200
BULK_BATCH_SIZE = 500

HEADER_ALIASES = {
    'ad': ('okul adı', 'okul adi', 'okul', 'ad', 'school', 'schoolname', 'school name'),
    'okul_turu': ('okul türü', 'okul turu', 'tür', 'tur', 'okul_turu', 'school type'),
    'il': ('il', 'city', 'şehir', 'sehir'),
    'ilce': ('ilçe', 'ilce', 'district'),
}


@dataclass
class BulkImportResult:
    toplam_satir: int = 0
    eklenen: int = 0
    guncellenen: int = 0
    atlanan: int = 0
    hatali: int = 0
    hatalar: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            'toplam_satir': self.toplam_satir,
            'eklenen': self.eklenen,
            'guncellenen': self.guncellenen,
            'atlanan': self.atlanan,
            'hatali': self.hatali,
            'hatalar': self.hatalar,
        }


def _ad_key(ad: str) -> str:
    return _normalize_header(normalize_okul_ad(ad))


def _load_existing_keys(sube_id: int) -> set[str]:
    return {_normalize_header(ad) for ad in Okul.objects.filter(sube_id=sube_id).values_list('ad', flat=True)}


def _parse_row_dict(raw: dict, satir_no: int) -> tuple[dict | None, str | None]:
    ad = normalize_okul_ad(str(raw.get('ad') or ''))
    if not ad:
        if any(str(raw.get(k) or '').strip() for k in ('okul_turu', 'il', 'ilce')):
            return None, 'Okul adı boş'
        return None, None  # tamamen boş satır

    if len(ad) > MAX_AD_LENGTH:
        return None, 'Geçersiz veri (okul adı çok uzun)'

    return {
        'satir_no': satir_no,
        'ad': ad,
        'okul_turu': str(raw.get('okul_turu') or '').strip()[:100],
        'il': str(raw.get('il') or '').strip()[:100],
        'ilce': str(raw.get('ilce') or '').strip()[:100],
    }, None


class BulkOkulImportService:
    def import_rows(
        self,
        rows: list[dict],
        *,
        kurum_id: int,
        sube_id: int,
        duplicate_msg: str = 'Aynı dosyada tekrar ediyor',
    ) -> BulkImportResult:
        result = BulkImportResult()
        existing_keys = _load_existing_keys(sube_id)
        seen_in_batch: set[str] = set()
        to_create: list[Okul] = []

        for idx, raw in enumerate(rows, start=1):
            parsed, err = _parse_row_dict(raw, idx)
            if err:
                result.toplam_satir += 1
                result.hatali += 1
                result.hatalar.append({
                    'satir': idx,
                    'ad': str(raw.get('ad') or '').strip(),
                    'neden': err,
                })
                continue
            if parsed is None:
                continue

            result.toplam_satir += 1
            key = _ad_key(parsed['ad'])

            if key in existing_keys:
                result.atlanan += 1
                seen_in_batch.add(key)
                continue

            if key in seen_in_batch:
                result.hatali += 1
                result.hatalar.append({
                    'satir': parsed['satir_no'],
                    'ad': parsed['ad'],
                    'neden': duplicate_msg,
                })
                continue

            seen_in_batch.add(key)

            to_create.append(Okul(
                kurum_id=kurum_id,
                sube_id=sube_id,
                ad=parsed['ad'],
                okul_turu=parsed['okul_turu'],
                il=parsed['il'],
                ilce=parsed['ilce'],
                aktif_mi=True,
            ))

        if to_create:
            Okul.objects.bulk_create(to_create, batch_size=BULK_BATCH_SIZE)
            result.eklenen = len(to_create)

        return result

    def import_ad_list(self, adlar: list[str], *, kurum_id: int, sube_id: int) -> BulkImportResult:
        rows = [{'ad': ad} for ad in adlar]
        return self.import_rows(
            rows,
            kurum_id=kurum_id,
            sube_id=sube_id,
            duplicate_msg='Aynı listede tekrar ediyor',
        )

    def parse_excel(self, file_obj: BinaryIO) -> list[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if not first_row:
            return []

        headers = [_cell_str(c) for c in first_row]
        col_map = _map_columns(headers)
        start_iter = rows_iter

        if col_map.get('ad') is None:
            # Başlık satırı yok — ilk satır veri, sütun sırası: ad, tür, il, ilçe
            col_map = {'ad': 0, 'okul_turu': 1, 'il': 2, 'ilce': 3}
            start_iter = _chain_first_row(first_row, rows_iter)

        parsed_rows: list[dict] = []
        for row in start_iter:
            if not row or all(_cell_str(c) == '' for c in row):
                continue
            parsed_rows.append({
                'ad': _cell_at(row, col_map.get('ad')),
                'okul_turu': _cell_at(row, col_map.get('okul_turu')),
                'il': _cell_at(row, col_map.get('il')),
                'ilce': _cell_at(row, col_map.get('ilce')),
            })
        wb.close()
        return parsed_rows

    def import_excel(self, file_obj: BinaryIO, *, kurum_id: int, sube_id: int) -> BulkImportResult:
        rows = self.parse_excel(file_obj)
        return self.import_rows(
            rows,
            kurum_id=kurum_id,
            sube_id=sube_id,
            duplicate_msg='Aynı Excel dosyasında tekrar ediyor',
        )


def build_excel_template() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = 'Okullar'
    ws.append(['Okul Adı', 'Okul Türü', 'İl', 'İlçe'])
    ws.append(['Atatürk Anadolu Lisesi', 'Anadolu Lisesi', 'Ankara', 'Çankaya'])
    ws.append(['Ankara Fen Lisesi', 'Fen Lisesi', 'Ankara', 'Altındağ'])
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _cell_at(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ''
    return _cell_str(row[index])


def _normalize_header(value: str) -> str:
    return (value or '').strip().casefold()


def _map_columns(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    normalized = [_normalize_header(h) for h in headers]
    for field_name, aliases in HEADER_ALIASES.items():
        alias_fold = {_normalize_header(a) for a in aliases}
        for idx, header in enumerate(normalized):
            if header in alias_fold or any(alias in header for alias in alias_fold):
                col_map[field_name] = idx
                break
    return col_map


def _chain_first_row(first_row, rows_iter):
    yield first_row
    yield from rows_iter
